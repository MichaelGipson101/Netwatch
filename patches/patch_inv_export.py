#!/usr/bin/env python3
"""
netwatch patch: inventory XLSX export.

Adds a "Export XLSX" button to the Inventory tab toolbar. Clicking it
downloads a spreadsheet with all current inventory records, using the
same column layout as the import flow so round-tripping works cleanly:
  - Export your current inventory
  - Edit the spreadsheet (bulk changes, find/replace, etc.)
  - Re-import with "Replace all" mode to sync changes back

The export also serves as a portable backup of just the inventory data
(separate from the full backup tarball, which is monitor + DB + etc.).

Filename includes hostname + ISO date:
  netwatch-inventory-<hostname>-<YYYY-MM-DD>.xlsx

Auth-gated: admin only, since inventory contains hardware identifiers
(MACs, serials) that you might not want to share casually.

Must be applied AFTER patch_ntfy.py.

Run once from ~/netwatch/:
    python3 patch_inv_export.py
    sudo systemctl restart netwatch

Backup of monitor.py saved to monitor.py.bak_invexport.
Idempotent - safe to re-run.
"""

import os, shutil, sys

TARGET = "monitor.py"
BACKUP = "monitor.py.bak_invexport"
SENTINEL = "export_inventory_to_xlsx"


# ─── Module-level helper: build the XLSX bytes ───────────────────────────────

NEW_EXPORT_HELPER = '''def export_inventory_to_xlsx(inventory_db):
    """Build an XLSX file in memory containing all inventory records.

    Column layout matches the import format so the same file can be
    re-imported (in 'Replace all' mode) for round-tripping.

    Returns (bytes, filename) on success, or (None, error_msg) on failure.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        return None, "openpyxl not available"

    import io
    import socket
    from datetime import datetime as _dt

    # Column layout (header -> field name in inventory record)
    # Order matches the user\'s original spreadsheet for familiarity.
    COLUMNS = [
        ("Category",             "category"),
        ("System",               "system"),
        ("Role / Status",        "role"),
        ("CPU",                  "cpu"),
        ("RAM_GB",               "ram_gb"),
        ("GPU",                  "gpu"),
        ("Architecture",         "architecture"),
        ("OS",                   "os"),
        ("Estimated_CPU_Score",  "cpu_score"),
        ("Max_TDP_Watts",        "tdp_watts"),
        ("TPM_Version",          "tpm"),
        ("MAC_Primary",          "mac"),
        ("IP_Address",           "ip"),
        ("Service_Tag_Serial",   "serial"),
        ("Notes",                "notes"),
    ]

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventory"

        # Header row
        for col_idx, (header, _) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")

        # Data rows
        records = inventory_db.list_all()
        for row_idx, rec in enumerate(records, start=2):
            for col_idx, (_, field) in enumerate(COLUMNS, start=1):
                val = rec.get(field)
                # openpyxl writes None as an empty cell, which is what we want
                ws.cell(row=row_idx, column=col_idx, value=val)

        # Auto-fit column widths based on max content length per column.
        # Cap at 50 chars to avoid runaway widths from long Notes fields.
        for col_idx, (header, field) in enumerate(COLUMNS, start=1):
            max_len = len(header)
            for rec in records:
                val = rec.get(field)
                if val is not None:
                    s = str(val)
                    if len(s) > max_len:
                        max_len = len(s)
            # openpyxl column letter
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        # Freeze the header row so it stays visible when scrolling
        ws.freeze_panes = "A2"

        # Write to BytesIO
        buf = io.BytesIO()
        wb.save(buf)

        hostname = socket.gethostname() or "unknown"
        date_str = _dt.now().strftime("%Y-%m-%d")
        filename = f"netwatch-inventory-{hostname}-{date_str}.xlsx"

        return buf.getvalue(), filename
    except Exception as e:
        return None, f"export failed: {e}"


'''


PATCHES = [
    # 1. Insert the helper function. Anchor on the existing
    # import_inventory_from_xlsx function so the new export helper is right
    # next to it logically.
    (
        '''def import_inventory_from_xlsx(inventory_db, xlsx_bytes, mode="add"):''',
        NEW_EXPORT_HELPER + '''def import_inventory_from_xlsx(inventory_db, xlsx_bytes, mode="add"):'''
    ),

    # 2. Add the GET /api/inventory-export endpoint. Anchor on the existing
    # /api/inventory GET handler (the open list endpoint) since that's
    # definitely in do_GET. Insert our new handler right before it.
    (
        '''            if self.path == "/api/inventory":
                # Inventory list is OPEN to viewers (matches the dashboard's view-mode)''',
        '''            if self.path == "/api/inventory-export":
                # Build XLSX in memory and stream it back as a download.
                # Admin-only because inventory contains hardware identifiers.
                if not self._require_auth(admin_only=True):
                    return
                if not inventory_db:
                    self._send_json(500, {"error": "inventory not available"})
                    return
                try:
                    data, result = export_inventory_to_xlsx(inventory_db)
                    if data is None:
                        self._send_json(500, {"error": result})
                        return
                    filename = result
                    self.send_response(200)
                    self.send_header("Content-Type",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    self.send_header("Content-Length", str(len(data)))
                    self.send_header("Content-Disposition",
                                     f'attachment; filename="{filename}"')
                    self.end_headers()
                    self.wfile.write(data)
                    logging.info(f"Inventory export: {filename} ({len(data)} bytes)")
                except Exception as e:
                    logging.exception("inventory export error")
                    self._send_json(500, {"error": str(e)})
                return

            if self.path == "/api/inventory":
                # Inventory list is OPEN to viewers (matches the dashboard's view-mode)'''
    ),

    # 3. Frontend: add "Export XLSX" button to the Inventory toolbar
    (
        '''        <button class="btn" onclick="openInventoryEditor()">+ Add</button>
        <button class="btn" onclick="openImportModal()">Import XLSX</button>''',
        '''        <button class="btn" onclick="openInventoryEditor()">+ Add</button>
        <button class="btn" onclick="openImportModal()">Import XLSX</button>
        <button class="btn" onclick="downloadInventoryExport(this)">Export XLSX</button>'''
    ),

    # 4. Frontend: add the JS handler. Anchor near the existing import logic.
    (
        '''async function submitImport(){''',
        '''async function downloadInventoryExport(btn){
  // Admin-only, so check auth first to give a friendly error
  if(!_authState.logged_in){
    if(_authState.setup_required) openSetup();
    else openLogin(() => downloadInventoryExport(btn));
    return;
  }
  if(!_authState.admin){
    alert("Inventory export requires admin access.");
    return;
  }
  const origText = btn ? btn.textContent : null;
  if(btn){ btn.disabled = true; btn.textContent = "Building..."; }
  try {
    const res = await fetch("/api/inventory-export");
    if(!res.ok){
      let msg = "Export failed (HTTP " + res.status + ")";
      try { const j = await res.json(); if(j.error) msg = j.error; } catch(e){}
      alert(msg);
      return;
    }
    const blob = await res.blob();
    const dispo = res.headers.get("Content-Disposition") || "";
    const m = dispo.match(/filename="?([^";]+)"?/);
    const filename = m ? m[1] : "netwatch-inventory.xlsx";
    const url = URL.createObjectURL(blob);
    const a = document.createElement("a");
    a.href = url; a.download = filename;
    document.body.appendChild(a);
    a.click();
    document.body.removeChild(a);
    setTimeout(() => URL.revokeObjectURL(url), 1000);
  } catch(e){
    alert("Export failed: " + e.message);
  } finally {
    if(btn){ btn.disabled = false; btn.textContent = origText || "Export XLSX"; }
  }
}

async function submitImport(){'''
    ),

    # 5. Bump version
    (
        'netwatch v3.14 - raspberry pi',
        'netwatch v3.15 - raspberry pi'
    ),
]


def main():
    if not os.path.isfile(TARGET):
        print(f"ERROR: {TARGET} not found.")
        sys.exit(1)

    content = open(TARGET).read()

    if SENTINEL in content:
        print(f"NOTE: '{SENTINEL}' found - patch already applied.")
        sys.exit(0)

    if "send_ntfy_alert" not in content:
        print("ERROR: This patch requires patch_ntfy first.")
        sys.exit(1)

    shutil.copy2(TARGET, BACKUP)
    print(f"[OK] Backed up {TARGET} -> {BACKUP}")

    applied = 0
    for i, (old, new) in enumerate(PATCHES, 1):
        count = content.count(old)
        if count == 0:
            print(f"[FAIL] Patch #{i}: target not found")
            shutil.copy2(BACKUP, TARGET); sys.exit(1)
        if count > 1:
            print(f"[FAIL] Patch #{i}: matches {count}x")
            shutil.copy2(BACKUP, TARGET); sys.exit(1)
        content = content.replace(old, new, 1)
        applied += 1

    if 'VERSION = "3.14"' in content:
        content = content.replace('VERSION = "3.14"', 'VERSION = "3.15"', 1)

    open(TARGET, "w").write(content)

    import ast
    try:
        ast.parse(open(TARGET).read())
        print("[OK] Resulting Python is valid")
    except SyntaxError as e:
        print(f"[FAIL] Syntax error: {e}")
        shutil.copy2(BACKUP, TARGET); sys.exit(1)

    print(f"[OK] Applied {applied} patches")
    print()
    print("Next steps:")
    print("  1. sudo systemctl restart netwatch")
    print("  2. Open the Inventory tab")
    print("  3. New 'Export XLSX' button appears next to 'Import XLSX'")
    print("  4. Click it - your inventory downloads as a spreadsheet")
    print()
    print(f"Rollback: cp {BACKUP} {TARGET} && sudo systemctl restart netwatch")


if __name__ == "__main__":
    main()
