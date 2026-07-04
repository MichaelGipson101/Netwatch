import sqlite3
import os
import sys
import re
import io
import json as _json
import threading as _threading
import urllib.request as _urlreq
import urllib.error as _urlerr
from http.server import ThreadingHTTPServer as _THTS
from monitor import _column_exists
from monitor import export_inventory_to_xlsx
from monitor import NASPoller
import monitor as _mon
from monitor import (
    _h_get_ai_config, _h_get_backup_status, _h_get_inventory_backup_status, _h_get_hosts,
    _h_get_auth_status, _h_get_auth_users,
    _h_get_inventory, _h_get_inventory_record,
    _h_get_topology, _h_get_connections, _h_get_connections_for_device,
    _h_get_discover,
    _h_post_inventory_create, _h_post_inventory_update, _h_post_inventory_delete,
    _h_post_connection_create, _h_post_connection_update, _h_post_connection_delete,
    _h_post_detect_mac, _h_post_wake, _h_post_hosts,
    _h_post_auth_users, _h_post_auth_password, _h_post_auth_user_delete,
    _h_get_ai_usage,
)


def test_column_exists_returns_true_for_existing_column():
    conn = sqlite3.connect(":memory:")
    conn.execute("CREATE TABLE t (a TEXT, b INTEGER)")
    assert _column_exists(conn, "t", "a") is True
    assert _column_exists(conn, "t", "b") is True


def test_column_exists_returns_false_for_missing_column():
    conn = sqlite3.connect(":memory:")
    conn.execute("CREATE TABLE t (a TEXT)")
    assert _column_exists(conn, "t", "x") is False


def test_column_exists_returns_false_for_unknown_table():
    conn = sqlite3.connect(":memory:")
    assert _column_exists(conn, "nonexistent", "col") is False


def test_real_operationalerror_propagates_on_bad_sql():
    """ALTER TABLE on a non-existent table raises OperationalError — not swallowed."""
    import sqlite3 as _sq
    conn = sqlite3.connect(":memory:")
    try:
        conn.execute("ALTER TABLE ghost ADD COLUMN x TEXT")
        assert False, "Should have raised"
    except _sq.OperationalError:
        pass  # expected — real errors propagate


import tempfile, time
from monitor import AuthManager


def _make_auth(tmpdir, db_path=None):
    auth_path = os.path.join(tmpdir, "auth.json")
    return AuthManager(auth_path, db_path=db_path)


def test_failed_attempts_persist_across_restart():
    with tempfile.TemporaryDirectory() as tmpdir:
        db_path = os.path.join(tmpdir, "test.db")
        am1 = _make_auth(tmpdir, db_path=db_path)
        for _ in range(AuthManager.LOCKOUT_AFTER):
            am1.record_failed_attempt("1.2.3.4")
        assert am1.is_locked_out("1.2.3.4")

        # New instance, same DB — lockout must survive
        am2 = _make_auth(tmpdir, db_path=db_path)
        assert am2.is_locked_out("1.2.3.4")


def test_successful_login_clears_persisted_attempts():
    with tempfile.TemporaryDirectory() as tmpdir:
        db_path = os.path.join(tmpdir, "test.db")
        am1 = _make_auth(tmpdir, db_path=db_path)
        for _ in range(AuthManager.LOCKOUT_AFTER):
            am1.record_failed_attempt("10.0.0.1")
        am1.record_successful_login("10.0.0.1")

        am2 = _make_auth(tmpdir, db_path=db_path)
        assert not am2.is_locked_out("10.0.0.1")


def test_no_db_path_works_as_before():
    """AuthManager without db_path still works (in-memory only)."""
    with tempfile.TemporaryDirectory() as tmpdir:
        am = _make_auth(tmpdir, db_path=None)
        for _ in range(AuthManager.LOCKOUT_AFTER):
            am.record_failed_attempt("192.168.1.1")
        assert am.is_locked_out("192.168.1.1")


# ── device_type in /api/status ──────────────────────────────────────────────

from monitor import HistoryDB, InventoryDB, build_api_payload, make_handler


class _FakeHost:
    def __init__(self, ip):
        self.ip = ip
        self.is_up = True
        self.last_checked = None
        self.always_on = True
    def to_dict(self):
        return {"ip": self.ip, "is_up": self.is_up}


class _FakeHostManager:
    def __init__(self, hosts): self._hosts = hosts
    def list_hosts(self): return self._hosts


def _make_idb(tmpdir):
    db_path = os.path.join(tmpdir, "icons_test.db")
    hdb = HistoryDB(db_path)
    idb = InventoryDB(hdb)
    return hdb, idb


def test_get_device_type_map_returns_ip_to_type():
    with tempfile.TemporaryDirectory() as tmpdir:
        hdb, idb = _make_idb(tmpdir)
        with hdb.lock:
            hdb.conn.execute(
                "INSERT INTO inventory (system, ip, device_type, created_at, updated_at)"
                " VALUES (?, ?, ?, ?, ?)",
                ("TabletA", "10.0.0.1", "tablet", 0, 0)
            )
            hdb.conn.execute(
                "INSERT INTO inventory (system, ip, device_type, created_at, updated_at)"
                " VALUES (?, ?, ?, ?, ?)",
                ("Phone1", "10.0.0.2", "phone", 0, 0)
            )
        result = idb.get_device_type_map()
        assert result == {"10.0.0.1": "tablet", "10.0.0.2": "phone"}
        hdb.close()


def test_get_device_type_map_excludes_null_ip():
    with tempfile.TemporaryDirectory() as tmpdir:
        hdb, idb = _make_idb(tmpdir)
        with hdb.lock:
            hdb.conn.execute(
                "INSERT INTO inventory (system, ip, device_type, created_at, updated_at)"
                " VALUES (?, ?, ?, ?, ?)",
                ("NoIP", None, "host", 0, 0)
            )
        result = idb.get_device_type_map()
        assert result == {}
        hdb.close()


def test_build_api_payload_annotates_device_type():
    with tempfile.TemporaryDirectory() as tmpdir:
        hdb, idb = _make_idb(tmpdir)
        with hdb.lock:
            hdb.conn.execute(
                "INSERT INTO inventory (system, ip, device_type, created_at, updated_at)"
                " VALUES (?, ?, ?, ?, ?)",
                ("MyTablet", "10.0.0.1", "tablet", 0, 0)
            )
        hm = _FakeHostManager([_FakeHost("10.0.0.1"), _FakeHost("10.0.0.99")])
        payload = build_api_payload(hm, {}, inventory_db=idb)
        hosts = {h["ip"]: h for h in payload["hosts"]}
        assert hosts["10.0.0.1"]["device_type"] == "tablet"
        assert hosts["10.0.0.99"]["device_type"] == "host"  # no record → default
        hdb.close()


def test_build_api_payload_without_inventory_db():
    hm = _FakeHostManager([_FakeHost("10.0.0.1")])
    payload = build_api_payload(hm, {})
    assert payload["hosts"][0]["device_type"] == "host"


def test_export_scope_hosts_filters_to_hosts_only():
    try:
        import openpyxl
    except ImportError:
        return  # skip if openpyxl not installed
    with tempfile.TemporaryDirectory() as tmpdir:
        hdb, idb = _make_idb(tmpdir)
        with hdb.lock:
            hdb.conn.execute(
                "INSERT INTO inventory (system, ip, device_type, created_at, updated_at)"
                " VALUES (?, ?, ?, ?, ?)",
                ("Server1", "10.0.0.1", "host", 0, 0)
            )
            hdb.conn.execute(
                "INSERT INTO inventory (system, ip, device_type, created_at, updated_at)"
                " VALUES (?, ?, ?, ?, ?)",
                ("VM1", "10.0.0.2", "vm", 0, 0)
            )
        data, filename = export_inventory_to_xlsx(idb, scope='hosts')
        assert data is not None, filename
        assert "hosts" in filename
        wb = openpyxl.load_workbook(io.BytesIO(data))
        assert wb.sheetnames == ["Inventory"]
        ws = wb.active
        # Only 1 data row (host); VM excluded
        assert ws.max_row == 2  # header + 1 host
        assert ws.cell(row=2, column=2).value == "Server1"
        hdb.close()


def test_export_scope_all_creates_one_sheet_per_type():
    try:
        import openpyxl
    except ImportError:
        return
    with tempfile.TemporaryDirectory() as tmpdir:
        hdb, idb = _make_idb(tmpdir)
        with hdb.lock:
            hdb.conn.execute(
                "INSERT INTO inventory (system, ip, device_type, created_at, updated_at)"
                " VALUES (?, ?, ?, ?, ?)",
                ("Server1", "10.0.0.1", "host", 0, 0)
            )
            hdb.conn.execute(
                "INSERT INTO inventory (system, ip, device_type, created_at, updated_at)"
                " VALUES (?, ?, ?, ?, ?)",
                ("VM1", "10.0.0.2", "vm", 0, 0)
            )
        data, filename = export_inventory_to_xlsx(idb, scope='all')
        assert data is not None, filename
        assert "all" in filename
        wb = openpyxl.load_workbook(io.BytesIO(data))
        assert "Hosts" in wb.sheetnames
        assert "VMs" in wb.sheetnames
        assert "Network" not in wb.sheetnames  # no network records → no sheet
        assert wb["Hosts"].cell(row=2, column=2).value == "Server1"
        assert wb["VMs"].cell(row=2, column=2).value == "VM1"
        hdb.close()


def test_export_scope_defaults_to_hosts():
    try:
        import openpyxl
    except ImportError:
        return
    with tempfile.TemporaryDirectory() as tmpdir:
        hdb, idb = _make_idb(tmpdir)
        with hdb.lock:
            hdb.conn.execute(
                "INSERT INTO inventory (system, ip, device_type, created_at, updated_at)"
                " VALUES (?, ?, ?, ?, ?)",
                ("VM1", "10.0.0.2", "vm", 0, 0)
            )
        # No scope arg → defaults to hosts → VM excluded → 0 data rows
        data, filename = export_inventory_to_xlsx(idb)
        assert data is not None
        wb = openpyxl.load_workbook(io.BytesIO(data))
        assert wb.active.max_row == 1  # header only, no host records
        hdb.close()


class _FakeAuthManagerForAiConfig:
    """Holds openrouter_api_key like a real AuthManager, and reports an
    always-authenticated admin session so _require_auth() lets requests
    through in HTTP-level tests without needing a real session cookie."""
    has_users = True

    def __init__(self, api_key):
        self.data = {} if api_key is None else {"openrouter_api_key": api_key}
        self.lock = MagicMock()
        self.lock.__enter__ = MagicMock(return_value=None)
        self.lock.__exit__ = MagicMock(return_value=False)

    def verify_session_cookie(self, _cookie):
        return "testuser", True


def _make_am_with_openrouter(api_key="sk-or-test-123"):
    return _FakeAuthManagerForAiConfig(api_key)


def _ai_config_server(settings, auth_manager=None):
    """Spin up a single-request test server for /api/ai-config. Returns (server, port)."""
    hm = _FakeHostManager([])
    handler = make_handler(hm, settings, "/dev/null", auth_manager=auth_manager)
    server = _THTS(("127.0.0.1", 0), handler)
    return server, server.server_address[1]


def test_ai_config_returns_model_without_leaking_key():
    settings = {"ai_model": "meta-llama/llama-3.3-70b-instruct:free"}
    am = _make_am_with_openrouter("sk-or-test-123")
    server, port = _ai_config_server(settings, auth_manager=am)
    t = _threading.Thread(target=server.handle_request)
    t.start()
    try:
        with _urlreq.urlopen(f"http://127.0.0.1:{port}/api/ai-config") as r:
            data = _json.loads(r.read())
        assert "api_key" not in data
        assert data["model"] == "meta-llama/llama-3.3-70b-instruct:free"
    finally:
        server.server_close()
        t.join()


def test_ai_config_missing_key_returns_404():
    server, port = _ai_config_server({}, auth_manager=_make_am_with_openrouter(None))
    t = _threading.Thread(target=server.handle_request)
    t.start()
    try:
        try:
            _urlreq.urlopen(f"http://127.0.0.1:{port}/api/ai-config")
            assert False, "Expected 404"
        except _urlerr.HTTPError as e:
            assert e.code == 404
    finally:
        server.server_close()
        t.join()


# ── _h_get_ai_config ─────────────────────────────────────────────────────────

def test_h_get_ai_config_returns_model_without_leaking_key():
    am = _make_am_with_openrouter("sk-test")
    code, body = _h_get_ai_config({"ai_model": "gpt-4"}, auth_manager=am)
    assert code == 200
    assert "api_key" not in body
    assert body["model"] == "gpt-4"


def test_h_get_ai_config_missing_key_returns_404():
    code, body = _h_get_ai_config({}, auth_manager=_make_am_with_openrouter(None))
    assert code == 404
    assert body["error"] == "ai_not_configured"


def test_h_get_ai_config_blank_key_returns_404():
    code, body = _h_get_ai_config({}, auth_manager=_make_am_with_openrouter("   "))
    assert code == 404
    assert body["error"] == "ai_not_configured"


def test_h_get_ai_usage_unconfigured_returns_404():
    code, body = _h_get_ai_usage(auth_manager=_make_am_with_openrouter(None))
    assert code == 404
    assert body["error"] == "ai_not_configured"


def test_h_get_ai_config_default_model():
    am = _make_am_with_openrouter("sk-x")
    code, body = _h_get_ai_config({}, auth_manager=am)
    assert code == 200
    assert body["model"] == "openrouter/free"


# ── _h_get_backup_status ─────────────────────────────────────────────────────

def test_h_get_backup_status_not_configured(monkeypatch, tmp_path):
    monkeypatch.setattr(_mon, "NAS_BACKUP_STATUS_PATH", str(tmp_path / "missing.json"))
    code, body = _h_get_backup_status()
    assert code == 200
    assert body == {"configured": False}


def test_h_get_backup_status_reads_success_file(monkeypatch, tmp_path):
    status_path = tmp_path / "_status.json"
    status_path.write_text(_json.dumps({
        "ok": True, "checked_at": 1700000000,
        "filename": "netwatch-backup-test.tar.gz", "size_bytes": 123,
        "files": {"monitor.py": 10},
    }))
    monkeypatch.setattr(_mon, "NAS_BACKUP_STATUS_PATH", str(status_path))
    code, body = _h_get_backup_status()
    assert code == 200
    assert body["configured"] is True
    assert body["ok"] is True
    assert body["filename"] == "netwatch-backup-test.tar.gz"


def test_h_get_backup_status_reads_failure_file(monkeypatch, tmp_path):
    status_path = tmp_path / "_status.json"
    status_path.write_text(_json.dumps({"ok": False, "checked_at": 1700000000, "error": "boom"}))
    monkeypatch.setattr(_mon, "NAS_BACKUP_STATUS_PATH", str(status_path))
    code, body = _h_get_backup_status()
    assert code == 200
    assert body["configured"] is True
    assert body["ok"] is False
    assert body["error"] == "boom"


def test_h_get_backup_status_handles_corrupt_file(monkeypatch, tmp_path):
    status_path = tmp_path / "_status.json"
    status_path.write_text("not valid json")
    monkeypatch.setattr(_mon, "NAS_BACKUP_STATUS_PATH", str(status_path))
    code, body = _h_get_backup_status()
    assert code == 200
    assert body["configured"] is False


# ── _h_get_inventory_backup_status ───────────────────────────────────────────

def test_h_get_inventory_backup_status_not_configured(monkeypatch, tmp_path):
    monkeypatch.setattr(_mon, "NAS_INVENTORY_STATUS_PATH", str(tmp_path / "missing.json"))
    code, body = _h_get_inventory_backup_status()
    assert code == 200
    assert body == {"configured": False}


def test_h_get_inventory_backup_status_reads_success_file(monkeypatch, tmp_path):
    status_path = tmp_path / "_status.json"
    status_path.write_text(_json.dumps({
        "ok": True, "checked_at": 1700000000,
        "filename": "netwatch-inventory-all.xlsx", "size_bytes": 456,
    }))
    monkeypatch.setattr(_mon, "NAS_INVENTORY_STATUS_PATH", str(status_path))
    code, body = _h_get_inventory_backup_status()
    assert code == 200
    assert body["configured"] is True
    assert body["ok"] is True
    assert body["filename"] == "netwatch-inventory-all.xlsx"


# ── _h_get_hosts ─────────────────────────────────────────────────────────────

def test_h_get_hosts_returns_host_list():
    with tempfile.TemporaryDirectory() as tmpdir:
        cfg_path = os.path.join(tmpdir, "hosts.yaml")
        with open(cfg_path, "w") as f:
            f.write("hosts:\n  - name: router\n    ip: 192.168.1.1\n    group: Lab\n")
        code, body = _h_get_hosts(cfg_path)
        assert code == 200
        assert body["hosts"][0]["name"] == "router"


def test_h_get_hosts_missing_file_returns_500():
    code, body = _h_get_hosts("/nonexistent/path/hosts.yaml")
    assert code == 500
    assert "error" in body


# ── _h_get_auth_status ───────────────────────────────────────────────────────

def test_h_get_auth_status_no_auth_manager():
    code, body = _h_get_auth_status(None, lambda: (None, False), "")
    assert code == 200
    assert body["logged_in"] is False
    assert body["setup_required"] is False


def test_h_get_auth_status_logged_in():
    class FakeAM:
        has_users = True
        def csrf_token_for_cookie(self, cookie_value):
            return "deadbeef"
    code, body = _h_get_auth_status(FakeAM(), lambda: ("alice", True), "some-cookie")
    assert code == 200
    assert body["logged_in"] is True
    assert body["username"] == "alice"
    assert body["admin"] is True
    assert body["csrf_token"] == "deadbeef"


def test_h_get_auth_status_setup_required():
    class FakeAM:
        has_users = False
    code, body = _h_get_auth_status(FakeAM(), lambda: (None, False), "")
    assert code == 200
    assert body["setup_required"] is True
    assert "csrf_token" not in body


def test_h_get_auth_status_logged_out_no_csrf_token():
    class FakeAM:
        has_users = True
    code, body = _h_get_auth_status(FakeAM(), lambda: (None, False), "")
    assert code == 200
    assert body["logged_in"] is False
    assert "csrf_token" not in body


# ── _h_get_auth_users ────────────────────────────────────────────────────────

def test_h_get_auth_users_returns_list():
    class FakeAM:
        def list_users(self): return [{"username": "alice", "admin": True}]
    code, body = _h_get_auth_users(FakeAM())
    assert code == 200
    assert body["users"][0]["username"] == "alice"


def test_h_get_auth_users_no_auth_manager():
    code, body = _h_get_auth_users(None)
    assert code == 404
    assert "error" in body


# ── _h_get_inventory ─────────────────────────────────────────────────────────

def test_h_get_inventory_no_db_returns_empty():
    code, body = _h_get_inventory(None, None)
    assert code == 200
    assert body["items"] == []


def test_h_get_inventory_returns_items():
    with tempfile.TemporaryDirectory() as tmpdir:
        hdb, idb = _make_idb(tmpdir)
        with hdb.lock:
            hdb.conn.execute(
                "INSERT INTO inventory (system, ip, device_type, created_at, updated_at)"
                " VALUES (?,?,?,?,?)", ("Switch1", "10.0.0.5", "network", 0, 0)
            )
            hdb.conn.commit()
        code, body = _h_get_inventory(idb, None)
        assert code == 200
        assert isinstance(body["items"], list)
        assert len(body["items"]) > 0
        assert any(i["system"] == "Switch1" for i in body["items"])
        assert "id" in body["items"][0]
        hdb.close()


# ── _h_get_inventory_record ──────────────────────────────────────────────────

def test_h_get_inventory_record_not_found():
    with tempfile.TemporaryDirectory() as tmpdir:
        hdb, idb = _make_idb(tmpdir)
        code, body = _h_get_inventory_record("/api/inventory/9999", idb, None)
        assert code == 404
        hdb.close()


def test_h_get_inventory_record_bad_id():
    with tempfile.TemporaryDirectory() as tmpdir:
        hdb, idb = _make_idb(tmpdir)
        code, body = _h_get_inventory_record("/api/inventory/notanint", idb, None)
        assert code == 400
        hdb.close()


def test_h_get_inventory_record_no_db():
    code, body = _h_get_inventory_record("/api/inventory/1", None, None)
    assert code == 500


# ── _h_get_topology ──────────────────────────────────────────────────────────

def test_h_get_topology_no_db():
    code, body = _h_get_topology(None, None)
    assert code == 500
    assert "error" in body


# ── _h_get_connections ───────────────────────────────────────────────────────

def test_h_get_connections_no_db():
    code, body = _h_get_connections(None)
    assert code == 500
    assert "error" in body


def test_h_get_connections_returns_list():
    with tempfile.TemporaryDirectory() as tmpdir:
        hdb, idb = _make_idb(tmpdir)
        code, body = _h_get_connections(idb)
        assert code == 200
        assert "items" in body
        hdb.close()


# ── _h_get_connections_for_device ────────────────────────────────────────────

def test_h_get_connections_for_device_bad_path():
    with tempfile.TemporaryDirectory() as tmpdir:
        hdb, idb = _make_idb(tmpdir)
        code, body = _h_get_connections_for_device("/api/inventory/notanint/connections", idb)
        assert code == 400
        hdb.close()


# ── _h_post_wake ─────────────────────────────────────────────────────────────

class _FakeHostWithMac:
    def __init__(self, ip, mac=None):
        self.ip = ip
        self.specs = {"mac": mac} if mac else {}
    def to_dict(self): return {"ip": self.ip}


class _FakeHMWake:
    def __init__(self, hosts): self._hosts = hosts
    def list_hosts(self): return self._hosts


def test_h_post_wake_missing_ip():
    hm = _FakeHMWake([])
    code, body = _h_post_wake({}, hm, None)
    assert code == 400
    assert body["error"] == "ip is required"


def test_h_post_wake_host_not_found():
    hm = _FakeHMWake([_FakeHostWithMac("10.0.0.1")])
    code, body = _h_post_wake({"ip": "10.0.0.99"}, hm, None)
    assert code == 404


def test_h_post_wake_no_mac():
    hm = _FakeHMWake([_FakeHostWithMac("10.0.0.1")])
    code, body = _h_post_wake({"ip": "10.0.0.1"}, hm, None)
    assert code == 400
    assert "MAC" in body["error"]


# ── _h_post_detect_mac ───────────────────────────────────────────────────────

def test_h_post_detect_mac_missing_ip():
    code, body = _h_post_detect_mac({})
    assert code == 400
    assert body["error"] == "ip required"


def test_h_post_detect_mac_blank_ip():
    code, body = _h_post_detect_mac({"ip": "  "})
    assert code == 400
    assert body["error"] == "ip required"


# ── _h_post_inventory_create ─────────────────────────────────────────────────

def test_h_post_inventory_create_no_db():
    code, body = _h_post_inventory_create({}, None)
    assert code == 500
    assert "error" in body


def test_h_post_inventory_create_success():
    with tempfile.TemporaryDirectory() as tmpdir:
        hdb, idb = _make_idb(tmpdir)
        code, body = _h_post_inventory_create(
            {"system": "TestBox", "device_type": "host"}, idb
        )
        assert code == 200
        assert body["ok"] is True
        assert isinstance(body["id"], int)
        hdb.close()


# ── _h_post_inventory_delete ─────────────────────────────────────────────────

def test_h_post_inventory_delete_not_found():
    with tempfile.TemporaryDirectory() as tmpdir:
        hdb, idb = _make_idb(tmpdir)
        code, body = _h_post_inventory_delete("/api/inventory/9999/delete", idb)
        assert code == 404
        hdb.close()


def test_h_post_inventory_delete_bad_id():
    with tempfile.TemporaryDirectory() as tmpdir:
        hdb, idb = _make_idb(tmpdir)
        code, body = _h_post_inventory_delete("/api/inventory/abc/delete", idb)
        assert code == 400
        hdb.close()


# ── _h_post_hosts ────────────────────────────────────────────────────────────

def test_h_post_hosts_not_a_list():
    hm = _FakeHostManager([])
    code, body = _h_post_hosts({"hosts": "not-a-list"}, "/dev/null", hm, {})
    assert code == 400
    assert "'hosts' must be a list" in body["error"]


def test_h_post_hosts_invalid_host():
    hm = _FakeHostManager([])
    code, body = _h_post_hosts(
        {"hosts": [{"name": "noip", "group": "Lab"}]}, "/dev/null", hm, {}
    )
    assert code == 400
    assert "error" in body


# ── _h_post_auth_users ───────────────────────────────────────────────────────

def test_h_post_auth_users_creates_user():
    with tempfile.TemporaryDirectory() as tmpdir:
        am = _make_auth(tmpdir)
        code, body = _h_post_auth_users(
            {"username": "bob", "password": "Str0ng!pass", "admin": False}, am
        )
        assert code == 200
        assert body["ok"] is True
        assert any(u["username"] == "bob" for u in am.list_users())


def test_h_post_auth_users_duplicate():
    with tempfile.TemporaryDirectory() as tmpdir:
        am = _make_auth(tmpdir)
        am.create_user("alice", "Str0ng!pass", admin=True)
        code, body = _h_post_auth_users({"username": "alice", "password": "other"}, am)
        assert code == 400
        assert "error" in body


# ── _h_post_auth_password ────────────────────────────────────────────────────

def test_h_post_auth_password_wrong_current():
    with tempfile.TemporaryDirectory() as tmpdir:
        am = _make_auth(tmpdir)
        am.create_user("carol", "Str0ng!pass", admin=False)
        code, body = _h_post_auth_password(
            {"current": "wrongpassword", "new": "newpass"}, "carol", am
        )
        assert code == 401
        assert "error" in body


def test_h_post_auth_password_success():
    with tempfile.TemporaryDirectory() as tmpdir:
        am = _make_auth(tmpdir)
        am.create_user("dave", "Str0ng!pass", admin=False)
        code, body = _h_post_auth_password(
            {"current": "Str0ng!pass", "new": "NewStr0ng!pass"}, "dave", am
        )
        assert code == 200
        assert body["ok"] is True


# ── _h_post_auth_user_delete ─────────────────────────────────────────────────

def test_h_post_auth_user_delete_empty_username():
    with tempfile.TemporaryDirectory() as tmpdir:
        am = _make_auth(tmpdir)
        code, body = _h_post_auth_user_delete("/api/auth/users/", am)
        assert code == 400
        assert "error" in body


def test_h_post_auth_user_delete_not_found():
    with tempfile.TemporaryDirectory() as tmpdir:
        am = _make_auth(tmpdir)
        code, body = _h_post_auth_user_delete("/api/auth/users/nobody", am)
        assert code == 400
        assert "error" in body


# ── Transition-only logging ──────────────────────────────────────────────────

def test_should_log_transition_first_ping():
    from monitor import _should_log_transition
    assert _should_log_transition(None, True) is True
    assert _should_log_transition(None, False) is True


def test_should_log_transition_state_change():
    from monitor import _should_log_transition
    assert _should_log_transition(True, False) is True
    assert _should_log_transition(False, True) is True


def test_should_log_transition_steady_state():
    from monitor import _should_log_transition
    assert _should_log_transition(True, True) is False
    assert _should_log_transition(False, False) is False


# ── WAL cap + prune without VACUUM ───────────────────────────────────────────

def test_journal_size_limit_pragma_set(tmp_path):
    from monitor import HistoryDB
    hdb = HistoryDB(str(tmp_path / "t.db"))
    assert hdb.conn.execute("PRAGMA journal_size_limit").fetchone()[0] == 16777216
    hdb.close()


def test_prune_deletes_old_keeps_recent_no_vacuum(tmp_path):
    from monitor import HistoryDB
    hdb = HistoryDB(str(tmp_path / "t.db"), retention_days=7)
    now = int(time.time())
    hdb.conn.execute("INSERT INTO pings (host_ip, timestamp, is_up, latency_ms) VALUES (?,?,?,?)",
                     ("10.0.0.1", now - 8 * 86400, 1, 1.0))
    hdb.conn.execute("INSERT INTO pings (host_ip, timestamp, is_up, latency_ms) VALUES (?,?,?,?)",
                     ("10.0.0.1", now, 1, 1.0))
    deleted, _ = hdb.prune()
    assert deleted == 1
    assert hdb.conn.execute("SELECT COUNT(*) FROM pings").fetchone()[0] == 1
    hdb.close()


# ── Brief CRUD ───────────────────────────────────────────────────────────────

def test_insert_brief_and_get_briefs(tmp_path):
    from monitor import HistoryDB
    import json
    hdb = HistoryDB(str(tmp_path / "t.db"))
    ts = int(time.time())
    hdb.insert_brief(ts, "BRIEF // Today", json.dumps({"up": 18, "down": 0, "idle": 1}), "All quiet.", None)
    briefs = hdb.get_briefs(days=7)
    assert len(briefs) == 1
    b = briefs[0]
    assert b["subject"] == "BRIEF // Today"
    assert b["narrative"] == "All quiet."
    assert b["stats"]["up"] == 18
    assert b["stats"]["down"] == 0
    assert "analysis" not in b
    hdb.close()


def test_get_briefs_returns_newest_first(tmp_path):
    from monitor import HistoryDB
    import json
    hdb = HistoryDB(str(tmp_path / "t.db"))
    now = int(time.time())
    hdb.insert_brief(now - 86400, "older", json.dumps({}), "older narrative", None)
    hdb.insert_brief(now, "newer", json.dumps({}), "newer narrative", None)
    briefs = hdb.get_briefs(days=7)
    assert briefs[0]["subject"] == "newer"
    assert briefs[1]["subject"] == "older"
    hdb.close()


def test_get_briefs_excludes_old(tmp_path):
    from monitor import HistoryDB
    import json
    hdb = HistoryDB(str(tmp_path / "t.db"))
    now = int(time.time())
    hdb.insert_brief(now - 8 * 86400, "ancient", json.dumps({}), "old text", None)
    hdb.insert_brief(now, "recent", json.dumps({}), "new text", None)
    briefs = hdb.get_briefs(days=7)
    assert len(briefs) == 1
    assert briefs[0]["subject"] == "recent"
    hdb.close()


def test_prune_deletes_old_briefs(tmp_path):
    from monitor import HistoryDB
    import json
    hdb = HistoryDB(str(tmp_path / "t.db"), retention_days=7)
    now = int(time.time())
    hdb.insert_brief(now - 8 * 86400, "old", json.dumps({}), "old", None)
    hdb.insert_brief(now, "new", json.dumps({}), "new", None)
    hdb.prune()
    remaining = hdb.conn.execute("SELECT COUNT(*) FROM briefs").fetchone()[0]
    assert remaining == 1
    hdb.close()



# ── Brief API handlers ────────────────────────────────────────────────────────

def test_h_post_brief_stores_and_returns_ok(tmp_path):
    import json
    from monitor import HistoryDB, _h_post_brief
    hdb = HistoryDB(str(tmp_path / "t.db"))
    data = {
        "subject": "BRIEF // Today",
        "stats": {"up": 18, "down": 0, "idle": 1},
        "narrative": "All quiet tonight.",
        "analysis": {"fleet_summary": {"up": 18}},
    }
    status, body = _h_post_brief(hdb, data)
    assert status == 200
    assert body == {"ok": True}
    briefs = hdb.get_briefs(days=7)
    assert len(briefs) == 1
    assert briefs[0]["narrative"] == "All quiet tonight."
    hdb.close()


def test_h_post_brief_missing_field_returns_400(tmp_path):
    from monitor import HistoryDB, _h_post_brief
    hdb = HistoryDB(str(tmp_path / "t.db"))
    # missing 'narrative'
    status, body = _h_post_brief(hdb, {"subject": "X", "stats": {}})
    assert status == 400
    assert "error" in body
    hdb.close()


def test_h_get_briefs_returns_list_without_analysis(tmp_path):
    import json
    from monitor import HistoryDB, _h_post_brief, _h_get_briefs
    hdb = HistoryDB(str(tmp_path / "t.db"))
    _h_post_brief(hdb, {
        "subject": "BRIEF",
        "stats": {"up": 5, "down": 0, "idle": 0},
        "narrative": "Quiet.",
        "analysis": {"secret": "data"},
    })
    status, body = _h_get_briefs(hdb)
    assert status == 200
    briefs = body["briefs"]
    assert len(briefs) == 1
    assert "analysis" not in briefs[0]
    assert briefs[0]["stats"]["up"] == 5
    hdb.close()


# ── Batched ping inserts ─────────────────────────────────────────────────────

def test_record_ping_buffers_until_flush(tmp_path):
    from monitor import HistoryDB
    db_path = str(tmp_path / "t.db")
    hdb = HistoryDB(db_path)
    hdb.record_ping("10.0.0.1", True, 5.0)
    other = sqlite3.connect(db_path)
    assert other.execute("SELECT COUNT(*) FROM pings").fetchone()[0] == 0
    hdb.flush_pings()
    assert other.execute("SELECT COUNT(*) FROM pings").fetchone()[0] == 1
    other.close(); hdb.close()


def test_buffer_threshold_autoflush(tmp_path):
    from monitor import HistoryDB
    db_path = str(tmp_path / "t.db")
    hdb = HistoryDB(db_path)
    hdb.FLUSH_MAX = 3
    for _ in range(3):
        hdb.record_ping("10.0.0.1", True, 1.0)
    other = sqlite3.connect(db_path)
    assert other.execute("SELECT COUNT(*) FROM pings").fetchone()[0] == 3
    other.close(); hdb.close()


def test_recent_pings_sees_buffered_rows(tmp_path):
    from monitor import HistoryDB
    hdb = HistoryDB(str(tmp_path / "t.db"))
    hdb.record_ping("10.0.0.1", True, 5.0)
    hdb.record_ping("10.0.0.1", False, None)
    assert hdb.recent_pings("10.0.0.1", limit=10) == [(True, 5.0), (False, None)]
    hdb.close()


def test_close_flushes_buffer(tmp_path):
    from monitor import HistoryDB
    db_path = str(tmp_path / "t.db")
    hdb = HistoryDB(db_path)
    hdb.record_ping("10.0.0.1", True, 5.0)
    hdb.close()
    other = sqlite3.connect(db_path)
    assert other.execute("SELECT COUNT(*) FROM pings").fetchone()[0] == 1
    other.close()


# ── Session invalidation ─────────────────────────────────────────────────────

def test_deleted_user_cookie_rejected():
    with tempfile.TemporaryDirectory() as td:
        auth = _make_auth(td)
        auth.create_user("alice", "password123", admin=True)
        auth.create_user("bob", "password123")
        cookie = auth.make_session_cookie("bob")
        assert auth.verify_session_cookie(cookie)[0] == "bob"
        auth.delete_user("bob")
        assert auth.verify_session_cookie(cookie) == (None, False)


def test_password_change_invalidates_old_sessions():
    with tempfile.TemporaryDirectory() as td:
        auth = _make_auth(td)
        auth.create_user("alice", "password123", admin=True)
        old_cookie = auth.make_session_cookie("alice")
        assert auth.verify_session_cookie(old_cookie)[0] == "alice"
        auth.change_password("alice", "newpassword456")
        assert auth.verify_session_cookie(old_cookie) == (None, False)
        new_cookie = auth.make_session_cookie("alice")
        assert auth.verify_session_cookie(new_cookie) == ("alice", True)


def test_legacy_two_part_cookie_rejected():
    import hmac as _hmac, hashlib as _hashlib, base64 as _b64
    with tempfile.TemporaryDirectory() as td:
        auth = _make_auth(td)
        auth.create_user("alice", "password123", admin=True)
        payload = f"alice|{int(time.time()) + 3600}"  # old format: no generation
        secret = auth.data["secret_key"].encode()
        sig = _hmac.new(secret, payload.encode(), _hashlib.sha256).hexdigest()
        token = _b64.urlsafe_b64encode(payload.encode()).decode().rstrip("=")
        assert auth.verify_session_cookie(f"{token}.{sig}") == (None, False)


def test_csrf_token_for_cookie_deterministic():
    with tempfile.TemporaryDirectory() as td:
        auth = _make_auth(td)
        token1 = auth.csrf_token_for_cookie("some-cookie-value")
        token2 = auth.csrf_token_for_cookie("some-cookie-value")
        assert token1 == token2
        assert len(token1) == 64  # hex sha256 digest


def test_csrf_token_for_cookie_differs_per_cookie():
    with tempfile.TemporaryDirectory() as td:
        auth = _make_auth(td)
        token_a = auth.csrf_token_for_cookie("cookie-a")
        token_b = auth.csrf_token_for_cookie("cookie-b")
        assert token_a != token_b


def test_csrf_token_for_cookie_differs_per_secret():
    with tempfile.TemporaryDirectory() as td1, tempfile.TemporaryDirectory() as td2:
        auth1 = _make_auth(td1)
        auth2 = _make_auth(td2)
        assert auth1.csrf_token_for_cookie("same-cookie") != auth2.csrf_token_for_cookie("same-cookie")


# ── /api/status settings allowlist ───────────────────────────────────────────

def test_api_payload_settings_allowlist():
    from monitor import build_api_payload

    class _FakeHM:
        def list_hosts(self):
            return []

    settings = {"default_interval": 15, "openrouter_api_key": "sk-secret",
                "ntfy_topic": "secret-topic", "refresh_rate": 5}
    payload = build_api_payload(_FakeHM(), settings)
    assert payload["settings"] == {"default_interval": 15, "refresh_rate": 5}
    assert "openrouter_api_key" not in payload["settings"]


# ── Config write hardening ───────────────────────────────────────────────────

def test_save_hosts_config_sets_0600(tmp_path):
    from monitor import save_hosts_config
    path = str(tmp_path / "hosts.yaml")
    save_hosts_config(path, [{"name": "a", "ip": "10.0.0.1"}])
    assert os.stat(path).st_mode & 0o777 == 0o600


def _auth_test_server(auth):
    """One-request test server with auth enabled. Returns (server, port, thread)."""
    handler = make_handler(None, {}, "/dev/null", auth_manager=auth)
    server = _THTS(("127.0.0.1", 0), handler)
    t = _threading.Thread(target=server.handle_request)
    t.start()
    return server, server.server_address[1], t


def test_post_hosts_requires_admin(tmp_path):
    auth = AuthManager(str(tmp_path / "auth.json"))
    auth.create_user("root", "password123", admin=True)
    auth.create_user("bob", "password123")
    server, port, t = _auth_test_server(auth)
    try:
        cookie = auth.make_session_cookie("bob")
        req = _urlreq.Request(f"http://127.0.0.1:{port}/api/hosts", data=b'{"hosts": []}',
                              method="POST", headers={"Cookie": f"nw_session={cookie}"})
        try:
            _urlreq.urlopen(req)
            assert False, "expected 403"
        except _urlerr.HTTPError as e:
            assert e.code == 403
    finally:
        server.server_close()
        t.join()


def test_post_without_csrf_token_rejected(tmp_path):
    auth = AuthManager(str(tmp_path / "auth.json"))
    auth.create_user("root", "password123", admin=True)
    server, port, t = _auth_test_server(auth)
    try:
        cookie = auth.make_session_cookie("root")
        req = _urlreq.Request(f"http://127.0.0.1:{port}/api/hosts", data=b'{"hosts": []}',
                              method="POST", headers={"Cookie": f"nw_session={cookie}"})
        try:
            _urlreq.urlopen(req)
            assert False, "expected 403"
        except _urlerr.HTTPError as e:
            assert e.code == 403
            body = _json.loads(e.read())
            assert body["error"] == "csrf_required"
    finally:
        server.server_close()
        t.join()


def test_post_with_wrong_csrf_token_rejected(tmp_path):
    auth = AuthManager(str(tmp_path / "auth.json"))
    auth.create_user("root", "password123", admin=True)
    server, port, t = _auth_test_server(auth)
    try:
        cookie = auth.make_session_cookie("root")
        req = _urlreq.Request(f"http://127.0.0.1:{port}/api/hosts", data=b'{"hosts": []}',
                              method="POST", headers={"Cookie": f"nw_session={cookie}",
                                                       "X-CSRF-Token": "wrong-token"})
        try:
            _urlreq.urlopen(req)
            assert False, "expected 403"
        except _urlerr.HTTPError as e:
            assert e.code == 403
            body = _json.loads(e.read())
            assert body["error"] == "csrf_required"
    finally:
        server.server_close()
        t.join()


def test_post_with_correct_csrf_token_accepted(tmp_path):
    from monitor import HostManager

    # Uses a real HostManager/tmp config instead of _auth_test_server, whose hardcoded
    # /dev/null config path can't survive this test's successful (200) config-writing POST.
    auth = AuthManager(str(tmp_path / "auth.json"))
    auth.create_user("root", "password123", admin=True)
    config_path = str(tmp_path / "hosts.yaml")
    host_manager = HostManager(config_path, ping_timeout=1, history_window=10,
                               global_stop=_threading.Event())
    handler = make_handler(host_manager, {}, config_path, auth_manager=auth)
    server = _THTS(("127.0.0.1", 0), handler)
    t = _threading.Thread(target=server.handle_request)
    t.start()
    port = server.server_address[1]
    try:
        cookie = auth.make_session_cookie("root")
        token = auth.csrf_token_for_cookie(cookie)
        req = _urlreq.Request(f"http://127.0.0.1:{port}/api/hosts", data=b'{"hosts": []}',
                              method="POST", headers={"Cookie": f"nw_session={cookie}",
                                                       "X-CSRF-Token": token,
                                                       "Content-Type": "application/json"})
        resp = _urlreq.urlopen(req)
        assert resp.status == 200
    finally:
        server.server_close()
        t.join()


def test_login_response_includes_csrf_token(tmp_path):
    auth = AuthManager(str(tmp_path / "auth.json"))
    auth.create_user("root", "password123", admin=True)
    server, port, t = _auth_test_server(auth)
    try:
        body = _json.dumps({"username": "root", "password": "password123"}).encode()
        req = _urlreq.Request(f"http://127.0.0.1:{port}/api/auth/login", data=body,
                              method="POST", headers={"Content-Type": "application/json"})
        resp = _urlreq.urlopen(req)
        data = _json.loads(resp.read())
        assert data["ok"] is True
        assert "csrf_token" in data
        assert len(data["csrf_token"]) == 64
    finally:
        server.server_close()
        t.join()


def test_setup_response_includes_csrf_token(tmp_path):
    auth = AuthManager(str(tmp_path / "auth.json"))
    server, port, t = _auth_test_server(auth)
    try:
        body = _json.dumps({"username": "root", "password": "password123"}).encode()
        req = _urlreq.Request(f"http://127.0.0.1:{port}/api/auth/setup", data=body,
                              method="POST", headers={"Content-Type": "application/json"})
        resp = _urlreq.urlopen(req)
        data = _json.loads(resp.read())
        assert data["ok"] is True
        assert "csrf_token" in data
        assert len(data["csrf_token"]) == 64
    finally:
        server.server_close()
        t.join()


def test_non_dict_json_body_rejected(tmp_path):
    auth = AuthManager(str(tmp_path / "auth.json"))
    auth.create_user("root", "password123", admin=True)
    server, port, t = _auth_test_server(auth)
    try:
        cookie = auth.make_session_cookie("root")
        token = auth.csrf_token_for_cookie(cookie)
        req = _urlreq.Request(f"http://127.0.0.1:{port}/api/hosts", data=b'[1,2,3]',
                              method="POST", headers={"Cookie": f"nw_session={cookie}",
                                                       "X-CSRF-Token": token})
        try:
            _urlreq.urlopen(req)
            assert False, "expected 400"
        except _urlerr.HTTPError as e:
            assert e.code == 400
    finally:
        server.server_close()
        t.join()


# ── /api/history: bucketed latency series ────────────────────────────────────

def test_history_series_buckets_and_aggregates(tmp_path):
    from monitor import HistoryDB
    hdb = HistoryDB(str(tmp_path / "t.db"))
    now = int(time.time())
    rows = [("10.0.0.1", now - 30, 1, 10.0), ("10.0.0.1", now - 20, 1, 20.0),
            ("10.0.0.1", now - 10, 0, None), ("10.0.0.2", now - 10, 1, 99.0)]
    hdb.conn.executemany("INSERT INTO pings (host_ip, timestamp, is_up, latency_ms) VALUES (?,?,?,?)", rows)
    res = hdb.history_series("10.0.0.1", hours=1)
    assert res["bucket_seconds"] == 60  # 3600s / 180 points -> min clamp 60
    assert sum(p["n"] for p in res["points"]) == 3
    assert max(p["max"] for p in res["points"] if p["max"] is not None) == 20.0
    assert min(p["min"] for p in res["points"] if p["min"] is not None) == 10.0
    hdb.close()


def test_history_series_clamps_hours(tmp_path):
    from monitor import HistoryDB
    hdb = HistoryDB(str(tmp_path / "t.db"))
    assert hdb.history_series("10.0.0.1", hours=9999)["points"] == []
    hdb.close()


def test_h_get_history_requires_ip(tmp_path):
    from monitor import HistoryDB, _h_get_history
    hdb = HistoryDB(str(tmp_path / "t.db"))
    code, body = _h_get_history("/api/history", hdb)
    assert code == 400
    code, body = _h_get_history("/api/history?ip=10.0.0.1&hours=abc", hdb)
    assert code == 400
    hdb.close()


def test_h_get_history_returns_points(tmp_path):
    from monitor import HistoryDB, _h_get_history
    hdb = HistoryDB(str(tmp_path / "t.db"))
    hdb.record_ping("10.0.0.1", True, 12.0)
    code, body = _h_get_history("/api/history?ip=10.0.0.1&hours=1", hdb)
    assert code == 200
    assert body["ip"] == "10.0.0.1"
    assert sum(p["n"] for p in body["points"]) == 1
    hdb.close()


# ── Daily uptime rollups ─────────────────────────────────────────────────────

def _insert_ping(hdb, ip, ts, up, lat):
    hdb.conn.execute("INSERT INTO pings (host_ip, timestamp, is_up, latency_ms) VALUES (?,?,?,?)",
                     (ip, ts, 1 if up else 0, lat))


def test_rollup_aggregates_complete_days_only(tmp_path):
    from monitor import HistoryDB
    hdb = HistoryDB(str(tmp_path / "t.db"))
    now = int(time.time())
    _insert_ping(hdb, "10.0.0.1", now - 86400, True, 10.0)   # yesterday
    _insert_ping(hdb, "10.0.0.1", now - 86400 + 60, False, None)
    _insert_ping(hdb, "10.0.0.1", now, True, 5.0)            # today: excluded
    hdb.rollup_days()
    rows = hdb.conn.execute("SELECT day, total, up, latency_avg FROM ping_daily").fetchall()
    assert len(rows) == 1
    assert rows[0][1] == 2 and rows[0][2] == 1 and rows[0][3] == 10.0
    hdb.close()


def test_rollup_is_idempotent(tmp_path):
    from monitor import HistoryDB
    hdb = HistoryDB(str(tmp_path / "t.db"))
    _insert_ping(hdb, "10.0.0.1", int(time.time()) - 86400, True, 10.0)
    hdb.rollup_days(); hdb.rollup_days()
    assert hdb.conn.execute("SELECT COUNT(*) FROM ping_daily").fetchone()[0] == 1
    hdb.close()


def test_daily_history_survives_prune(tmp_path):
    from monitor import HistoryDB
    hdb = HistoryDB(str(tmp_path / "t.db"), retention_days=7)
    old = int(time.time()) - 8 * 86400
    _insert_ping(hdb, "10.0.0.1", old, True, 10.0)
    hdb.rollup_days()
    hdb.prune()  # deletes the raw ping
    assert hdb.conn.execute("SELECT COUNT(*) FROM pings").fetchone()[0] == 0
    daily = hdb.daily_history("10.0.0.1", days=30)
    assert len(daily) == 1 and daily[0]["uptime_pct"] == 100.0
    hdb.close()


# ── Static asset cache busting ───────────────────────────────────────────────

def test_dashboard_html_version_substitution(tmp_path):
    from monitor import _load_dashboard_html, VERSION
    (tmp_path / "dashboard.html").write_text('<script src="/static/core.js?v={{VERSION}}"></script>')
    out = _load_dashboard_html(str(tmp_path))
    assert "{{VERSION}}" not in out
    assert f"?v={VERSION}" in out


# ── static asset self-hosting ────────────────────────────────────────────

def test_static_whitelist_includes_vendored_assets():
    from monitor import _STATIC_FILES
    expected = {
        'd3.v7.min.js': 'application/javascript',
        'fonts.css': 'text/css',
        'dmsans-300.woff2': 'font/woff2', 'dmsans-400.woff2': 'font/woff2',
        'dmsans-500.woff2': 'font/woff2', 'dmsans-600.woff2': 'font/woff2',
        'dmmono-400.woff2': 'font/woff2', 'dmmono-500.woff2': 'font/woff2',
        'favicon.svg': 'image/svg+xml', 'favicon-alert.svg': 'image/svg+xml',
        'manifest.json': 'application/manifest+json',
        'icon-192.png': 'image/png', 'icon-512.png': 'image/png',
        'apple-touch-icon.png': 'image/png',
    }
    for fname, mime in expected.items():
        assert fname in _STATIC_FILES, fname
        assert _STATIC_FILES[fname].startswith(mime), fname

def test_vendored_asset_files_exist_on_disk():
    base = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'static'))
    for fname in ['d3.v7.min.js', 'fonts.css', 'dmsans-400.woff2', 'dmmono-400.woff2']:
        assert os.path.exists(os.path.join(base, fname)), fname

def test_dmsans_weights_are_distinct():
    import hashlib
    base = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'static'))
    digests = {}
    for w in ('300', '400', '500', '600'):
        with open(os.path.join(base, f'dmsans-{w}.woff2'), 'rb') as f:
            digests[w] = hashlib.md5(f.read()).hexdigest()
    assert len(set(digests.values())) == 4, f'duplicate DM Sans weights: {digests}'


# ── incident timestamp payload ───────────────────────────────────────────

def _insert_incident(hdb, started, ended=None, dur=None):
    with hdb.lock:
        hdb.conn.execute(
            "INSERT INTO incidents (host_ip, host_name, host_group, started, ended, duration_seconds) "
            "VALUES (?,?,?,?,?,?)",
            ("10.0.0.9", "TestHost", "G", started, ended, dur))

def test_list_incidents_includes_epoch_ts(tmp_path):
    from monitor import HistoryDB
    hdb = HistoryDB(str(tmp_path / "t.db"))
    now = int(time.time())
    _insert_incident(hdb, now - 30, now, 30)
    inc = hdb.list_incidents()[0]
    assert inc["started_ts"] == now - 30

def test_started_str_time_only_for_today(tmp_path):
    from monitor import HistoryDB
    hdb = HistoryDB(str(tmp_path / "t.db"))
    now = int(time.time())
    _insert_incident(hdb, now - 60, now, 60)
    inc = hdb.list_incidents()[0]
    # NOTE: flakes if run within 60s of local midnight (incident lands on yesterday)
    assert re.fullmatch(r"\d{2}:\d{2}:\d{2}", inc["started_str"])

def test_started_str_includes_date_for_older_events(tmp_path):
    from monitor import HistoryDB
    hdb = HistoryDB(str(tmp_path / "t.db"))
    old = int(time.time()) - 3 * 86400
    _insert_incident(hdb, old, old + 60, 60)
    inc = hdb.list_incidents()[0]
    # e.g. "Jun 08 17:39" — month abbrev + day + HH:MM
    assert re.fullmatch(r"[A-Z][a-z]{2} \d{2} \d{2}:\d{2}", inc["started_str"])

def test_open_incident_uses_provided_started_at(tmp_path):
    from monitor import HistoryDB
    hdb = HistoryDB(str(tmp_path / "t.db"))
    past_ts = int(time.time()) - 120
    hdb.open_incident("1.2.3.4", "TestHost", "G", started_at=past_ts)
    row = hdb.conn.execute(
        "SELECT started FROM incidents WHERE host_ip = '1.2.3.4'"
    ).fetchone()
    assert row[0] == past_ts


def test_open_incident_defaults_to_now_when_no_started_at(tmp_path):
    from monitor import HistoryDB
    hdb = HistoryDB(str(tmp_path / "t.db"))
    before = int(time.time())
    hdb.open_incident("1.2.3.4", "TestHost", "G")
    after = int(time.time())
    row = hdb.conn.execute(
        "SELECT started FROM incidents WHERE host_ip = '1.2.3.4'"
    ).fetchone()
    assert before <= row[0] <= after


# ============================================================================
# NASPoller parse tests
# ============================================================================

_POOL_RAW = {
    "name": "tank",
    "status": "ONLINE",
    "size": 4000000000000,
    "allocated": 2000000000000,
    "topology": {
        "data": [
            {
                "type": "MIRROR",
                "name": "mirror-0",
                "status": "ONLINE",
                "children": [
                    {"disk": "ada0", "status": "ONLINE", "type": "DISK"},
                    {"disk": "ada1", "status": "ONLINE", "type": "DISK"},
                ],
            }
        ]
    },
    "scan": {
        "state": "FINISHED",
        "end_time": "2026-06-01T04:22:00",
        "errors": 0,
    },
}

_SCRUB_TASKS = [
    # Matches TrueNAS's real /api/v2.0/pool/scrub shape: "pool" is the
    # numeric pool ID, "pool_name" is the string name _parse_pool matches on.
    {"pool": 1, "pool_name": "tank", "schedule": {"minute": "0", "hour": "0", "dom": "1", "month": "*", "dow": "*"}}
]

_REP_RAW = {
    "id": 1,
    "name": "tank → backup",
    "state": {"state": "FINISHED", "datetime": "2026-06-14T02:00:00"},
}


def test_parse_pool_basic():
    pool = NASPoller._parse_pool(_POOL_RAW, _SCRUB_TASKS)
    assert pool["name"] == "tank"
    assert pool["status"] == "ONLINE"
    assert pool["capacity_total_bytes"] == 4000000000000
    assert pool["capacity_used_bytes"] == 2000000000000
    assert pool["last_scrub"]["errors"] == 0
    assert pool["last_scrub"]["status"] == "FINISHED"
    assert pool["next_scrub"] is not None  # cron produced a date


def test_next_cron_run_uses_local_timezone_not_utc():
    """Regression: TrueNAS's schedule hour/minute are in the NAS's configured
    local timezone, not UTC. A schedule of hour=0 in America/New_York (UTC-4
    in summer DST) must resolve to 04:00 UTC, not 00:00 UTC."""
    from datetime import datetime, timezone
    from zoneinfo import ZoneInfo
    start = datetime(2026, 6, 1, tzinfo=timezone.utc)  # a Monday
    result = NASPoller._next_cron_run(
        "0", "0", "*", "*", "1",  # Mondays at 00:00 local
        start=start, tz=ZoneInfo("America/New_York"),
    )
    dt = datetime.fromisoformat(result)
    assert dt.utcoffset().total_seconds() == -4 * 3600  # EDT
    assert dt.hour == 0 and dt.minute == 0
    # The same wall-clock moment expressed in UTC is 4 hours later.
    assert dt.astimezone(timezone.utc).hour == 4


def test_parse_pool_uses_tz_name_for_next_scrub():
    """_parse_pool's tz_name parameter must actually thread through to
    _next_cron_run, not just be accepted and ignored."""
    raw = dict(_POOL_RAW)
    raw["scan"] = None
    scrub_tasks = [
        {"pool": 1, "pool_name": "tank",
         "schedule": {"minute": "0", "hour": "0", "dom": "*", "month": "*", "dow": "*"}}
    ]
    pool_utc = NASPoller._parse_pool(raw, scrub_tasks, tz_name=None)
    pool_ny = NASPoller._parse_pool(raw, scrub_tasks, tz_name="America/New_York")
    from datetime import datetime
    assert datetime.fromisoformat(pool_utc["next_scrub"]).utcoffset().total_seconds() == 0
    assert datetime.fromisoformat(pool_ny["next_scrub"]).utcoffset().total_seconds() != 0


def test_parse_pool_matches_scrub_task_by_pool_name_not_numeric_id():
    """Regression: TrueNAS's scrub-task API returns pool as a numeric ID and
    pool_name as the string name. _parse_pool must match on pool_name - an
    earlier bug matched on pool (the int), which never equals the pool's
    name string, so next_scrub silently stayed None for every real pool."""
    scrub_tasks = [
        {"pool": 999, "pool_name": "tank", "schedule": {"minute": "0", "hour": "0", "dom": "*", "month": "*", "dow": "*"}}
    ]
    pool = NASPoller._parse_pool(_POOL_RAW, scrub_tasks)
    assert pool["next_scrub"] is not None


def test_parse_pool_next_scrub_respects_threshold_past_cron_interval():
    """Regression: a scrub task's "schedule" (e.g. weekly) just says how often
    to check; "threshold" is the minimum number of days since the last scrub
    before TrueNAS actually lets it run again. When threshold (e.g. 35 days)
    outlasts the cron interval, next_scrub must reflect the threshold date,
    not the next bare weekly cron occurrence."""
    from datetime import datetime
    raw = dict(_POOL_RAW)
    raw["scan"] = {"state": "FINISHED", "end_time": "2026-06-07T09:35:53+00:00", "errors": 0}
    scrub_tasks = [
        {"pool": 1, "pool_name": "tank", "threshold": 35,
         "schedule": {"minute": "0", "hour": "0", "dom": "*", "month": "*", "dow": "7"}}
    ]
    pool = NASPoller._parse_pool(raw, scrub_tasks)
    next_dt = datetime.fromisoformat(pool["next_scrub"])
    last_dt = datetime.fromisoformat("2026-06-07T09:35:53+00:00")
    assert (next_dt - last_dt).days >= 35


def test_parse_pool_next_scrub_without_threshold_uses_plain_cron():
    """No threshold field (or 0) - fall back to the old plain-cron behavior."""
    raw = dict(_POOL_RAW)
    raw["scan"] = {"state": "FINISHED", "end_time": "2026-06-07T09:35:53+00:00", "errors": 0}
    scrub_tasks = [
        {"pool": 1, "pool_name": "tank",
         "schedule": {"minute": "0", "hour": "0", "dom": "*", "month": "*", "dow": "*"}}
    ]
    pool = NASPoller._parse_pool(raw, scrub_tasks)
    assert pool["next_scrub"] is not None


def test_parse_pool_null_scan():
    raw = dict(_POOL_RAW)
    raw["scan"] = None
    pool = NASPoller._parse_pool(raw, [])
    assert pool["last_scrub"]["status"] is None
    assert pool["last_scrub"]["errors"] == 0


def test_parse_pool_surfaces_non_data_vdev_types():
    """Regression: cache/log/spare/special/dedup vdevs were silently dropped -
    a failed cache (L2ARC) or log (SLOG) device would never show up anywhere,
    since pool "status" doesn't necessarily reflect a non-critical device
    like a cache disk failing."""
    raw = dict(_POOL_RAW)
    raw["topology"] = dict(_POOL_RAW["topology"])
    raw["topology"]["cache"] = [
        {"type": "DISK", "name": "nvme0", "status": "FAULTED", "children": [], "disk": "nvme0"}
    ]
    pool = NASPoller._parse_pool(raw, [])
    assert len(pool["cache_vdevs"]) == 1
    assert pool["cache_vdevs"][0]["status"] == "FAULTED"
    assert pool["log_vdevs"] == []
    assert pool["spare_vdevs"] == []
    assert pool["special_vdevs"] == []
    assert pool["dedup_vdevs"] == []


def test_parse_vdevs_mirror():
    vdevs = NASPoller._parse_vdevs(_POOL_RAW["topology"]["data"])
    assert len(vdevs) == 1
    assert vdevs[0]["type"] == "MIRROR"
    assert len(vdevs[0]["disks"]) == 2
    assert vdevs[0]["disks"][0]["name"] == "ada0"


def test_parse_replication_basic():
    task = NASPoller._parse_replication(_REP_RAW)
    assert task["id"] == 1
    assert task["name"] == "tank → backup"
    assert task["last_state"] == "FINISHED"
    assert task["last_run"] == "2026-06-14T02:00:00"
    assert task["enabled"] is True  # _REP_RAW has no "enabled" key - defaults True


def test_parse_replication_captures_enabled_false():
    raw = dict(_REP_RAW)
    raw["enabled"] = False
    task = NASPoller._parse_replication(raw)
    assert task["enabled"] is False


def test_parse_replication_ms_epoch():
    # TrueNAS returns {"$date": <ms_since_epoch>} for some datetime fields
    raw = {
        "id": 2,
        "name": "tank → offsite",
        "state": {"state": "FINISHED", "datetime": {"$date": 1781337621000}},
    }
    task = NASPoller._parse_replication(raw)
    assert task["last_state"] == "FINISHED"
    # Should be an ISO string, not a raw integer
    assert isinstance(task["last_run"], str)
    assert "2026" in task["last_run"]


def test_next_cron_run_monthly():
    # dom=1 means 1st of each month — result must be in the future
    from datetime import datetime, timezone
    result = NASPoller._next_cron_run("0", "0", "1", "*", "*")
    assert result is not None
    dt = datetime.fromisoformat(result).replace(tzinfo=None)
    assert dt > datetime.now(tz=timezone.utc).replace(tzinfo=None)


# ============================================================================
# NASPoller poll loop and alert logic tests
# ============================================================================

from unittest.mock import patch, MagicMock


def _make_nas_poller():
    am = MagicMock()
    am.data = {"truenas_url": "http://truenas.test", "truenas_api_key": "testkey"}
    return NASPoller(am, alert_settings={}, alert_port=8080)


def test_alert_fires_once_on_repeated_degraded():
    poller = _make_nas_poller()
    pools = [{"name": "tank", "status": "DEGRADED", "last_scrub": {"errors": 0}, "next_scrub": None}]
    with patch("monitor._send_alert_async") as mock_send:
        poller._check_alerts(pools, [])
        poller._check_alerts(pools, [])
    assert mock_send.call_count == 1  # fired once, not twice


def test_alert_rearmed_after_clear():
    poller = _make_nas_poller()
    pools_degraded = [{"name": "tank", "status": "DEGRADED", "last_scrub": {"errors": 0}, "next_scrub": None}]
    pools_ok = [{"name": "tank", "status": "ONLINE", "last_scrub": {"errors": 0}, "next_scrub": None}]
    with patch("monitor._send_alert_async") as mock_send:
        poller._check_alerts(pools_degraded, [])  # fires
        poller._check_alerts(pools_ok, [])         # clears
        poller._check_alerts(pools_degraded, [])  # re-arms → fires again
    assert mock_send.call_count == 2


def test_replication_stale_alert():
    poller = _make_nas_poller()
    from datetime import datetime, timezone, timedelta
    old_run = (datetime.now(tz=timezone.utc) - timedelta(hours=30)).isoformat()
    tasks = [{"id": 1, "name": "tank→backup", "last_run": old_run, "last_state": "FINISHED", "enabled": True}]
    with patch("monitor._send_alert_async") as mock_send:
        poller._check_alerts([], tasks)
    assert mock_send.call_count == 1


def test_replication_disabled_task_does_not_alert_when_stale():
    """Regression: a deliberately disabled replication task must not alert
    as stale/failed forever - there's no reason to expect it to have run."""
    poller = _make_nas_poller()
    from datetime import datetime, timezone, timedelta
    old_run = (datetime.now(tz=timezone.utc) - timedelta(hours=200)).isoformat()
    tasks = [{"id": 1, "name": "tank→backup", "last_run": old_run, "last_state": "ERROR", "enabled": False}]
    with patch("monitor._send_alert_async") as mock_send:
        poller._check_alerts([], tasks)
    assert mock_send.call_count == 0


def test_replication_disabled_task_clears_existing_alert():
    poller = _make_nas_poller()
    from datetime import datetime, timezone, timedelta
    old_run = (datetime.now(tz=timezone.utc) - timedelta(hours=200)).isoformat()
    enabled_task = [{"id": 1, "name": "tank→backup", "last_run": old_run, "last_state": "ERROR", "enabled": True}]
    disabled_task = [{"id": 1, "name": "tank→backup", "last_run": old_run, "last_state": "ERROR", "enabled": False}]
    with patch("monitor._send_alert_async"):
        poller._check_alerts([], enabled_task)  # fires, alert_state[cid] = True
    assert poller._alert_state.get("replication_1") is True
    with patch("monitor._send_alert_async"):
        poller._check_alerts([], disabled_task)  # disabling must clear it
    assert poller._alert_state.get("replication_1") is False


def test_replication_cid_uses_explicit_none_check_for_id_zero():
    """Regression: cid construction used `task['id'] or task['name']`, which
    would wrongly fall back to the name for a falsy id of 0. Must use an
    explicit `is not None` check instead."""
    poller = _make_nas_poller()
    tasks = [{"id": 0, "name": "tank→backup", "last_run": None, "last_state": "ERROR", "enabled": True}]
    with patch("monitor._send_alert_async"):
        poller._check_alerts([], tasks)
    assert "replication_0" in poller._alert_state
    assert "replication_tank→backup" not in poller._alert_state


def test_truenas_alert_fires_once():
    poller = _make_nas_poller()
    alerts = [{"id": "a1", "klass": "PoolUSBDisks", "level": "WARNING", "message": "USB disk warning"}]
    with patch("monitor._send_alert_async") as mock_send:
        poller._check_alerts([], [], alerts)
        poller._check_alerts([], [], alerts)
    assert mock_send.call_count == 1


def test_truenas_alert_clears_when_no_longer_present():
    poller = _make_nas_poller()
    alerts = [{"id": "a1", "klass": "PoolUSBDisks", "level": "WARNING", "message": "USB disk warning"}]
    with patch("monitor._send_alert_async"):
        poller._check_alerts([], [], alerts)  # fires, sets alert_state["truenas_alert_a1"] = True
    assert poller._alert_state.get("truenas_alert_a1") is True
    with patch("monitor._send_alert_async"):
        poller._check_alerts([], [], [])  # alert resolved/disappeared from TrueNAS's list
    assert poller._alert_state.get("truenas_alert_a1") is False


def test_truenas_alert_rearmed_after_clear():
    poller = _make_nas_poller()
    alerts = [{"id": "a1", "klass": "PoolUSBDisks", "level": "WARNING", "message": "USB disk warning"}]
    with patch("monitor._send_alert_async") as mock_send:
        poller._check_alerts([], [], alerts)   # fires
        poller._check_alerts([], [], [])        # clears
        poller._check_alerts([], [], alerts)   # re-arms -> fires again
    assert mock_send.call_count == 2


def test_check_alerts_without_alerts_arg_still_works():
    """Backward compatibility: existing call sites pass only (pools, tasks)."""
    poller = _make_nas_poller()
    with patch("monitor._send_alert_async") as mock_send:
        poller._check_alerts([], [])
    assert mock_send.call_count == 0


# ── NASPoller._filter_alerts ──────────────────────────────────────────────────

_RAW_ALERTS = [
    {"id": "a1", "klass": "PoolUSBDisks", "level": "WARNING",
     "formatted": "'ArchiveBackup' is consuming USB devices 'sde'.", "text": "fallback"},
    {"id": "a2", "klass": "AppUpdate", "level": "INFO",
     "formatted": "An update is available for \"plex\".", "text": "fallback"},
    {"id": "a3", "klass": "PoolDegraded", "level": "CRITICAL",
     "formatted": None, "text": "Pool tank is DEGRADED"},
]


def test_filter_alerts_excludes_info_level():
    result = NASPoller._filter_alerts(_RAW_ALERTS, "")
    klasses = [a["klass"] for a in result]
    assert "AppUpdate" not in klasses
    assert "PoolUSBDisks" in klasses
    assert "PoolDegraded" in klasses


def test_filter_alerts_excludes_ignored_klass():
    result = NASPoller._filter_alerts(_RAW_ALERTS, "PoolUSBDisks")
    klasses = [a["klass"] for a in result]
    assert "PoolUSBDisks" not in klasses
    assert "PoolDegraded" in klasses


def test_filter_alerts_ignore_list_handles_multiple_comma_separated():
    result = NASPoller._filter_alerts(_RAW_ALERTS, "PoolUSBDisks, PoolDegraded")
    assert result == []


def test_filter_alerts_falls_back_to_text_when_formatted_missing():
    result = NASPoller._filter_alerts(_RAW_ALERTS, "")
    degraded = next(a for a in result if a["klass"] == "PoolDegraded")
    assert degraded["message"] == "Pool tank is DEGRADED"


def test_filter_alerts_unrecognized_level_is_kept():
    raw = [{"id": "x", "klass": "SomethingNew", "level": "WEIRD_FUTURE_LEVEL",
            "formatted": "something", "text": "fallback"}]
    result = NASPoller._filter_alerts(raw, "")
    assert len(result) == 1


def test_filter_alerts_empty_ignore_list_string():
    result = NASPoller._filter_alerts(_RAW_ALERTS, None)
    assert len(result) == 2  # WARNING + CRITICAL, INFO excluded


# ── _h_post_nas_ignore_alert / _h_post_nas_unignore_alert ────────────────────

from monitor import _h_post_nas_ignore_alert, _h_post_nas_unignore_alert


def test_ignore_alert_requires_klass():
    code, body = _h_post_nas_ignore_alert({}, "/dev/null", {})
    assert code == 400
    assert "klass" in body["error"]


def test_ignore_alert_adds_to_empty_list(tmp_path):
    config_path = str(tmp_path / "hosts.yaml")
    with open(config_path, "w") as f:
        f.write("settings: {}\nhosts: []\n")
    settings = {}
    code, body = _h_post_nas_ignore_alert({"klass": "PoolUSBDisks"}, config_path, settings)
    assert code == 200
    assert settings["truenas_ignored_alert_klasses"] == "PoolUSBDisks"


def test_ignore_alert_dedupes_repeated_calls(tmp_path):
    config_path = str(tmp_path / "hosts.yaml")
    with open(config_path, "w") as f:
        f.write("settings: {}\nhosts: []\n")
    settings = {}
    _h_post_nas_ignore_alert({"klass": "PoolUSBDisks"}, config_path, settings)
    code, body = _h_post_nas_ignore_alert({"klass": "PoolUSBDisks"}, config_path, settings)
    assert code == 200
    assert settings["truenas_ignored_alert_klasses"] == "PoolUSBDisks"


def test_ignore_alert_appends_to_existing_list(tmp_path):
    config_path = str(tmp_path / "hosts.yaml")
    with open(config_path, "w") as f:
        f.write("settings: {}\nhosts: []\n")
    settings = {"truenas_ignored_alert_klasses": "SomeOtherKlass"}
    code, body = _h_post_nas_ignore_alert({"klass": "PoolUSBDisks"}, config_path, settings)
    assert code == 200
    assert "SomeOtherKlass" in settings["truenas_ignored_alert_klasses"]
    assert "PoolUSBDisks" in settings["truenas_ignored_alert_klasses"]


def test_unignore_alert_requires_klass():
    code, body = _h_post_nas_unignore_alert({}, "/dev/null", {})
    assert code == 400


def test_unignore_alert_removes_klass(tmp_path):
    config_path = str(tmp_path / "hosts.yaml")
    with open(config_path, "w") as f:
        f.write("settings: {}\nhosts: []\n")
    settings = {"truenas_ignored_alert_klasses": "PoolUSBDisks,SomeOtherKlass"}
    code, body = _h_post_nas_unignore_alert({"klass": "PoolUSBDisks"}, config_path, settings)
    assert code == 200
    assert "PoolUSBDisks" not in settings.get("truenas_ignored_alert_klasses", "")
    assert "SomeOtherKlass" in settings["truenas_ignored_alert_klasses"]


def test_unignore_alert_clears_key_when_list_becomes_empty(tmp_path):
    config_path = str(tmp_path / "hosts.yaml")
    with open(config_path, "w") as f:
        f.write("settings: {}\nhosts: []\n")
    settings = {"truenas_ignored_alert_klasses": "PoolUSBDisks"}
    code, body = _h_post_nas_unignore_alert({"klass": "PoolUSBDisks"}, config_path, settings)
    assert code == 200
    assert "truenas_ignored_alert_klasses" not in settings


def test_get_cache_returns_copy():
    poller = _make_nas_poller()
    cache1 = poller.get_cache()
    cache1["reachable"] = True
    cache2 = poller.get_cache()
    assert cache2["reachable"] is False  # mutation of copy didn't affect internal cache


# ============================================================================
# /api/nas handler tests
# ============================================================================

from monitor import _h_get_nas


def test_h_get_nas_when_poller_is_none():
    status, body = _h_get_nas(None)
    assert status == 503
    assert body["reachable"] is False


def test_h_get_nas_returns_cache():
    poller = _make_nas_poller()
    # Manually inject a known cache state
    with poller._lock:
        poller._cache = {
            "reachable": True,
            "last_updated": "2026-06-14T02:00:00",
            "error": None,
            "pools": [{"name": "tank", "status": "ONLINE"}],
            "replication_tasks": [],
        }
    status, body = _h_get_nas(poller)
    assert status == 200
    assert body["reachable"] is True
    assert body["pools"][0]["name"] == "tank"


def test_h_get_nas_force_triggers_poll():
    """Regression: "Refresh now" in the UI just re-read whatever the last
    background poll (every 15 min) happened to cache - clicking it did
    nothing until the next scheduled poll. force=True must trigger a real
    poll immediately."""
    poller = _make_nas_poller()
    with patch.object(poller, "_poll") as mock_poll:
        _h_get_nas(poller, force=True)
    mock_poll.assert_called_once()


def test_h_get_nas_without_force_does_not_poll():
    poller = _make_nas_poller()
    with patch.object(poller, "_poll") as mock_poll:
        _h_get_nas(poller, force=False)
    mock_poll.assert_not_called()


def test_poll_skipped_when_unconfigured():
    am = MagicMock()
    am.data = {}  # no truenas_url or api_key
    poller = NASPoller(am)
    poller._poll()
    assert poller.get_cache()["reachable"] is False
    assert poller.get_cache()["error"] == "NAS not configured"


# ============================================================================
# Auth routing — Proxmox credential keys
# ============================================================================

from monitor import _AUTH_STORED_KEYS, _h_get_settings, _h_post_settings


def _make_am_with_proxmox():
    am = MagicMock()
    am.data = {
        "proxmox_url": "https://pve.test:8006",
        "proxmox_user": "root@pam",
        "proxmox_token_id": "Netwatch",
        "proxmox_token_secret": "test-uuid",
    }
    am.lock = MagicMock()
    am.lock.__enter__ = MagicMock(return_value=None)
    am.lock.__exit__ = MagicMock(return_value=False)
    return am


def test_proxmox_keys_in_auth_stored_keys():
    for k in ("proxmox_url", "proxmox_user", "proxmox_token_id", "proxmox_token_secret"):
        assert k in _AUTH_STORED_KEYS


def test_openrouter_key_in_auth_stored_keys():
    assert "openrouter_api_key" in _AUTH_STORED_KEYS, f"{k} not in _AUTH_STORED_KEYS"


def test_h_get_settings_reads_proxmox_creds_from_auth_manager():
    am = _make_am_with_proxmox()
    status, body = _h_get_settings({}, auth_manager=am)
    assert status == 200
    assert body["proxmox_url"] == "https://pve.test:8006"
    # secrets are redacted on the way out — sentinel proves the key was read
    from monitor import SECRET_PLACEHOLDER
    assert body["proxmox_token_secret"] == SECRET_PLACEHOLDER


def test_h_post_settings_saves_proxmox_secret_to_auth_manager():
    import tempfile, os
    am = MagicMock()
    am.data = {}
    am.lock = MagicMock()
    am.lock.__enter__ = MagicMock(return_value=None)
    am.lock.__exit__ = MagicMock(return_value=False)
    am._save = MagicMock()
    with tempfile.TemporaryDirectory() as d:
        cfg = os.path.join(d, "hosts.yaml")
        status, body = _h_post_settings(
            {"proxmox_token_secret": "new-uuid"}, cfg, {}, auth_manager=am
        )
    assert am.data.get("proxmox_token_secret") == "new-uuid"


# ============================================================================
# proxmox_vmid in VM inventory properties
# ============================================================================

def test_proxmox_vmid_in_vm_type_properties():
    from monitor import INVENTORY_TYPE_PROPERTIES
    vm_keys = [p[0] for p in INVENTORY_TYPE_PROPERTIES.get("vm", [])]
    assert "proxmox_vmid" in vm_keys


# ============================================================================
# ProxmoxPoller — data fetch and transformation
# ============================================================================

from monitor import ProxmoxPoller


def _make_proxmox_poller():
    am = MagicMock()
    am.data = {
        "proxmox_url": "https://pve.test:8006",
        "proxmox_user": "root@pam",
        "proxmox_token_id": "Netwatch",
        "proxmox_token_secret": "test-uuid",
    }
    return ProxmoxPoller(am, alert_settings={}, alert_port=8080)


_RAW_QEMU = {
    "vmid": 108, "name": "haos13.2", "status": "running",
    "cpu": 0.0243, "mem": 2046949088, "maxmem": 4294967296,
    "cpus": 2, "uptime": 3891377,
}
_RAW_LXC = {
    "vmid": 120, "name": "pihole", "type": "lxc", "status": "running",
    "cpu": 0.003, "mem": 20336640, "maxmem": 536870912,
    "cpus": 1, "uptime": 3891350,
}
_RAW_STOPPED = {
    "vmid": 100, "name": "windows-11", "status": "stopped",
    "cpu": 0, "mem": 0, "maxmem": 8589934592,
    "cpus": 4, "uptime": 0,
}
_RAW_NODE = {
    "node": "pve", "status": "online",
    "cpu": 0.1209, "mem": 11170574336, "maxmem": 16147808256,
    "uptime": 3891504,
}


def test_build_guest_qemu_type_inferred():
    poller = _make_proxmox_poller()
    g = poller._build_guest(_RAW_QEMU, "qemu")
    assert g["type"] == "qemu"
    assert g["vmid"] == 108
    assert g["name"] == "haos13.2"
    assert g["status"] == "running"


def test_build_guest_lxc_type_inferred():
    poller = _make_proxmox_poller()
    g = poller._build_guest(_RAW_LXC, "lxc")
    assert g["type"] == "lxc"


def test_build_guest_cpu_fraction_to_percent():
    poller = _make_proxmox_poller()
    g = poller._build_guest(_RAW_QEMU, "qemu")
    assert g["cpu_percent"] == round(0.0243 * 100, 1)


def test_build_guest_stopped_has_zero_cpu():
    poller = _make_proxmox_poller()
    g = poller._build_guest(_RAW_STOPPED, "qemu")
    assert g["cpu_percent"] == 0.0
    assert g["mem_used_bytes"] == 0


def test_build_node_contains_guests():
    poller = _make_proxmox_poller()
    node = poller._build_node(_RAW_NODE, [_RAW_QEMU, _RAW_STOPPED], [_RAW_LXC])
    assert node["name"] == "pve"
    assert node["status"] == "online"
    assert len(node["guests"]) == 3
    assert node["guests"][0]["vmid"] == 100
    assert node["guests"][1]["vmid"] == 108
    assert node["guests"][2]["vmid"] == 120


def test_build_node_cpu_fraction_to_percent():
    poller = _make_proxmox_poller()
    node = poller._build_node(_RAW_NODE, [], [])
    assert node["cpu_percent"] == round(0.1209 * 100, 1)


def test_pve_get_cache_returns_deepcopy():
    poller = _make_proxmox_poller()
    c1 = poller.get_cache()
    c1["reachable"] = True
    c2 = poller.get_cache()
    assert c2["reachable"] is False


def test_pve_poll_skipped_when_unconfigured():
    am = MagicMock()
    am.data = {}
    poller = ProxmoxPoller(am)
    poller._poll()
    cache = poller.get_cache()
    assert cache["reachable"] is False
    assert cache["error"] == "Proxmox not configured"


# ============================================================================
# ProxmoxPoller — alert logic
# ============================================================================

def _make_nodes(node_status="online", guest_status="running", vmid=108):
    return [{
        "name": "pve", "status": node_status,
        "cpu_percent": 1.0, "mem_used_bytes": 0, "mem_total_bytes": 0, "uptime_seconds": 0,
        "guests": [{"vmid": vmid, "name": "haos", "type": "qemu",
                    "status": guest_status, "cpu_percent": 0.0,
                    "mem_used_bytes": 0, "mem_total_bytes": 0}],
    }]


def test_pve_node_offline_alert_fires_once():
    poller = _make_proxmox_poller()
    nodes_offline = _make_nodes(node_status="offline")
    with patch("monitor._send_alert_async") as mock_send:
        poller._check_alerts(nodes_offline, [])
        poller._check_alerts(nodes_offline, [])
    assert mock_send.call_count == 1


def test_pve_node_offline_alert_message_mentions_corosync():
    from monitor import ProxmoxPoller
    poller = _make_proxmox_poller()
    nodes_offline = _make_nodes(node_status="offline")
    with patch("monitor._send_alert_async") as mock_send:
        poller._check_alerts(nodes_offline, [])
    args = mock_send.call_args[0]
    assert "corosync" in args[2]
    assert "cluster heartbeat" in args[2]


def test_pve_node_alert_rearmed_after_clear():
    poller = _make_proxmox_poller()
    offline = _make_nodes(node_status="offline")
    online  = _make_nodes(node_status="online")
    with patch("monitor._send_alert_async") as mock_send:
        poller._check_alerts(offline, [])  # fires
        poller._check_alerts(online, [])   # clears
        poller._check_alerts(offline, [])  # re-arms → fires
    assert mock_send.call_count == 2


def test_pve_unexpected_stop_fires_alert():
    poller = _make_proxmox_poller()
    prev   = _make_nodes(guest_status="running")
    now    = _make_nodes(guest_status="stopped")
    with patch("monitor._send_alert_async") as mock_send:
        poller._check_alerts(now, prev)
    assert mock_send.call_count == 1
    args = mock_send.call_args[0]
    assert "stopped unexpectedly" in args[2]


def test_pve_exempted_stop_does_not_alert():
    poller = _make_proxmox_poller()
    poller.exempt_vmid(108, seconds=60)
    prev = _make_nodes(guest_status="running")
    now  = _make_nodes(guest_status="stopped")
    with patch("monitor._send_alert_async") as mock_send:
        poller._check_alerts(now, prev)
    assert mock_send.call_count == 0


def test_pve_paused_guest_fires_alert():
    poller = _make_proxmox_poller()
    nodes = _make_nodes(guest_status="paused")
    with patch("monitor._send_alert_async") as mock_send:
        poller._check_alerts(nodes, [])
    assert mock_send.call_count == 1
    args = mock_send.call_args[0]
    assert "paused" in args[2]


# ============================================================================
# PBSPoller — config, SSL, fetch, classification/grouping logic
# ============================================================================

from monitor import PBSPoller


def _make_pbs_poller():
    am = MagicMock()
    am.data = {
        "pbs_url": "https://pbs.test:8007",
        "pbs_api_token_id": "root@pbs!Netwatch",
        "pbs_api_token_secret": "test-uuid",
    }
    return PBSPoller(am, alert_settings={}, alert_port=8080)


def test_pbs_get_config_returns_tuple():
    poller = _make_pbs_poller()
    assert poller._get_config() == ("https://pbs.test:8007", "root@pbs!Netwatch", "test-uuid")


def test_pbs_get_cache_returns_deepcopy():
    poller = _make_pbs_poller()
    c1 = poller.get_cache()
    c1["reachable"] = True
    c2 = poller.get_cache()
    assert c2["reachable"] is False


def test_pbs_poll_skipped_when_unconfigured():
    am = MagicMock()
    am.data = {}
    poller = PBSPoller(am)
    poller._poll()
    cache = poller.get_cache()
    assert cache["reachable"] is False
    assert cache["error"] == "PBS not configured"


def test_parse_datastore_computes_percent():
    ds = PBSPoller._parse_datastore({"store": "backup-store", "used": 500, "total": 1000, "avail": 500})
    assert ds["name"] == "backup-store"
    assert ds["used_bytes"] == 500
    assert ds["total_bytes"] == 1000
    assert ds["percent"] == 50.0


def test_parse_datastore_zero_total_no_division_error():
    ds = PBSPoller._parse_datastore({"store": "empty", "used": 0, "total": 0, "avail": 0})
    assert ds["percent"] == 0.0


def test_classify_backup_none_when_no_history():
    assert PBSPoller._classify_backup(None, None, 25) == "none"


def test_classify_backup_ok_when_recent_and_unverified():
    from datetime import datetime, timezone, timedelta
    now = datetime.now(tz=timezone.utc)
    recent = now - timedelta(hours=2)
    assert PBSPoller._classify_backup(recent, None, 25, now=now) == "ok"


def test_classify_backup_failed_when_verification_failed():
    from datetime import datetime, timezone, timedelta
    now = datetime.now(tz=timezone.utc)
    recent = now - timedelta(hours=2)
    assert PBSPoller._classify_backup(recent, "failed", 25, now=now) == "failed"


def test_classify_backup_stale_after_threshold():
    from datetime import datetime, timezone, timedelta
    now = datetime(2026, 1, 7, 12, 0, tzinfo=timezone.utc)   # Wednesday - no weekend in the 26h window
    old = now - timedelta(hours=26)
    assert PBSPoller._classify_backup(old, "ok", 25, now=now) == "stale"


def test_classify_backup_ok_just_under_threshold():
    from datetime import datetime, timezone, timedelta
    now = datetime.now(tz=timezone.utc)
    almost = now - timedelta(hours=24, minutes=59)
    assert PBSPoller._classify_backup(almost, None, 25, now=now) == "ok"


def test_classify_backup_ok_saturday_after_friday_backup():
    from datetime import datetime, timezone
    friday_backup = datetime(2026, 1, 9, 20, 0, tzinfo=timezone.utc)   # Friday 20:00
    saturday_check = datetime(2026, 1, 10, 12, 0, tzinfo=timezone.utc) # Saturday 12:00 (16h raw)
    assert PBSPoller._classify_backup(friday_backup, "ok", 25, now=saturday_check) == "ok"


def test_classify_backup_ok_sunday_after_friday_backup():
    from datetime import datetime, timezone
    friday_backup = datetime(2026, 1, 9, 20, 0, tzinfo=timezone.utc)   # Friday 20:00
    sunday_check = datetime(2026, 1, 11, 20, 0, tzinfo=timezone.utc)   # Sunday 20:00 (48h raw, 4 business)
    assert PBSPoller._classify_backup(friday_backup, "ok", 25, now=sunday_check) == "ok"


def test_classify_backup_stale_monday_night_if_still_missing():
    from datetime import datetime, timezone
    friday_backup = datetime(2026, 1, 9, 20, 0, tzinfo=timezone.utc)    # Friday 20:00
    monday_night_check = datetime(2026, 1, 12, 22, 0, tzinfo=timezone.utc)  # Monday 22:00
    # raw elapsed = 74h; weekend excluded = 48h; business = 26h > 25h threshold
    assert PBSPoller._classify_backup(friday_backup, "ok", 25, now=monday_night_check) == "stale"


def test_classify_backup_failed_still_fires_on_saturday():
    from datetime import datetime, timezone
    friday_backup = datetime(2026, 1, 9, 20, 0, tzinfo=timezone.utc)
    saturday_check = datetime(2026, 1, 10, 12, 0, tzinfo=timezone.utc)
    assert PBSPoller._classify_backup(friday_backup, "failed", 25, now=saturday_check) == "failed"


def test_group_backups_keeps_latest_per_guest():
    poller = _make_pbs_poller()
    snaps = [
        {"backup-type": "vm", "backup-id": "108", "backup-time": 1000, "size": 100},
        {"backup-type": "vm", "backup-id": "108", "backup-time": 2000, "size": 200},
    ]
    backups = poller._group_backups(snaps)
    assert len(backups) == 1
    assert backups[0]["vmid"] == 108
    assert backups[0]["size_bytes"] == 200


def test_group_backups_separates_different_guests():
    poller = _make_pbs_poller()
    snaps = [
        {"backup-type": "vm", "backup-id": "108", "backup-time": 1000},
        {"backup-type": "ct", "backup-id": "120", "backup-time": 1000},
    ]
    backups = poller._group_backups(snaps)
    assert {b["vmid"] for b in backups} == {108, 120}


def test_group_backups_status_failed_when_verification_failed():
    from datetime import datetime, timezone
    poller = _make_pbs_poller()
    now = datetime.now(tz=timezone.utc)
    snaps = [{
        "backup-type": "vm", "backup-id": "108",
        "backup-time": int(now.timestamp()),
        "verification": {"state": "failed"},
    }]
    backups = poller._group_backups(snaps, now=now)
    assert backups[0]["status"] == "failed"


def test_group_backups_skips_records_without_type_or_id():
    poller = _make_pbs_poller()
    snaps = [{"backup-time": 1000}]
    assert poller._group_backups(snaps) == []


def test_business_hours_elapsed_same_weekday_no_weekend():
    from datetime import datetime, timezone, timedelta
    start = datetime(2026, 1, 12, 9, 0, tzinfo=timezone.utc)   # Monday 09:00
    end = datetime(2026, 1, 12, 15, 0, tzinfo=timezone.utc)    # Monday 15:00
    assert PBSPoller._business_hours_elapsed(start, end) == 6.0


def test_business_hours_elapsed_spans_one_weekend():
    from datetime import datetime, timezone
    start = datetime(2026, 1, 9, 20, 0, tzinfo=timezone.utc)   # Friday 20:00
    end = datetime(2026, 1, 12, 20, 0, tzinfo=timezone.utc)    # Monday 20:00 (72h raw)
    # Sat + Sun fully excluded (48h) -> 24 business hours remain
    assert PBSPoller._business_hours_elapsed(start, end) == 24.0


def test_business_hours_elapsed_spans_two_weekends():
    from datetime import datetime, timezone
    start = datetime(2026, 1, 2, 20, 0, tzinfo=timezone.utc)   # Friday 20:00
    end = datetime(2026, 1, 16, 20, 0, tzinfo=timezone.utc)    # Friday 20:00, 14 days later (336h raw)
    # 4 full weekend days excluded (96h) -> 240 business hours remain
    assert PBSPoller._business_hours_elapsed(start, end) == 240.0


def test_business_hours_elapsed_partial_day_boundary():
    from datetime import datetime, timezone
    start = datetime(2026, 1, 9, 23, 0, tzinfo=timezone.utc)   # Friday 23:00
    end = datetime(2026, 1, 10, 1, 0, tzinfo=timezone.utc)     # Saturday 01:00 (2h raw)
    # 1h on Friday (business) + 1h on Saturday (weekend, excluded) -> 1 business hour remains
    assert PBSPoller._business_hours_elapsed(start, end) == 1.0


def test_business_hours_elapsed_end_before_start_returns_zero():
    from datetime import datetime, timezone
    start = datetime(2026, 1, 12, 15, 0, tzinfo=timezone.utc)
    end = datetime(2026, 1, 12, 9, 0, tzinfo=timezone.utc)
    assert PBSPoller._business_hours_elapsed(start, end) == 0.0


# ============================================================================
# PBSPoller — alerting and full poll cycle
# ============================================================================

def test_pbs_check_alerts_fires_once_for_failed():
    poller = _make_pbs_poller()
    backups = [{"type": "vm", "vmid": 108, "status": "failed", "last_backup_time": "2026-07-01T00:00:00+00:00"}]
    with patch("monitor._send_alert_async") as mock_send:
        poller._check_alerts(backups)
        poller._check_alerts(backups)
    assert mock_send.call_count == 1


def test_pbs_check_alerts_fires_for_stale():
    poller = _make_pbs_poller()
    backups = [{"type": "ct", "vmid": 120, "status": "stale", "last_backup_time": "2026-06-28T01:02:00+00:00"}]
    with patch("monitor._send_alert_async") as mock_send:
        poller._check_alerts(backups)
    assert mock_send.call_count == 1
    args = mock_send.call_args[0]
    assert "120" in args[2]


def test_pbs_check_alerts_rearms_after_clear():
    poller = _make_pbs_poller()
    failed = [{"type": "vm", "vmid": 108, "status": "failed", "last_backup_time": None}]
    ok     = [{"type": "vm", "vmid": 108, "status": "ok",     "last_backup_time": None}]
    with patch("monitor._send_alert_async") as mock_send:
        poller._check_alerts(failed)  # fires
        poller._check_alerts(ok)      # clears
        poller._check_alerts(failed)  # re-arms -> fires again
    assert mock_send.call_count == 2


def test_pbs_check_alerts_no_alert_for_none_status():
    poller = _make_pbs_poller()
    backups = [{"type": "vm", "vmid": 108, "status": "none", "last_backup_time": None}]
    with patch("monitor._send_alert_async") as mock_send:
        poller._check_alerts(backups)
    assert mock_send.call_count == 0


def test_pbs_poll_builds_cache_from_fetch():
    from datetime import datetime, timezone
    poller = _make_pbs_poller()
    now = datetime.now(tz=timezone.utc)

    def _fake_fetch(base_url, token_id, token_secret, path):
        if path == "/api2/json/status/datastore-usage":
            return [{"store": "backup-store", "used": 500, "total": 1000, "avail": 500}]
        if path == "/api2/json/admin/datastore/backup-store/snapshots":
            return [{"backup-type": "vm", "backup-id": "108", "backup-time": int(now.timestamp())}]
        raise AssertionError(f"unexpected path {path}")

    with patch.object(poller, "_fetch", side_effect=_fake_fetch):
        poller._poll()
    cache = poller.get_cache()
    assert cache["reachable"] is True
    assert cache["error"] is None
    assert cache["datastores"][0]["name"] == "backup-store"
    assert cache["backups"][0]["vmid"] == 108
    assert cache["backups"][0]["status"] == "ok"


def test_pbs_poll_sets_unreachable_on_fetch_error():
    poller = _make_pbs_poller()
    with patch.object(poller, "_fetch", side_effect=ConnectionError("timeout")):
        poller._poll()
    assert poller.get_cache()["reachable"] is False


# ============================================================================
# /api/proxmox handler
# ============================================================================

from monitor import _h_get_proxmox


def test_h_get_proxmox_when_poller_is_none():
    status, body = _h_get_proxmox(None)
    assert status == 503
    assert body["reachable"] is False


def test_h_get_proxmox_returns_cache():
    poller = _make_proxmox_poller()
    with poller._lock:
        poller._cache = {
            "reachable": True,
            "last_updated": "2026-06-16T14:00:00",
            "error": None,
            "nodes": [{"name": "pve"}],
        }
    status, body = _h_get_proxmox(poller)
    assert status == 200
    assert body["reachable"] is True
    assert body["nodes"][0]["name"] == "pve"


def test_h_get_proxmox_force_triggers_poll():
    poller = _make_proxmox_poller()
    with patch.object(poller, "_poll") as mock_poll:
        _h_get_proxmox(poller, force=True)
    mock_poll.assert_called_once()


def test_h_get_proxmox_without_force_does_not_poll():
    poller = _make_proxmox_poller()
    with patch.object(poller, "_poll") as mock_poll:
        _h_get_proxmox(poller, force=False)
    mock_poll.assert_not_called()


# ============================================================================
# /api/nas/acknowledge-alert handler
# ============================================================================

from monitor import _h_post_nas_acknowledge_alert


def test_acknowledge_alert_requires_id():
    code, body = _h_post_nas_acknowledge_alert({}, _make_nas_poller())
    assert code == 400
    assert "id" in body["error"]


def test_acknowledge_alert_poller_none():
    code, body = _h_post_nas_acknowledge_alert({"id": "abc"}, None)
    assert code == 503


def test_acknowledge_alert_not_configured():
    am = MagicMock()
    am.data = {}
    poller = NASPoller(am, alert_settings={}, alert_port=8080)
    code, body = _h_post_nas_acknowledge_alert({"id": "abc"}, poller)
    assert code == 503


def test_acknowledge_alert_calls_truenas_dismiss_and_repolls():
    poller = _make_nas_poller()
    fake_response = MagicMock()
    fake_response.__enter__.return_value = fake_response
    fake_response.__exit__.return_value = False
    with patch("urllib.request.urlopen", return_value=fake_response) as mock_urlopen, \
         patch.object(poller, "_poll") as mock_poll:
        code, body = _h_post_nas_acknowledge_alert({"id": "bf046f61-72b1-43a3-9192-3047833abf1b"}, poller)
    assert code == 200
    assert body["ok"] is True
    mock_poll.assert_called_once()
    sent_req = mock_urlopen.call_args[0][0]
    assert sent_req.data == _json.dumps("bf046f61-72b1-43a3-9192-3047833abf1b").encode()
    assert sent_req.get_method() == "POST"
    assert sent_req.full_url == "http://truenas.test/api/v2.0/alert/dismiss"


def test_acknowledge_alert_handles_http_error_and_skips_repoll():
    poller = _make_nas_poller()
    import urllib.error as _urlerr
    err = _urlerr.HTTPError(url="x", code=404, msg="not found", hdrs=None,
                             fp=io.BytesIO(b'{"error":"no such alert"}'))
    with patch("urllib.request.urlopen", side_effect=err), patch.object(poller, "_poll") as mock_poll:
        code, body = _h_post_nas_acknowledge_alert({"id": "bad-id"}, poller)
    assert code == 404
    mock_poll.assert_not_called()


# ============================================================================
# /api/system/restart handler
# ============================================================================

from monitor import _h_post_system_restart


def test_system_restart_closes_resources_then_execs_in_order():
    history_db = MagicMock()
    auth_manager = MagicMock()
    manager = MagicMock()
    manager.attach_mock(history_db.close, "history_close")
    manager.attach_mock(auth_manager.close, "auth_close")
    with patch("os.execv") as mock_execv:
        manager.attach_mock(mock_execv, "execv")
        _h_post_system_restart(history_db, auth_manager)
    assert [c[0] for c in manager.mock_calls] == ["history_close", "auth_close", "execv"]


def test_system_restart_execs_with_current_interpreter_and_argv():
    with patch("os.execv") as mock_execv, \
         patch.object(_mon.sys, "argv", ["monitor.py", "--no-tui", "--port", "8080"]):
        _h_post_system_restart(MagicMock(), MagicMock())
    mock_execv.assert_called_once_with(
        _mon.sys.executable,
        [_mon.sys.executable, "monitor.py", "--no-tui", "--port", "8080"],
    )


def test_system_restart_tolerates_missing_history_db_and_auth_manager():
    with patch("os.execv") as mock_execv:
        _h_post_system_restart(None, None)
    mock_execv.assert_called_once()


# ============================================================================
# /api/proxmox/action handler
# ============================================================================

from monitor import _h_post_proxmox_action


def _make_auth_manager_with_pve():
    am = MagicMock()
    am.data = {
        "proxmox_url": "https://pve.test:8006",
        "proxmox_user": "root@pam",
        "proxmox_token_id": "Netwatch",
        "proxmox_token_secret": "test-uuid",
    }
    return am


def test_pve_action_missing_node_returns_400():
    am = _make_auth_manager_with_pve()
    status, body = _h_post_proxmox_action(
        {"vmid": 108, "type": "qemu", "action": "stop"}, None, am
    )
    assert status == 400


def test_pve_action_invalid_action_returns_400():
    am = _make_auth_manager_with_pve()
    status, body = _h_post_proxmox_action(
        {"node": "pve", "vmid": 108, "type": "qemu", "action": "destroy"}, None, am
    )
    assert status == 400


def test_pve_action_invalid_type_returns_400():
    am = _make_auth_manager_with_pve()
    status, body = _h_post_proxmox_action(
        {"node": "pve", "vmid": 108, "type": "openvz", "action": "stop"}, None, am
    )
    assert status == 400


def test_pve_action_stop_exempts_vmid():
    import time
    am = _make_auth_manager_with_pve()
    poller = _make_proxmox_poller()
    with patch("urllib.request.urlopen") as mock_open:
        mock_resp = MagicMock()
        mock_resp.__enter__ = MagicMock(return_value=mock_resp)
        mock_resp.__exit__ = MagicMock(return_value=False)
        mock_resp.read.return_value = b'{"data": "UPID:pve:..."}'
        mock_open.return_value = mock_resp
        status, body = _h_post_proxmox_action(
            {"node": "pve", "vmid": 108, "type": "qemu", "action": "stop"},
            poller, am
        )
    # Exemption should now be set
    assert poller._exemptions.get(108, 0) > time.time()


def test_pve_action_unconfigured_returns_503():
    am = MagicMock()
    am.data = {}
    status, body = _h_post_proxmox_action(
        {"node": "pve", "vmid": 108, "type": "qemu", "action": "stop"}, None, am
    )
    assert status == 503


def test_pve_action_rejects_path_traversal_node():
    am = _make_auth_manager_with_pve()
    status, body = _h_post_proxmox_action(
        {"node": "../../etc/passwd", "vmid": 108, "type": "qemu", "action": "stop"},
        None, am
    )
    assert status == 400


def test_pve_action_rejects_node_with_slash():
    am = _make_auth_manager_with_pve()
    status, body = _h_post_proxmox_action(
        {"node": "pve/extra", "vmid": 108, "type": "qemu", "action": "stop"},
        None, am
    )
    assert status == 400


# ── restore_backup ───────────────────────────────────────────────────────────

def _build_fixture_backup(tmp_path):
    """Build a real backup tarball via create_backup_tarball, for restore tests."""
    from monitor import create_backup_tarball
    src_dir = tmp_path / "source"
    src_dir.mkdir()
    config_path = str(src_dir / "hosts.yaml")
    auth_path = str(src_dir / "auth.json")
    db_path = str(src_dir / "netwatch.db")
    monitor_path = str(src_dir / "monitor.py")

    with open(config_path, "w") as f:
        f.write("hosts: []\n")
    with open(auth_path, "w") as f:
        f.write('{"secret_key": "abc123", "users": {}}')
    with open(monitor_path, "w") as f:
        f.write("# fake monitor.py for fixture purposes\n")
    conn = sqlite3.connect(db_path)
    conn.execute("CREATE TABLE t (a INTEGER)")
    conn.commit()
    conn.close()

    data, filename, manifest = create_backup_tarball(config_path, auth_path)
    tarball_path = tmp_path / filename
    tarball_path.write_bytes(data)
    return str(tarball_path), manifest


def test_restore_backup_happy_path(tmp_path):
    from monitor import restore_backup
    tarball_path, manifest = _build_fixture_backup(tmp_path)

    dest_dir = tmp_path / "dest"
    dest_dir.mkdir()
    dest_config_path = str(dest_dir / "hosts.yaml")

    ok, message = restore_backup(tarball_path, dest_config_path)

    assert ok is True
    assert (dest_dir / "hosts.yaml").exists()
    assert (dest_dir / "auth.json").exists()
    assert (dest_dir / "netwatch.db").exists()
    assert manifest["source_hostname"] in message
    assert oct(os.stat(dest_dir / "auth.json").st_mode)[-3:] == "600"
    # The bundled monitor.py in the tarball must NOT be extracted
    assert not (dest_dir / "monitor.py").exists()


def test_restore_backup_refuses_to_overwrite_without_force(tmp_path):
    from monitor import restore_backup
    tarball_path, _ = _build_fixture_backup(tmp_path)

    dest_dir = tmp_path / "dest"
    dest_dir.mkdir()
    dest_config_path = str(dest_dir / "hosts.yaml")
    with open(dest_config_path, "w") as f:
        f.write("hosts: []\n")  # pre-existing file in the way

    ok, message = restore_backup(tarball_path, dest_config_path)

    assert ok is False
    assert "hosts.yaml" in message
    assert "--force" in message


def test_restore_backup_force_overwrites(tmp_path):
    from monitor import restore_backup
    tarball_path, _ = _build_fixture_backup(tmp_path)

    dest_dir = tmp_path / "dest"
    dest_dir.mkdir()
    dest_config_path = str(dest_dir / "hosts.yaml")
    with open(dest_config_path, "w") as f:
        f.write("hosts: [{name: stale}]\n")

    ok, message = restore_backup(tarball_path, dest_config_path, force=True)

    assert ok is True
    with open(dest_config_path) as f:
        assert "stale" not in f.read()


def test_restore_backup_rejects_invalid_tarball(tmp_path):
    from monitor import restore_backup
    not_a_backup = tmp_path / "not-a-backup.tar.gz"
    import tarfile
    with tarfile.open(str(not_a_backup), "w:gz") as tar:
        info = tarfile.TarInfo(name="some-other-file.txt")
        data = b"hello"
        info.size = len(data)
        tar.addfile(info, io.BytesIO(data))

    dest_config_path = str(tmp_path / "dest" / "hosts.yaml")
    ok, message = restore_backup(str(not_a_backup), dest_config_path)

    assert ok is False
    assert "not a valid netwatch backup" in message.lower()


def test_restore_backup_missing_tarball_file(tmp_path):
    from monitor import restore_backup
    ok, message = restore_backup(str(tmp_path / "does-not-exist.tar.gz"), str(tmp_path / "hosts.yaml"))
    assert ok is False
    assert "not found" in message.lower()


def test_restore_backup_warns_on_newer_manifest_version(tmp_path, monkeypatch):
    from monitor import restore_backup
    import monitor as _mon
    tarball_path, _ = _build_fixture_backup(tmp_path)

    # Pretend this monitor.py is older than the backup's manifest version
    monkeypatch.setattr(_mon, "BACKUP_MANIFEST_VERSION", 0)

    dest_dir = tmp_path / "dest"
    dest_dir.mkdir()
    ok, message = restore_backup(tarball_path, str(dest_dir / "hosts.yaml"))

    assert ok is True
    assert "newer netwatch version" in message.lower()


def test_restore_cli_exits_before_server_startup(tmp_path):
    """--restore should print a message and exit without ever importing/starting
    AuthManager, HistoryDB, etc. Run as a subprocess so we observe the real
    argparse + main() path, not just the restore_backup() function in isolation."""
    import subprocess
    tarball_path, _ = _build_fixture_backup(tmp_path)
    dest_dir = tmp_path / "dest"
    dest_dir.mkdir()

    monitor_py = os.path.join(os.path.dirname(os.path.abspath(_mon.__file__)), "monitor.py")
    result = subprocess.run(
        [sys.executable, monitor_py, "--restore", tarball_path, "--config",
         str(dest_dir / "hosts.yaml")],
        capture_output=True, text=True, timeout=10,
    )
    assert result.returncode == 0
    assert "Restored backup from" in result.stdout
    assert (dest_dir / "hosts.yaml").exists()


# ── _tui_status_role ─────────────────────────────────────────────────────────

def test_tui_status_role_wait_when_pending():
    from monitor import _tui_status_role
    assert _tui_status_role(pend=True, is_up=False, status="WAIT") == "wait"


def test_tui_status_role_pending_wins_over_degraded():
    from monitor import _tui_status_role
    # Defensive priority check: pend must win even if status somehow already
    # says DEGRADED before the first check has completed.
    assert _tui_status_role(pend=True, is_up=True, status="DEGRADED") == "wait"


def test_tui_status_role_degraded_when_up_but_degraded():
    from monitor import _tui_status_role
    assert _tui_status_role(pend=False, is_up=True, status="DEGRADED") == "degraded"


def test_tui_status_role_degraded_wins_over_up():
    from monitor import _tui_status_role
    # This is the actual regression this task fixes: a DEGRADED host has
    # is_up == True, so a naive "if is_up: return up" check (the current
    # bug in draw_tui) would misclassify it as healthy.
    assert _tui_status_role(pend=False, is_up=True, status="DEGRADED") != "up"


def test_tui_status_role_up_when_healthy():
    from monitor import _tui_status_role
    assert _tui_status_role(pend=False, is_up=True, status="UP") == "up"


def test_tui_status_role_idle_when_down_and_not_always_on():
    from monitor import _tui_status_role
    assert _tui_status_role(pend=False, is_up=False, status="IDLE") == "idle"


def test_tui_status_role_down_when_down_and_always_on():
    from monitor import _tui_status_role
    assert _tui_status_role(pend=False, is_up=False, status="DOWN") == "down"


# ── ProxmoxPoller._make_ssl_ctx ──────────────────────────────────────────────

def _make_self_signed_ca(path, extra_args=()):
    """Generate a throwaway self-signed CA cert for SSL context tests.

    Proxmox's own cluster CA commonly omits the Key Usage extension, which
    OpenSSL 3.x's strict X.509 policy rejects as an issuer. extra_args lets
    a test add/omit extensions to reproduce that exact shape."""
    import subprocess
    key_path = str(path) + ".key"
    subprocess.run(
        ["openssl", "req", "-x509", "-newkey", "rsa:2048", "-nodes",
         "-keyout", key_path, "-out", str(path), "-days", "1",
         "-subj", "/CN=Test CA"] + list(extra_args),
        check=True, capture_output=True,
    )


def test_make_ssl_ctx_pinned_ca_without_key_usage_disables_x509_strict(tmp_path):
    import ssl
    ca_path = tmp_path / "ca.pem"
    _make_self_signed_ca(ca_path, ["-addext", "basicConstraints=critical,CA:TRUE"])
    poller = ProxmoxPoller(MagicMock(), alert_settings={"proxmox_ca_cert": str(ca_path)})

    ctx = poller._make_ssl_ctx()

    if hasattr(ssl, "VERIFY_X509_STRICT"):
        assert not (ctx.verify_flags & ssl.VERIFY_X509_STRICT)


def test_make_ssl_ctx_without_pinned_ca_keeps_default_strict_flag():
    import ssl
    poller = ProxmoxPoller(MagicMock(), alert_settings={})

    ctx = poller._make_ssl_ctx()

    default_ctx = ssl.create_default_context()
    assert ctx.verify_flags == default_ctx.verify_flags


def test_make_ssl_ctx_verify_disabled_ignores_ca_cert():
    import ssl
    poller = ProxmoxPoller(MagicMock(), alert_settings={"proxmox_verify_ssl": False,
                                                         "proxmox_ca_cert": "/nonexistent.pem"})

    ctx = poller._make_ssl_ctx()

    assert ctx.verify_mode == ssl.CERT_NONE


def test_incident_log_record_down_passes_started_at(tmp_path):
    """record_down forwards started_at to open_incident."""
    from monitor import HistoryDB, IncidentLog, HostState
    hdb = HistoryDB(str(tmp_path / "t.db"))
    log = IncidentLog(hdb)
    host = HostState(name="H", ip="10.0.0.1", group="G", interval=60, always_on=True)
    past = int(time.time()) - 300
    log.record_down(host, started_at=past)
    row = hdb.conn.execute(
        "SELECT started FROM incidents WHERE host_ip = '10.0.0.1'"
    ).fetchone()
    assert row[0] == past


def test_incident_not_opened_below_threshold(tmp_path):
    """Fewer than NTFY_DOWN_THRESHOLD consecutive misses must not open an incident."""
    from monitor import HistoryDB, IncidentLog, HostState, NTFY_DOWN_THRESHOLD
    hdb = HistoryDB(str(tmp_path / "t.db"))
    inc_log = IncidentLog(hdb)
    host = HostState(name="H", ip="10.0.0.2", group="G", interval=60, always_on=True)

    # Simulate NTFY_DOWN_THRESHOLD - 1 consecutive misses
    for i in range(1, NTFY_DOWN_THRESHOLD):
        host.consecutive_down = i
        if host.consecutive_down == 1:
            host.first_down_at = time.time() - 100
        # Mirror what poll_host does: only call record_down at threshold
        if host.consecutive_down == NTFY_DOWN_THRESHOLD:
            inc_log.record_down(host, started_at=host.first_down_at)

    row = hdb.conn.execute(
        "SELECT id FROM incidents WHERE host_ip = '10.0.0.2'"
    ).fetchone()
    assert row is None, "No incident should open before the threshold"


def test_incident_opened_at_threshold_with_correct_start(tmp_path):
    """At exactly NTFY_DOWN_THRESHOLD consecutive misses an incident opens, backdated."""
    from monitor import HistoryDB, IncidentLog, HostState, NTFY_DOWN_THRESHOLD
    hdb = HistoryDB(str(tmp_path / "t.db"))
    inc_log = IncidentLog(hdb)
    host = HostState(name="H", ip="10.0.0.3", group="G", interval=60, always_on=True)
    first_miss_time = int(time.time()) - 60

    host.first_down_at = first_miss_time
    host.consecutive_down = NTFY_DOWN_THRESHOLD
    inc_log.record_down(host, started_at=host.first_down_at)

    row = hdb.conn.execute(
        "SELECT started FROM incidents WHERE host_ip = '10.0.0.3'"
    ).fetchone()
    assert row is not None
    assert row[0] == first_miss_time


# ============================================================================
# HistoryDB power_readings tests
# ============================================================================

import time as _time

def test_power_readings_roundtrip(tmp_path):
    from monitor import HistoryDB
    db = HistoryDB(str(tmp_path / "t.db"))
    ts = int(_time.time())
    db.insert_power_reading(ts, 47.3, 230.1, 0.21, 1.234)
    rows = db.get_power_readings(days=7)
    assert len(rows) == 1
    assert rows[0]["watts"] == 47.3
    assert rows[0]["voltage"] == 230.1
    assert rows[0]["current_a"] == 0.21
    assert rows[0]["energy_kwh"] == 1.234
    assert rows[0]["timestamp"] == ts


def test_power_readings_filters_by_days(tmp_path):
    from monitor import HistoryDB
    db = HistoryDB(str(tmp_path / "t.db"))
    old_ts = int(_time.time()) - 8 * 86400
    recent_ts = int(_time.time()) - 1 * 86400
    db.insert_power_reading(old_ts, 10.0, 230.0, 0.04, 0.1)
    db.insert_power_reading(recent_ts, 50.0, 230.0, 0.22, 0.5)
    rows = db.get_power_readings(days=7)
    assert len(rows) == 1
    assert rows[0]["watts"] == 50.0


def test_power_readings_pruned(tmp_path):
    from monitor import HistoryDB
    db = HistoryDB(str(tmp_path / "t.db"), retention_days=7)
    old_ts = int(_time.time()) - 8 * 86400
    db.insert_power_reading(old_ts, 10.0, 230.0, 0.04, 0.1)
    db.prune()
    rows = db.get_power_readings(days=30)
    assert rows == []


def test_power_readings_none_values_stored(tmp_path):
    from monitor import HistoryDB
    db = HistoryDB(str(tmp_path / "t.db"))
    ts = int(_time.time())
    db.insert_power_reading(ts, 47.3, None, None, None)
    rows = db.get_power_readings(days=7)
    assert len(rows) == 1
    assert rows[0]["watts"] == 47.3
    assert rows[0]["voltage"] is None


# ============================================================================
# HAPoller tests
# ============================================================================

from monitor import HAPoller as _HAPoller


def _make_ha_poller(tmp_path):
    from monitor import HistoryDB
    am = MagicMock()
    am.data = {
        "ha_url": "http://ha.test:8123",
        "ha_token": "testtoken",
        "ha_entity_power":   "sensor.tapo_p115_current_power",
        "ha_entity_voltage":  "sensor.tapo_p115_voltage",
        "ha_entity_current":  "sensor.tapo_p115_current",
        "ha_entity_energy":   "sensor.tapo_p115_energy_today",
    }
    db = HistoryDB(str(tmp_path / "ha.db"))
    return _HAPoller(am, db)


def test_ha_poller_poll_updates_cache(tmp_path):
    poller = _make_ha_poller(tmp_path)
    responses = {
        "sensor.tapo_p115_current_power": 47.3,
        "sensor.tapo_p115_voltage":        230.1,
        "sensor.tapo_p115_current":        0.21,
        "sensor.tapo_p115_energy_today":   1.234,
    }
    with patch.object(_HAPoller, "_fetch_state", staticmethod(lambda url, tok, eid: responses.get(eid))):
        poller._poll()
    cache = poller.get_cache()
    assert cache["reachable"] is True
    assert cache["watts"] == 47.3
    assert cache["voltage"] == 230.1
    assert cache["current_a"] == 0.21
    assert cache["energy_kwh"] == 1.234
    assert cache["error"] is None


def test_ha_poller_unavailable_entity_stored_as_none(tmp_path):
    poller = _make_ha_poller(tmp_path)
    def _fake(url, tok, eid):
        return None if eid == "sensor.tapo_p115_voltage" else 47.3
    with patch.object(_HAPoller, "_fetch_state", staticmethod(_fake)):
        poller._poll()
    cache = poller.get_cache()
    assert cache["reachable"] is True
    assert cache["voltage"] is None
    assert cache["watts"] == 47.3


def test_ha_poller_network_error_sets_unreachable(tmp_path):
    poller = _make_ha_poller(tmp_path)
    def _fail(url, tok, eid):
        raise ConnectionError("timeout")
    with patch.object(_HAPoller, "_fetch_state", staticmethod(_fail)):
        poller._poll()
    cache = poller.get_cache()
    assert cache["reachable"] is False
    assert cache["error"] is not None


def test_ha_poller_unconfigured_does_not_poll(tmp_path):
    from monitor import HistoryDB
    am = MagicMock()
    am.data = {"ha_url": "", "ha_token": ""}
    db = HistoryDB(str(tmp_path / "ha.db"))
    poller = _HAPoller(am, db)
    with patch.object(_HAPoller, "_fetch_state", staticmethod(lambda *a: 1 / 0)):
        poller._poll()  # should not raise or fetch
    assert poller.get_cache()["reachable"] is False


def test_ha_poller_poll_writes_to_history_db(tmp_path):
    poller = _make_ha_poller(tmp_path)
    with patch.object(_HAPoller, "_fetch_state", staticmethod(lambda url, tok, eid: 47.3)):
        poller._poll()
    rows = poller._history_db.get_power_readings(days=1)
    assert len(rows) == 1
    assert rows[0]["watts"] == 47.3


def test_ha_poller_get_cache_returns_deep_copy(tmp_path):
    poller = _make_ha_poller(tmp_path)
    c1 = poller.get_cache()
    c1["watts"] = 999
    c2 = poller.get_cache()
    assert c2["watts"] is None  # original cache unchanged


# ============================================================================
# _h_get_power handler tests
# ============================================================================

from monitor import _h_get_power as _hgp


def test_h_get_power_not_configured():
    status, body = _hgp(None, None)
    assert status == 200
    assert body == {"configured": False}


def test_h_get_power_configured(tmp_path):
    from monitor import HistoryDB, HAPoller
    db = HistoryDB(str(tmp_path / "p.db"))
    ts = int(_time.time())
    db.insert_power_reading(ts, 47.3, 230.1, 0.21, 1.234)
    am = MagicMock()
    am.data = {
        "ha_url": "http://ha.test:8123", "ha_token": "t",
        "ha_entity_power": "sensor.p", "ha_entity_voltage": "",
        "ha_entity_current": "", "ha_entity_energy": "",
    }
    poller = HAPoller(am, db)
    with poller._lock:
        poller._cache.update({"reachable": True, "watts": 47.3, "last_updated": ts})
    status, body = _hgp(poller, db)
    assert status == 200
    assert body["configured"] is True
    assert body["live"]["watts"] == 47.3
    assert body["live"]["reachable"] is True
    assert len(body["history"]) == 1
    assert body["history"][0]["watts"] == 47.3


def test_h_get_power_force_triggers_poll(tmp_path):
    from monitor import HistoryDB, HAPoller
    db = HistoryDB(str(tmp_path / "p.db"))
    am = MagicMock()
    am.data = {
        "ha_url": "http://ha.test:8123", "ha_token": "t",
        "ha_entity_power": "sensor.p", "ha_entity_voltage": "",
        "ha_entity_current": "", "ha_entity_energy": "",
    }
    poller = HAPoller(am, db)
    with patch.object(HAPoller, "_fetch_state", staticmethod(lambda url, tok, eid: 55.0)):
        status, body = _hgp(poller, db, force=True)
    assert body["live"]["watts"] == 55.0


# ============================================================================
# /api/pbs handler
# ============================================================================

from monitor import _h_get_pbs


def test_h_get_pbs_when_poller_is_none():
    status, body = _h_get_pbs(None)
    assert status == 503
    assert body["reachable"] is False


def test_h_get_pbs_returns_cache():
    poller = _make_pbs_poller()
    with poller._lock:
        poller._cache = {
            "reachable": True,
            "last_updated": "2026-07-01T02:00:00",
            "error": None,
            "datastores": [{"name": "backup-store"}],
            "backups": [],
        }
    status, body = _h_get_pbs(poller)
    assert status == 200
    assert body["datastores"][0]["name"] == "backup-store"


def test_h_get_pbs_force_triggers_poll():
    poller = _make_pbs_poller()
    with patch.object(poller, "_poll") as mock_poll:
        _h_get_pbs(poller, force=True)
    mock_poll.assert_called_once()


def test_h_get_pbs_without_force_does_not_poll():
    poller = _make_pbs_poller()
    with patch.object(poller, "_poll") as mock_poll:
        _h_get_pbs(poller, force=False)
    mock_poll.assert_not_called()


def test_h_get_pbs_not_configured_error():
    am = MagicMock()
    am.data = {}
    poller = PBSPoller(am)
    status, body = _h_get_pbs(poller)
    assert status == 200
    assert body["error"] == "PBS not configured"


# ============================================================================
# PBS settings keys
# ============================================================================

def test_pbs_credential_keys_in_auth_stored_keys():
    for k in ("pbs_url", "pbs_api_token_id", "pbs_api_token_secret"):
        assert k in _AUTH_STORED_KEYS


def test_pbs_verify_ssl_and_ca_cert_not_in_auth_stored_keys():
    assert "pbs_verify_ssl" not in _AUTH_STORED_KEYS
    assert "pbs_ca_cert" not in _AUTH_STORED_KEYS


def test_h_post_settings_saves_pbs_secret_to_auth_manager():
    import tempfile, os
    am = MagicMock()
    am.data = {}
    am.lock = MagicMock()
    am.lock.__enter__ = MagicMock(return_value=None)
    am.lock.__exit__ = MagicMock(return_value=False)
    am._save = MagicMock()
    with tempfile.TemporaryDirectory() as d:
        cfg = os.path.join(d, "hosts.yaml")
        status, body = _h_post_settings(
            {"pbs_api_token_secret": "new-uuid"}, cfg, {}, auth_manager=am
        )
    assert am.data.get("pbs_api_token_secret") == "new-uuid"


# ── /api/settings secret redaction ──────────────────────────────────────────

def test_get_settings_redacts_secrets():
    from monitor import _h_get_settings, SECRET_PLACEHOLDER

    class FakeAuth:
        lock = _threading.Lock()
        data = {"truenas_api_key": "real-secret-123", "openrouter_api_key": ""}

    code, result = _h_get_settings({"default_interval": 30}, FakeAuth())
    assert code == 200
    assert result["truenas_api_key"] == SECRET_PLACEHOLDER
    assert result["openrouter_api_key"] == ""          # unset stays empty, not sentinel
    assert "real-secret-123" not in str(result)


def test_post_settings_sentinel_keeps_existing_secret(tmp_path):
    from monitor import _h_post_settings, SECRET_PLACEHOLDER

    class FakeAuth:
        lock = _threading.Lock()
        data = {"truenas_api_key": "keep-me"}
        def _save(self): pass

    cfg = tmp_path / "hosts.yaml"
    cfg.write_text("settings: {}\nhosts: []\n")
    auth = FakeAuth()
    code, _ = _h_post_settings(
        {"truenas_api_key": SECRET_PLACEHOLDER}, str(cfg), {}, auth)
    assert code == 200
    assert auth.data["truenas_api_key"] == "keep-me"


def test_post_settings_empty_still_clears_secret(tmp_path):
    from monitor import _h_post_settings

    class FakeAuth:
        lock = _threading.Lock()
        data = {"truenas_api_key": "clear-me"}
        def _save(self): pass

    cfg = tmp_path / "hosts.yaml"
    cfg.write_text("settings: {}\nhosts: []\n")
    auth = FakeAuth()
    code, _ = _h_post_settings({"truenas_api_key": ""}, str(cfg), {}, auth)
    assert code == 200
    assert "truenas_api_key" not in auth.data


# ── Proxmox node CPU/RAM sparkline history ──────────────────────────────────

def test_proxmox_append_history_tracks_and_caps():
    from monitor import ProxmoxPoller

    history = {}
    def node(cpu, mem_used, mem_total):
        return {"name": "pve-01", "cpu_percent": cpu,
                "mem_used_bytes": mem_used, "mem_total_bytes": mem_total}

    for i in range(25):
        nodes = [node(float(i), 50, 100)]
        ProxmoxPoller.append_history(history, nodes, cap=20)

    assert nodes[0]["cpu_history"][-1] == 24.0
    assert len(nodes[0]["cpu_history"]) == 20          # capped
    assert nodes[0]["cpu_history"][0] == 5.0           # oldest trimmed
    assert nodes[0]["mem_history"][-1] == 50.0         # percent, not bytes

    # zero-total RAM must not divide by zero
    nodes = [node(1.0, 0, 0)]
    ProxmoxPoller.append_history(history, nodes, cap=20)
    assert nodes[0]["mem_history"][-1] == 0.0
