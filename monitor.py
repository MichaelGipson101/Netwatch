#!/usr/bin/env python3
"""
netwatch - Homelab ping monitor with btop-style TUI, web dashboard,
and web-based hosts.yaml editor. (Version: see VERSION below.)

Usage:
    python monitor.py                  # TUI + web server
    python monitor.py --no-tui         # Headless + web server (for systemd)
    python monitor.py --no-web         # TUI only, no web
    python monitor.py --port 8080      # Custom port

Config: hosts.yaml in the same directory.
Dashboard: http://<pi-ip>:8080
"""

import os, time, json, shutil, subprocess, threading, curses, yaml, argparse, logging
from datetime import datetime, timezone
from collections import deque
from dataclasses import dataclass, field
from typing import Optional
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer

BRAND   = "NETWATCH"
VERSION = "3.39"


def _column_exists(conn: "sqlite3.Connection", table: str, column: str) -> bool:
    """Return True if `column` exists in `table`. Both must be code-controlled identifiers."""
    rows = conn.execute(f"PRAGMA table_info({table})").fetchall()
    return any(row[1] == column for row in rows)


# ============================================================================
# Host state
# ============================================================================

@dataclass
class HostState:
    name: str
    ip: str
    group: str
    interval: int
    always_on: bool = True
    alert: bool = True
    specs: dict = field(default_factory=dict)
    notes: str = ""
    links: dict = field(default_factory=dict)
    services: list = field(default_factory=list)
    strict: bool = False
    service_results: dict = field(default_factory=dict)
    history: deque = field(default_factory=lambda: deque(maxlen=100))
    last_latency_ms: Optional[float] = None
    last_checked: Optional[datetime] = None
    last_seen_up: Optional[datetime] = None
    consecutive_down: int = 0
    lock: threading.Lock = field(default_factory=threading.Lock)
    thread: Optional[threading.Thread] = None
    stop_event: Optional[threading.Event] = None

    @property
    def is_up(self):
        return bool(self.history) and self.history[-1]

    @property
    def uptime_pct(self):
        if not self.history:
            return None
        return (sum(self.history) / len(self.history)) * 100

    @property
    def status_str(self):
        if not self.last_checked:
            return "WAIT"
        if self.is_up:
            # Ping is fine. If strict mode is on AND any configured service is
            # currently failing, surface DEGRADED. Otherwise UP.
            if self.strict and self.services:
                for svc in self.services:
                    port = svc.get("port")
                    if port is None:
                        continue
                    res = self.service_results.get(port)
                    # Only count a service as failing if we have a result and
                    # it's a failure. While we're still waiting on the first
                    # check, we treat the service as not-yet-known.
                    if res is not None and not res.get("ok", True):
                        return "DEGRADED"
            return "UP"
        return "DOWN" if self.always_on else "IDLE"

    @property
    def latency_str(self):
        if not self.is_up or self.last_latency_ms is None:
            return "- ms"
        return f"{self.last_latency_ms:.1f} ms"

    @property
    def uptime_str(self):
        if self.uptime_pct is None:
            return "-%"
        return f"{self.uptime_pct:.1f}%"

    @property
    def checked_str(self):
        if not self.last_checked:
            return "-"
        return self.last_checked.strftime("%H:%M:%S")

    def spark_str(self, width=20):
        hist = list(self.history)[-width:]
        while len(hist) < width:
            hist.insert(0, None)
        result = []
        for v in hist:
            if v is None:
                result.append(" ")
            elif v:
                result.append("\u2588")
            else:
                result.append("\u2581")
        return "".join(result)

    def to_dict(self):
        with self.lock:
            last_seen_secs = None
            if self.last_seen_up:
                last_seen_secs = int((datetime.now() - self.last_seen_up).total_seconds())
            primary_url = (self.links or {}).get("primary", "").strip()
            if not primary_url:
                primary_url = f"http://{self.ip}"
            extra_links = []
            for entry in (self.links or {}).get("extras", []) or []:
                if isinstance(entry, dict) and entry.get("url") and entry.get("name"):
                    extra_links.append({"name": str(entry["name"]), "url": str(entry["url"])})
            services_out = []
            for svc in (self.services or []):
                port = svc.get("port")
                if port is None:
                    continue
                res = self.service_results.get(port, {})
                services_out.append({
                    "port":     port,
                    "name":     svc.get("name", f"port {port}"),
                    "ok":       res.get("ok"),
                    "error":    res.get("error"),
                    "checked":  res.get("checked"),
                })
            return {
                "name":         self.name,
                "ip":           self.ip,
                "group":        self.group,
                "interval":     self.interval,
                "always_on":    self.always_on,
                "specs":        dict(self.specs) if self.specs else {},
                "notes":        self.notes,
                "links":        {"primary": primary_url, "extras": extra_links},
                "services":     services_out,
                "strict":       self.strict,
                "is_pi":        _is_local_ip(self.ip),
                "status":       self.status_str,
                "is_up":        self.is_up,
                "latency_ms":   round(self.last_latency_ms, 2) if self.last_latency_ms else None,
                "uptime_pct":   round(self.uptime_pct, 1) if self.uptime_pct is not None else None,
                "last_checked": self.checked_str,
                "last_seen_up_seconds": last_seen_secs,
                "history":      list(self.history)[-50:],
                "consecutive_down": self.consecutive_down,
            }


# ============================================================================
# Ping logic
# ============================================================================

def check_tcp_service(ip, port, timeout=2):
    """Attempt a TCP connect to ip:port. Returns (ok, error_str_or_None).

    A successful connect is enough - we close the socket immediately. We
    do not send any data, since some services (SSH, IMAP, etc.) don't like
    being connected-to without a proper handshake.
    """
    import socket
    try:
        port = int(port)
    except (ValueError, TypeError):
        return False, "invalid port"
    if port < 1 or port > 65535:
        return False, "port out of range"
    sock = None
    try:
        sock = socket.create_connection((ip, port), timeout=timeout)
        return True, None
    except socket.timeout:
        return False, "timeout"
    except ConnectionRefusedError:
        return False, "refused"
    except OSError as e:
        return False, str(e.strerror or e)
    finally:
        if sock is not None:
            try: sock.close()
            except Exception: pass


def ping_host(ip, timeout):
    try:
        result = subprocess.run(
            ["ping", "-c", "1", "-W", str(timeout), "--", ip],
            stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=timeout + 1
        )
        if result.returncode == 0:
            output = result.stdout.decode()
            for part in output.split():
                if part.startswith("time="):
                    try:
                        return True, float(part.split("=")[1])
                    except ValueError:
                        pass
            return True, None
        return False, None
    except (subprocess.TimeoutExpired, FileNotFoundError):
        return False, None


def _should_log_transition(was_up, is_up):
    """INFO-log only state changes (and the first ping). Steady-state pings
    go to DEBUG so monitor.log doesn't grow ~12MB/day."""
    return was_up is None or bool(is_up) != bool(was_up)


def poll_host(host, timeout, global_stop, incident_log=None, history_db=None, config_path_for_arp=None, alert_settings=None, alert_port=8080):
    while not global_stop.is_set() and not host.stop_event.is_set():
        was_up = host.is_up if host.history else None
        is_up, latency = ping_host(host.ip, timeout)
        with host.lock:
            host.history.append(is_up)
            host.last_latency_ms = latency if is_up else None
            host.last_checked = datetime.now()
            if is_up:
                host.last_seen_up = datetime.now()
                host.consecutive_down = 0
            else:
                host.consecutive_down += 1

        # Persist to DB if available
        if history_db is not None:
            try:
                history_db.record_ping(host.ip, is_up, latency)
            except Exception as e:
                logging.warning(f"HistoryDB record_ping failed: {e}")

        # ARP-based MAC auto-detection (only on successful pings, since
        # only successful pings populate the kernel's ARP cache fresh)
        if is_up and host.ip not in _ARP_SAVED_THIS_SESSION:
            detected = _detect_mac_for_ip(host.ip)
            if detected:
                with host.lock:
                    configured = (host.specs or {}).get("mac", "")
                if not configured:
                    # No MAC configured - save the detected one
                    if _save_detected_mac(config_path_for_arp, host.ip, detected):
                        with host.lock:
                            if not isinstance(host.specs, dict):
                                host.specs = {}
                            host.specs["mac"] = detected
                            host.specs["mac_auto"] = True
                        _ARP_SAVED_THIS_SESSION.add(host.ip)
                elif _normalise_mac(configured) != _normalise_mac(detected):
                    # Mismatch - log a warning but keep the user's value.
                    # We add to the session set so we don't spam this every
                    # ping cycle.
                    logging.warning(
                        f"ARP detect: configured MAC {configured} for {host.ip} "
                        f"({host.name}) differs from network-detected MAC {detected}. "
                        f"Keeping configured value. If the host's NIC was changed, "
                        f"clear the MAC in the editor and let it re-detect."
                    )
                    _ARP_SAVED_THIS_SESSION.add(host.ip)
                else:
                    # Match - no need to keep checking this session
                    _ARP_SAVED_THIS_SESSION.add(host.ip)

        # TCP service checks (only if ping succeeded - no point checking ports
        # if the host itself is offline).
        if host.services:
            now_str = datetime.now().strftime("%H:%M:%S")
            for svc in host.services:
                port = svc.get("port")
                if port is None:
                    continue
                if is_up:
                    ok, err = check_tcp_service(host.ip, port, timeout=timeout)
                else:
                    ok, err = False, "host offline"
                with host.lock:
                    host.service_results[port] = {
                        "ok": ok, "error": err, "checked": now_str,
                    }

        # Track incidents (only for always_on hosts)
        if incident_log is not None and host.always_on:
            if not is_up and (was_up or was_up is None):
                # Down transition (or first ping that's down)
                if not is_up:
                    incident_log.record_down(host)
            elif is_up and was_up is False:
                # Capture alert status BEFORE record_up closes the incident
                was_alerted = None
                if history_db is not None:
                    was_alerted = history_db.get_open_incident_alert_status(host.ip)
                # Recovery
                incident_log.record_up(host)
                # Recovery alert: only if down alert was actually delivered
                if (alert_settings and alert_settings.get("ntfy_topic")
                        and host.alert and was_alerted is True):
                    base = _get_dashboard_url(alert_settings, alert_port)
                    click_url = f"{base}/?host={host.ip}" if base else None
                    title = f"Recovered: {host.name}"
                    msg = (f"{host.name} ({host.ip}) is back online.\n"
                           f"Group: {host.group}")
                    _send_alert_async(alert_settings, title, msg,
                                      priority="default",
                                      tags="white_check_mark,green_circle",
                                      click_url=click_url)

        # Down alert: fire when consecutive_down hits threshold
        if (incident_log is not None and host.always_on and not is_up
                and alert_settings and alert_settings.get("ntfy_topic")
                and host.alert and history_db is not None):
            with host.lock:
                cd = host.consecutive_down
            if cd == NTFY_DOWN_THRESHOLD:
                already = history_db.get_open_incident_alert_status(host.ip)
                if not already:
                    base = _get_dashboard_url(alert_settings, alert_port)
                    click_url = f"{base}/?host={host.ip}" if base else None
                    title = f"Host down: {host.name}"
                    msg = (f"{host.name} ({host.ip}) failed {cd} consecutive pings.\n"
                           f"Group: {host.group}")
                    host_ip_for_cb = host.ip
                    _send_alert_async(
                        alert_settings, title, msg,
                        priority="high",
                        tags="warning,red_circle",
                        click_url=click_url,
                        on_success=lambda: history_db.mark_incident_alerted(host_ip_for_cb)
                    )

        line = (
            f"{'UP  ' if is_up else 'DOWN'} | {host.name:<20} | {host.ip:<16} | "
            f"{f'{latency:.1f}ms' if latency else '-':>8}"
        )
        if _should_log_transition(was_up, is_up):
            logging.info(line)
        else:
            logging.debug(line)
        elapsed = 0
        while elapsed < host.interval and not global_stop.is_set() and not host.stop_event.is_set():
            time.sleep(0.5)
            elapsed += 0.5


# ============================================================================
# Host manager: thread-safe add/remove, hot-reload
# ============================================================================

class HostManager:
    def __init__(self, config_path, ping_timeout, history_window, global_stop, incident_log=None, history_db=None, alert_settings=None, alert_port=8080):
        self.config_path = config_path
        self.ping_timeout = ping_timeout
        self.history_window = history_window
        self.global_stop = global_stop
        self.incident_log = incident_log
        self.history_db = history_db
        self.alert_settings = alert_settings or {}
        self.alert_port = alert_port
        self.hosts = []
        self.lock = threading.Lock()

    def _spawn(self, name, ip, group, interval, always_on=True, specs=None, notes="", links=None, services=None, strict=False, alert=True):
        history_window = self.history_window
        host = HostState(
            name=name, ip=ip, group=group, interval=interval,
            always_on=always_on,
            alert=bool(alert),
            specs=specs or {},
            notes=notes or "",
            links=links or {},
            services=services or [],
            strict=bool(strict),
            history=deque(maxlen=history_window),
            stop_event=threading.Event(),
        )
        # Restore history from DB if available
        if self.history_db:
            try:
                recent = self.history_db.recent_pings(ip, limit=history_window)
                for is_up, latency_ms in recent:
                    host.history.append(is_up)
                # Also restore last_latency_ms / last_seen_up from latest ping
                latest = self.history_db.latest_ping(ip)
                if latest is not None:
                    last_up, last_lat, last_ts = latest
                    if last_up:
                        host.last_seen_up = datetime.fromtimestamp(last_ts)
                        host.last_latency_ms = last_lat
            except Exception as e:
                logging.warning(f"HistoryDB restore failed for {ip}: {e}")
        host.thread = threading.Thread(
            target=poll_host,
            args=(host, self.ping_timeout, self.global_stop, self.incident_log, self.history_db, self.config_path, self.alert_settings, self.alert_port),
            daemon=True, name=f"ping-{name}",
        )
        host.thread.start()
        return host

    def load_initial(self, raw_hosts, default_interval):
        with self.lock:
            config_ips = set()
            for h in raw_hosts:
                config_ips.add(h["ip"])
                self.hosts.append(self._spawn(
                    h["name"], h["ip"], h.get("group", "General"),
                    h.get("interval", default_interval),
                    bool(h.get("always_on", True)),
                    h.get("specs") if isinstance(h.get("specs"), dict) else None,
                    h.get("notes", "") if isinstance(h.get("notes"), str) else "",
                    h.get("links") if isinstance(h.get("links"), dict) else None,
                    h.get("services") if isinstance(h.get("services"), list) else None,
                    bool(h.get("strict", False)),
                    bool(h.get("alert", True))
                ))
            # Reconcile orphans: any ongoing incident for an IP not in the
            # current config is stale (host was removed/renamed while netwatch
            # was offline). Close them so the Events tab is honest.
            if self.history_db and self.incident_log:
                try:
                    orphan_count = 0
                    for inc in self.history_db.list_incidents(limit=500):
                        if inc.get("ongoing") and inc.get("host_ip") not in config_ips:
                            self.incident_log._close_orphaned_incident(inc["host_ip"])
                            orphan_count += 1
                    if orphan_count:
                        logging.info(f"Reconciled {orphan_count} orphaned ongoing incident(s) at startup")
                except Exception as e:
                    logging.warning(f"Orphan reconciliation failed: {e}")

    def list_hosts(self):
        with self.lock:
            return list(self.hosts)

    def reload_from_config(self, new_hosts_config, default_interval):
        with self.lock:
            current_by_ip = {h.ip: h for h in self.hosts}
            new_ips = set()
            rebuilt = []
            for h in new_hosts_config:
                ip = h["ip"]
                new_ips.add(ip)
                interval = h.get("interval", default_interval)
                group = h.get("group", "General")
                name = h["name"]
                always_on = bool(h.get("always_on", True))
                specs = h.get("specs") if isinstance(h.get("specs"), dict) else {}
                notes = h.get("notes", "") if isinstance(h.get("notes"), str) else ""
                links = h.get("links") if isinstance(h.get("links"), dict) else {}
                services = h.get("services") if isinstance(h.get("services"), list) else []
                strict = bool(h.get("strict", False))
                alert = bool(h.get("alert", True))
                if ip in current_by_ip:
                    existing = current_by_ip[ip]
                    name_changed = existing.name != name or existing.group != group
                    monitoring_disabled = existing.always_on and not always_on
                    existing.name = name
                    existing.group = group
                    existing.interval = interval
                    existing.always_on = always_on
                    existing.specs = specs
                    existing.notes = notes
                    existing.links = links
                    # Reset service_results for ports that have been removed
                    new_ports = {s.get("port") for s in services if s.get("port") is not None}
                    existing.service_results = {p: v for p, v in existing.service_results.items() if p in new_ports}
                    existing.services = services
                    existing.strict = strict
                    existing.alert = alert
                    if name_changed and self.incident_log:
                        self.incident_log.update_host_info(existing)
                    if monitoring_disabled and self.incident_log:
                        self.incident_log._close_orphaned_incident(ip)
                    rebuilt.append(existing)
                else:
                    rebuilt.append(self._spawn(name, ip, group, interval, always_on, specs, notes, links, services, strict, alert))
            for ip, existing in current_by_ip.items():
                if ip not in new_ips:
                    existing.stop_event.set()
                    # Close any ongoing incident for this IP. Without this, the
                    # incident stays "ongoing" forever in the Events tab even
                    # though the host is no longer monitored.
                    if self.incident_log:
                        try:
                            self.incident_log._close_orphaned_incident(ip)
                        except Exception as e:
                            logging.warning(f"Could not close orphaned incident for {ip}: {e}")
            self.hosts = rebuilt


# ============================================================================
# Incident log
# ============================================================================

class IncidentLog:
    """Records down/up incidents to the HistoryDB. Thread-safe.

    With SQLite-backed history, this becomes a thin wrapper around HistoryDB.
    The previous in-memory implementation was capped at 100 incidents and
    reset on restart - the DB-backed version preserves history forever
    (subject to the retention_days setting)."""

    def __init__(self, history_db=None):
        self.history_db = history_db
        self.lock = threading.Lock()

    def record_down(self, host):
        if self.history_db:
            self.history_db.open_incident(host.ip, host.name, host.group)

    def record_up(self, host):
        if self.history_db:
            self.history_db.close_incident(host.ip)

    def _close_orphaned_incident(self, ip):
        """Close any ongoing incident for an IP, used when a host has been
        removed from the config or had its IP changed. Unlike record_up, this
        doesn't need a HostState object."""
        if self.history_db:
            self.history_db.close_incident(ip)

    def update_host_info(self, host):
        """Called when a host is renamed/regrouped so existing open incidents
        keep matching the host's current display values."""
        if self.history_db:
            self.history_db.update_incident_host_info(host.ip, host.name, host.group)

    def list_incidents(self):
        if not self.history_db:
            return []
        return self.history_db.list_incidents(limit=100)


# ============================================================================
# YAML config: read, validate, save (with backups)
# ============================================================================

def load_yaml(path):
    with open(path) as f:
        return yaml.safe_load(f)


def _validate_ip_or_hostname(s):
    """Returns True if s is a valid IPv4/IPv6 address or DNS-safe hostname.

    Used to reject argument-injection attempts via the 'ip' field in
    hosts.yaml entries. The ping subprocess takes the IP as a positional
    argument; without this, a leading '-' could be interpreted as a flag
    by some ping implementations. We pass '--' separator in ping_host
    too as belt-and-suspenders.
    """
    if not s or not isinstance(s, str):
        return False
    s = s.strip()
    if not s or len(s) > 253:
        return False
    # Reject anything starting with '-' immediately (argument injection guard)
    if s.startswith('-'):
        return False
    # Try as IP first
    try:
        import ipaddress
        ipaddress.ip_address(s)
        return True
    except (ValueError, ImportError):
        pass
    # Otherwise validate as hostname (RFC 1123-ish)
    import re as _re
    label_re = _re.compile(r'^[a-zA-Z0-9]([a-zA-Z0-9\-]{0,61}[a-zA-Z0-9])?$')
    labels = s.split('.')
    return all(label_re.match(label) for label in labels)


def _validate_url(s):
    """Allow only http/https URLs. Rejects javascript:, file:, etc."""
    if not isinstance(s, str):
        return False
    s = s.strip()
    return s.startswith("http://") or s.startswith("https://")


def validate_hosts_config(config):
    if not isinstance(config, dict):
        return False, "Config must be a YAML mapping."
    if "hosts" not in config or not isinstance(config["hosts"], list):
        return False, "Missing or invalid 'hosts' list."
    seen = set()
    for i, h in enumerate(config["hosts"]):
        if not isinstance(h, dict):
            return False, f"Host #{i+1} is not a mapping."
        if not str(h.get("name", "")).strip():
            return False, f"Host #{i+1} is missing a name."
        if not str(h.get("ip", "")).strip():
            return False, f"Host '{h.get('name','?')}' is missing an ip."
        ip = str(h["ip"]).strip()
        if not _validate_ip_or_hostname(ip):
            return False, f"Host '{h.get('name','?')}': '{ip}' is not a valid IP address or hostname"
        if ip in seen:
            return False, f"Duplicate IP: {ip}"
        seen.add(ip)
        if "interval" in h:
            try:
                iv = int(h["interval"])
                if iv < 5:
                    return False, f"Host '{h['name']}': interval must be >= 5 seconds."
            except (ValueError, TypeError):
                return False, f"Host '{h['name']}': interval must be an integer."
        if "always_on" in h and not isinstance(h["always_on"], bool):
            return False, f"Host '{h['name']}': always_on must be true or false."
        if "specs" in h:
            if not isinstance(h["specs"], dict):
                return False, f"Host '{h['name']}': specs must be a mapping."
            mac = h["specs"].get("mac")
            if mac:
                m = str(mac).strip()
                import re as _re
                if not _re.match(r"^([0-9a-fA-F]{2}[:-]){5}[0-9a-fA-F]{2}$", m) and not _re.match(r"^[0-9a-fA-F]{12}$", m):
                    return False, f"Host '{h['name']}': MAC address format looks invalid."
        if "notes" in h and not isinstance(h["notes"], str):
            return False, f"Host '{h['name']}': notes must be a string."
        if "strict" in h and not isinstance(h["strict"], bool):
            return False, f"Host '{h['name']}': strict must be true or false."
        if "alert" in h and not isinstance(h["alert"], bool):
            return False, f"Host '{h['name']}': alert must be true or false."
        if "services" in h:
            if not isinstance(h["services"], list):
                return False, f"Host '{h['name']}': services must be a list."
            for i, svc in enumerate(h["services"]):
                if not isinstance(svc, dict):
                    return False, f"Host '{h['name']}': service #{i+1} must be a mapping."
                port = svc.get("port")
                try:
                    p = int(port)
                except (ValueError, TypeError):
                    return False, f"Host '{h['name']}': service #{i+1} port must be an integer."
                if p < 1 or p > 65535:
                    return False, f"Host '{h['name']}': service #{i+1} port {p} is out of range (1-65535)."
                if "name" in svc and not isinstance(svc["name"], str):
                    return False, f"Host '{h['name']}': service #{i+1} name must be a string."
        if "links" in h:
            if not isinstance(h["links"], dict):
                return False, f"Host '{h['name']}': links must be a mapping."
            primary = h["links"].get("primary", "")
            if primary and not _validate_url(primary):
                return False, f"Host '{h['name']}': primary URL must start with http:// or https://"
            extras = h["links"].get("extras", []) or []
            if extras and not isinstance(extras, list):
                return False, f"Host '{h['name']}': links.extras must be a list."
            for i, x in enumerate(extras):
                if not isinstance(x, dict):
                    return False, f"Host '{h['name']}': link #{i+1} must be a mapping."
                if not x.get("name") or not isinstance(x["name"], str):
                    return False, f"Host '{h['name']}': link #{i+1} is missing a name."
                if not x.get("url") or not _validate_url(x["url"]):
                    return False, f"Host '{h['name']}': link '{x.get('name')}' has an invalid URL."
    return True, None


def save_hosts_config(path, new_hosts):
    try:
        existing = load_yaml(path) or {}
    except Exception:
        existing = {}
    settings = existing.get("settings", {})

    if os.path.exists(path):
        backup_dir = os.path.join(os.path.dirname(path), "backups")
        os.makedirs(backup_dir, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        shutil.copy2(path, os.path.join(backup_dir, f"hosts-{stamp}.yaml"))
        backups = sorted(os.listdir(backup_dir))
        for old in backups[:-10]:
            try:
                os.remove(os.path.join(backup_dir, old))
            except OSError:
                pass

    new_config = {}
    if settings:
        new_config["settings"] = settings
    new_config["hosts"] = new_hosts

    tmp_path = path + ".tmp"
    with open(tmp_path, "w") as f:
        yaml.safe_dump(new_config, f, sort_keys=False, default_flow_style=False)
    os.chmod(tmp_path, 0o600)  # hosts.yaml carries the OpenRouter key + ntfy topic
    os.replace(tmp_path, path)


# ============================================================================
# Authentication
# ============================================================================

import hashlib
import hmac
import secrets
import base64

class AuthManager:
    """Manages users, sessions, and brute-force protection.

    Storage: ~/netwatch/auth.json with structure:
        {
            "secret_key": "<hex string>",
            "users": {
                "alice": {
                    "salt": "<hex>",
                    "hash": "<hex>",
                    "admin": true,
                    "created": <unix epoch>
                }
            }
        }

    Sessions are stateless: the cookie itself contains username + expiry,
    signed with HMAC-SHA256 using secret_key. We don't track sessions
    server-side, so logout is just "delete the cookie" client-side
    (with a logout endpoint that issues a Set-Cookie deleting it).
    """

    SCRYPT_N = 16384  # cost factor; ~1 second on a Pi 5
    SCRYPT_R = 8
    SCRYPT_P = 1
    SESSION_DAYS = 7
    LOCKOUT_AFTER = 5
    LOCKOUT_MINUTES = 15

    def __init__(self, auth_path, db_path=None):
        self.path = auth_path
        self.lock = threading.Lock()
        self._failed_attempts = {}  # ip -> [timestamp, ...]
        self._attempts_db = self._open_attempts_db(db_path) if db_path else None
        self.data = self._load()
        if self._attempts_db:
            self._load_attempts()

    def _open_attempts_db(self, db_path):
        import sqlite3
        try:
            conn = sqlite3.connect(db_path, check_same_thread=False, isolation_level=None)
            conn.execute("PRAGMA journal_mode=WAL")
            conn.execute("PRAGMA synchronous=NORMAL")
            conn.execute(
                "CREATE TABLE IF NOT EXISTS login_attempts "
                "(ip TEXT NOT NULL, timestamp INTEGER NOT NULL)"
            )
            conn.execute(
                "CREATE INDEX IF NOT EXISTS idx_attempts_ip ON login_attempts(ip)"
            )
            return conn
        except Exception as e:
            logging.warning(f"AuthManager: could not open attempts DB at {db_path}: {e}; brute-force attempts will not persist")
            return None

    def _load_attempts(self):
        cutoff = time.time() - self.LOCKOUT_MINUTES * 60
        rows = self._attempts_db.execute(
            "SELECT ip, timestamp FROM login_attempts WHERE timestamp > ?", (cutoff,)
        ).fetchall()
        for ip, ts in rows:
            self._failed_attempts.setdefault(ip, []).append(ts)

    def _load(self):
        if not os.path.isfile(self.path):
            return {"secret_key": secrets.token_hex(32), "users": {}}
        try:
            with open(self.path) as f:
                data = json.load(f)
        except (json.JSONDecodeError, OSError):
            logging.warning(f"AuthManager: could not read {self.path}, starting fresh")
            return {"secret_key": secrets.token_hex(32), "users": {}}
        # Make sure required fields exist
        if "secret_key" not in data:
            data["secret_key"] = secrets.token_hex(32)
        if "users" not in data:
            data["users"] = {}
        return data

    def _save(self):
        # Atomic write so we never leave a corrupt file
        tmp = self.path + ".tmp"
        with open(tmp, "w") as f:
            json.dump(self.data, f, indent=2)
        os.chmod(tmp, 0o600)
        os.replace(tmp, self.path)

    @property
    def has_users(self):
        with self.lock:
            return len(self.data.get("users", {})) > 0

    def list_users(self):
        with self.lock:
            return [
                {"username": u, "admin": d.get("admin", False), "created": d.get("created", 0)}
                for u, d in self.data["users"].items()
            ]

    def _hash_password(self, password, salt_hex=None):
        if salt_hex is None:
            salt_hex = secrets.token_hex(16)
        salt = bytes.fromhex(salt_hex)
        h = hashlib.scrypt(
            password.encode("utf-8"), salt=salt,
            n=self.SCRYPT_N, r=self.SCRYPT_R, p=self.SCRYPT_P,
        )
        return salt_hex, h.hex()

    def create_user(self, username, password, admin=False):
        username = username.strip().lower()
        if not username or not username.replace("_", "").replace("-", "").isalnum():
            return False, "username must be alphanumeric (underscores and hyphens allowed)"
        if len(username) > 32:
            return False, "username too long (max 32 chars)"
        if len(password) < 8:
            return False, "password must be at least 8 characters"
        with self.lock:
            if username in self.data["users"]:
                return False, "user already exists"
            salt, pwhash = self._hash_password(password)
            self.data["users"][username] = {
                "salt":    salt,
                "hash":    pwhash,
                "admin":   bool(admin),
                "created": int(time.time()),
            }
            self._save()
        logging.info(f"AuthManager: created user '{username}' (admin={admin})")
        return True, None

    def delete_user(self, username):
        username = username.strip().lower()
        with self.lock:
            if username not in self.data["users"]:
                return False, "user not found"
            if len(self.data["users"]) == 1:
                return False, "cannot delete the last user"
            del self.data["users"][username]
            self._save()
        logging.info(f"AuthManager: deleted user '{username}'")
        return True, None

    def change_password(self, username, new_password):
        username = username.strip().lower()
        if len(new_password) < 8:
            return False, "password must be at least 8 characters"
        with self.lock:
            user = self.data["users"].get(username)
            if not user:
                return False, "user not found"
            salt, pwhash = self._hash_password(new_password)
            user["salt"] = salt
            user["hash"] = pwhash
            user["session_gen"] = int(user.get("session_gen", 0)) + 1
            self._save()
        return True, None

    def verify_password(self, username, password):
        username = username.strip().lower()
        with self.lock:
            user = self.data["users"].get(username)
            if not user:
                # Constant-time delay to prevent username enumeration
                self._hash_password(password, "00" * 16)
                return False
            _, expected = self._hash_password(password, user["salt"])
        return hmac.compare_digest(expected, user["hash"])

    def is_admin(self, username):
        with self.lock:
            user = self.data["users"].get(username)
            return bool(user and user.get("admin"))

    # ── Brute-force protection ──────────────────────────────────────────────

    def _prune_old_attempts(self, ip, now):
        cutoff = now - self.LOCKOUT_MINUTES * 60
        self._failed_attempts[ip] = [
            t for t in self._failed_attempts.get(ip, []) if t > cutoff
        ]
        if not self._failed_attempts[ip]:
            del self._failed_attempts[ip]
        if self._attempts_db:
            self._attempts_db.execute(
                "DELETE FROM login_attempts WHERE ip = ? AND timestamp <= ?",
                (ip, cutoff),
            )

    def is_locked_out(self, ip):
        now = time.time()
        with self.lock:
            self._prune_old_attempts(ip, now)
            attempts = self._failed_attempts.get(ip, [])
            return len(attempts) >= self.LOCKOUT_AFTER

    def record_failed_attempt(self, ip):
        now = time.time()
        with self.lock:
            self._prune_old_attempts(ip, now)
            self._failed_attempts.setdefault(ip, []).append(now)
            if self._attempts_db:
                self._attempts_db.execute(
                    "INSERT INTO login_attempts (ip, timestamp) VALUES (?, ?)",
                    (ip, int(now)),
                )

    def record_successful_login(self, ip):
        with self.lock:
            self._failed_attempts.pop(ip, None)
            if self._attempts_db:
                self._attempts_db.execute(
                    "DELETE FROM login_attempts WHERE ip = ?", (ip,)
                )

    def close(self):
        with self.lock:
            if self._attempts_db:
                self._attempts_db.close()
                self._attempts_db = None

    # ── Session cookie signing ──────────────────────────────────────────────

    def make_session_cookie(self, username):
        expiry = int(time.time()) + self.SESSION_DAYS * 86400
        with self.lock:
            user = self.data["users"].get(username, {})
            gen = int(user.get("session_gen", 0))
        payload = f"{username}|{expiry}|{gen}"
        secret = self.data["secret_key"].encode()
        sig = hmac.new(secret, payload.encode(), hashlib.sha256).hexdigest()
        token = base64.urlsafe_b64encode(payload.encode()).decode().rstrip("=")
        return f"{token}.{sig}"

    def verify_session_cookie(self, cookie_value):
        """Return (username, admin_bool) if valid, or (None, False).

        The payload carries a per-user session generation; bumping it on
        password change (or deleting the user) invalidates old cookies."""
        if not cookie_value or "." not in cookie_value:
            return None, False
        try:
            token, sig = cookie_value.rsplit(".", 1)
            padded = token + "=" * (-len(token) % 4)
            payload = base64.urlsafe_b64decode(padded.encode()).decode()
            username, expiry_s, gen_s = payload.split("|")
            expiry = int(expiry_s)
            gen = int(gen_s)
        except (ValueError, UnicodeDecodeError, base64.binascii.Error):
            return None, False
        if expiry < time.time():
            return None, False
        secret = self.data["secret_key"].encode()
        expected = hmac.new(secret, payload.encode(), hashlib.sha256).hexdigest()
        if not hmac.compare_digest(expected, sig):
            return None, False
        with self.lock:
            user = self.data["users"].get(username)
            if not user or int(user.get("session_gen", 0)) != gen:
                return None, False
            return username, bool(user.get("admin"))


def parse_cookies(cookie_header):
    """Parse a Cookie: header into a dict. Stdlib does this but we keep it tiny."""
    cookies = {}
    if not cookie_header:
        return cookies
    for pair in cookie_header.split(";"):
        if "=" in pair:
            k, v = pair.split("=", 1)
            cookies[k.strip()] = v.strip()
    return cookies


# ============================================================================
# Persistent history (SQLite)
# ============================================================================

class HistoryDB:
    """Thread-safe SQLite store for ping results and incidents.

    Schema:
        pings(id, host_ip, timestamp, is_up, latency_ms)
        incidents(id, host_ip, host_name, host_group, started, ended, duration_seconds)

    All timestamps are unix epoch seconds (integer).
    Pruning runs nightly to drop data older than retention_days.
    """

    SCHEMA = """
    CREATE TABLE IF NOT EXISTS pings (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        host_ip     TEXT NOT NULL,
        timestamp   INTEGER NOT NULL,
        is_up       INTEGER NOT NULL,
        latency_ms  REAL
    );
    CREATE INDEX IF NOT EXISTS idx_pings_host_time ON pings(host_ip, timestamp);
    CREATE INDEX IF NOT EXISTS idx_pings_time ON pings(timestamp);

    CREATE TABLE IF NOT EXISTS incidents (
        id               INTEGER PRIMARY KEY AUTOINCREMENT,
        host_ip          TEXT NOT NULL,
        host_name        TEXT NOT NULL,
        host_group       TEXT NOT NULL,
        started          INTEGER NOT NULL,
        ended            INTEGER,
        duration_seconds INTEGER
    );
    CREATE INDEX IF NOT EXISTS idx_incidents_host ON incidents(host_ip);
    CREATE INDEX IF NOT EXISTS idx_incidents_started ON incidents(started);
    CREATE INDEX IF NOT EXISTS idx_incidents_ended ON incidents(ended);

    CREATE TABLE IF NOT EXISTS ping_daily (
        day          TEXT NOT NULL,
        host_ip      TEXT NOT NULL,
        total        INTEGER NOT NULL,
        up           INTEGER NOT NULL,
        latency_avg  REAL,
        latency_min  REAL,
        latency_max  REAL,
        PRIMARY KEY (day, host_ip)
    );
    """

    FLUSH_MAX = 200          # safety flush if the 30s flusher falls behind
    BUFFER_HARD_CAP = 5000   # drop oldest beyond this if SQLite is wedged

    def __init__(self, db_path, retention_days=30):
        import sqlite3
        self.db_path = db_path
        self.retention_days = retention_days
        self.lock = threading.Lock()
        self._ping_buffer = []
        # check_same_thread=False because we share the connection across threads,
        # and we serialize writes with self.lock.
        self.conn = sqlite3.connect(db_path, check_same_thread=False, isolation_level=None)
        self.conn.execute("PRAGMA journal_mode=WAL")
        self.conn.execute("PRAGMA synchronous=NORMAL")
        # Cap the WAL file: without this SQLite never shrinks the -wal file
        # below its high-water mark (it hit 226MB via daily VACUUM).
        self.conn.execute("PRAGMA journal_size_limit=16777216")  # 16MB
        self.conn.executescript(self.SCHEMA)
        if not _column_exists(self.conn, "incidents", "alert_sent"):
            self.conn.execute("ALTER TABLE incidents ADD COLUMN alert_sent INTEGER DEFAULT 0")
            logging.info("HistoryDB: added alert_sent column to incidents")
        logging.info(f"HistoryDB: opened {db_path} (retention {retention_days} days)")

    def close(self):
        with self.lock:
            try:
                self._flush_pings_locked()
            except Exception:
                pass
            try:
                self.conn.close()
            except Exception:
                pass

    # ── Pings ───────────────────────────────────────────────────────────────

    def record_ping(self, host_ip, is_up, latency_ms):
        ts = int(time.time())
        with self.lock:
            self._ping_buffer.append((host_ip, ts, 1 if is_up else 0, latency_ms))
            if len(self._ping_buffer) >= self.FLUSH_MAX:
                self._flush_pings_locked()

    def _flush_pings_locked(self):
        """Write all buffered pings in one transaction. Caller holds self.lock.
        Batching cuts WAL write amplification ~10-30x vs per-ping commits."""
        if not self._ping_buffer:
            return
        if len(self._ping_buffer) > self.BUFFER_HARD_CAP:
            dropped = len(self._ping_buffer) - self.BUFFER_HARD_CAP
            del self._ping_buffer[:dropped]
            logging.warning(f"HistoryDB: dropped {dropped} buffered pings (DB unavailable?)")
        self.conn.execute("BEGIN")
        try:
            self.conn.executemany(
                "INSERT INTO pings (host_ip, timestamp, is_up, latency_ms) VALUES (?, ?, ?, ?)",
                self._ping_buffer,
            )
            self.conn.execute("COMMIT")
        except Exception:
            try: self.conn.execute("ROLLBACK")
            except Exception: pass
            raise
        self._ping_buffer.clear()

    def flush_pings(self):
        with self.lock:
            self._flush_pings_locked()

    def recent_pings(self, host_ip, limit=100):
        """Return up to `limit` most-recent pings for a host, oldest first.
        Used to repopulate the in-memory history deque on startup."""
        with self.lock:
            self._flush_pings_locked()
            cur = self.conn.execute(
                "SELECT is_up, latency_ms FROM pings "
                "WHERE host_ip = ? ORDER BY timestamp DESC LIMIT ?",
                (host_ip, limit),
            )
            rows = cur.fetchall()
        # Reverse so it's oldest-first (matches deque chronological order)
        return [(bool(r[0]), r[1]) for r in reversed(rows)]

    def latest_ping(self, host_ip):
        """Return (is_up, latency_ms, timestamp) of the most recent ping for
        the host, or None if no pings recorded."""
        with self.lock:
            self._flush_pings_locked()
            cur = self.conn.execute(
                "SELECT is_up, latency_ms, timestamp FROM pings "
                "WHERE host_ip = ? ORDER BY timestamp DESC LIMIT 1",
                (host_ip,),
            )
            row = cur.fetchone()
        if row is None:
            return None
        return (bool(row[0]), row[1], row[2])

    def history_series(self, host_ip, hours=24, target_points=180):
        """Bucketed latency/uptime series for charting, oldest first.

        Buckets are aligned absolute-time windows of span/target_points
        (min 60s). avg/min/max ignore down pings (NULL latency); up_pct is
        the fraction of up pings in the bucket."""
        hours = max(1, min(int(hours), 168))
        span = hours * 3600
        bucket = max(60, span // target_points)
        since = int(time.time()) - span
        with self.lock:
            self._flush_pings_locked()
            cur = self.conn.execute(
                "SELECT (timestamp / ?) * ? AS bucket_ts, "
                "AVG(latency_ms), MIN(latency_ms), MAX(latency_ms), AVG(is_up), COUNT(*) "
                "FROM pings WHERE host_ip = ? AND timestamp >= ? "
                "GROUP BY bucket_ts ORDER BY bucket_ts",
                (bucket, bucket, host_ip, since),
            )
            rows = cur.fetchall()
        points = [
            {"t": r[0],
             "avg": round(r[1], 2) if r[1] is not None else None,
             "min": round(r[2], 2) if r[2] is not None else None,
             "max": round(r[3], 2) if r[3] is not None else None,
             "up_pct": round((r[4] or 0) * 100, 1),
             "n": r[5]}
            for r in rows
        ]
        return {"bucket_seconds": bucket, "points": points}

    # ── Daily rollups ───────────────────────────────────────────────────────

    def rollup_days(self):
        """Aggregate complete (past, local-time) days into ping_daily.
        Idempotent (INSERT OR REPLACE re-rolls partial days). Runs daily
        from _prune_loop, BEFORE prune deletes the raw rows. ~29 rows/day,
        kept forever — months of uptime trends for pennies of storage."""
        with self.lock:
            self._flush_pings_locked()
            cur = self.conn.execute(
                "INSERT OR REPLACE INTO ping_daily "
                "(day, host_ip, total, up, latency_avg, latency_min, latency_max) "
                "SELECT date(timestamp, 'unixepoch', 'localtime'), host_ip, "
                "COUNT(*), SUM(is_up), AVG(latency_ms), MIN(latency_ms), MAX(latency_ms) "
                "FROM pings "
                "WHERE date(timestamp, 'unixepoch', 'localtime') < date('now', 'localtime') "
                "GROUP BY date(timestamp, 'unixepoch', 'localtime'), host_ip"
            )
            n = cur.rowcount
        if n:
            logging.info(f"HistoryDB: rolled up {n} host-day row(s)")
        return n

    def daily_history(self, host_ip, days=60):
        """Daily uptime/latency rollups for a host, oldest first."""
        days = max(1, min(int(days), 365))
        with self.lock:
            cur = self.conn.execute(
                "SELECT day, total, up, latency_avg, latency_min, latency_max "
                "FROM ping_daily WHERE host_ip = ? AND day >= date('now', 'localtime', ?) "
                "ORDER BY day",
                (host_ip, f"-{days} days"),
            )
            rows = cur.fetchall()
        return [
            {"day": r[0], "total": r[1], "up": r[2],
             "uptime_pct": round(r[2] / r[1] * 100, 2) if r[1] else None,
             "latency_avg": round(r[3], 2) if r[3] is not None else None,
             "latency_min": round(r[4], 2) if r[4] is not None else None,
             "latency_max": round(r[5], 2) if r[5] is not None else None}
            for r in rows
        ]

    # ── Incidents ───────────────────────────────────────────────────────────

    def open_incident(self, host_ip, host_name, host_group):
        """Open a new incident if there's no ongoing one for this host."""
        ts = int(time.time())
        with self.lock:
            # Check for an existing open incident
            cur = self.conn.execute(
                "SELECT id FROM incidents WHERE host_ip = ? AND ended IS NULL LIMIT 1",
                (host_ip,),
            )
            if cur.fetchone():
                return  # already an ongoing incident
            self.conn.execute(
                "INSERT INTO incidents (host_ip, host_name, host_group, started) "
                "VALUES (?, ?, ?, ?)",
                (host_ip, host_name, host_group, ts),
            )

    def close_incident(self, host_ip):
        """Close any open incident for this host."""
        ts = int(time.time())
        with self.lock:
            cur = self.conn.execute(
                "SELECT id, started FROM incidents WHERE host_ip = ? AND ended IS NULL LIMIT 1",
                (host_ip,),
            )
            row = cur.fetchone()
            if row is None:
                return
            inc_id, started = row
            duration = ts - started
            self.conn.execute(
                "UPDATE incidents SET ended = ?, duration_seconds = ? WHERE id = ?",
                (ts, duration, inc_id),
            )

    def list_incidents(self, limit=100):
        """Return incidents most-recent-first as a list of dicts. Ongoing ones
        come first; resolved ones follow ordered by start time descending."""
        from datetime import datetime as _dt
        now = int(time.time())
        with self.lock:
            cur = self.conn.execute(
                "SELECT id, host_ip, host_name, host_group, started, ended, duration_seconds "
                "FROM incidents ORDER BY ended IS NULL DESC, started DESC LIMIT ?",
                (limit,),
            )
            rows = cur.fetchall()
        result = []
        for inc_id, host_ip, host_name, host_group, started, ended, duration in rows:
            ongoing = ended is None
            dur = (now - started) if ongoing else (duration or 0)
            st = _dt.fromtimestamp(started)
            # Time-only for today's events; month+day prefix once a midnight
            # has passed so the list never shows ambiguous bare times.
            # NOTE: started_str is SERVER-local; clients grouping by started_ts
            # (browser-local) should derive display labels from started_ts.
            fmt = "%H:%M:%S" if st.date() == _dt.now().date() else "%b %d %H:%M"
            result.append({
                "host_ip":          host_ip,
                "host_name":        host_name,
                "host_group":       host_group,
                "started_ts":       started,
                "started_str":      st.strftime(fmt),
                "started_iso":      st.isoformat(),
                "ended_iso":        _dt.fromtimestamp(ended).isoformat() if ended else None,
                "duration_seconds": dur,
                "ongoing":          ongoing,
            })
        return result

    def update_incident_host_info(self, host_ip, host_name, host_group):
        """Keep host_name/host_group up to date on existing open incidents
        when a host gets renamed or moved between groups."""
        with self.lock:
            self.conn.execute(
                "UPDATE incidents SET host_name = ?, host_group = ? "
                "WHERE host_ip = ? AND ended IS NULL",
                (host_name, host_group, host_ip),
            )

    def mark_incident_alerted(self, host_ip):
        """Set alert_sent=1 for the open incident for host_ip."""
        with self.lock:
            self.conn.execute(
                "UPDATE incidents SET alert_sent = 1 "
                "WHERE host_ip = ? AND ended IS NULL",
                (host_ip,),
            )

    def get_open_incident_alert_status(self, host_ip):
        """Return alert_sent (bool) for the open incident, or None."""
        with self.lock:
            cur = self.conn.execute(
                "SELECT alert_sent FROM incidents "
                "WHERE host_ip = ? AND ended IS NULL LIMIT 1",
                (host_ip,),
            )
            row = cur.fetchone()
        if row is None:
            return None
        return bool(row[0])

    def get_last_closed_incident_alert_status(self, host_ip):
        """Return alert_sent (bool) for the most recently closed incident."""
        with self.lock:
            cur = self.conn.execute(
                "SELECT alert_sent FROM incidents "
                "WHERE host_ip = ? AND ended IS NOT NULL "
                "ORDER BY ended DESC LIMIT 1",
                (host_ip,),
            )
            row = cur.fetchone()
        if row is None:
            return None
        return bool(row[0])

    # ── Pruning ─────────────────────────────────────────────────────────────

    def prune(self):
        """Delete rows older than retention_days. Returns (pings_deleted, incidents_deleted)."""
        cutoff = int(time.time()) - self.retention_days * 86400
        with self.lock:
            r1 = self.conn.execute(
                "DELETE FROM pings WHERE timestamp < ?", (cutoff,)
            )
            pings_deleted = r1.rowcount
            r2 = self.conn.execute(
                "DELETE FROM incidents WHERE ended IS NOT NULL AND ended < ?", (cutoff,)
            )
            incidents_deleted = r2.rowcount
            # No VACUUM: with fixed retention the DB is steady-state and
            # freed pages get reused. Daily VACUUM rewrote the whole DB
            # through the WAL (~300MB/day of SD writes) for nothing.
            # Manual reclaim if ever needed: sqlite3 netwatch.db VACUUM.
            self.conn.execute("PRAGMA wal_checkpoint(TRUNCATE)")
        if pings_deleted or incidents_deleted:
            logging.info(
                f"HistoryDB: pruned {pings_deleted} pings, "
                f"{incidents_deleted} incidents older than {self.retention_days}d"
            )
        return pings_deleted, incidents_deleted


def _flush_loop(history_db, stop_event):
    """Flush buffered pings every 30s; final flush on shutdown."""
    while not stop_event.wait(30):
        try:
            history_db.flush_pings()
        except Exception as e:
            logging.warning(f"HistoryDB flush failed: {e}")
    try:
        history_db.flush_pings()
    except Exception:
        pass


def _prune_loop(history_db, stop_event):
    """Run prune() once a day until stop_event is set."""
    SECONDS_PER_DAY = 86400
    # Run first prune ~60s after startup so the system isn't busy at boot
    elapsed = SECONDS_PER_DAY - 60
    while not stop_event.is_set():
        if elapsed >= SECONDS_PER_DAY:
            try:
                history_db.rollup_days()
            except Exception as e:
                logging.warning(f"HistoryDB rollup failed: {e}")
            try:
                history_db.prune()
            except Exception as e:
                logging.warning(f"HistoryDB prune failed: {e}")
            elapsed = 0
        time.sleep(5)
        elapsed += 5


# ============================================================================
# Inventory (CMDB)
# ============================================================================

# Inventory type taxonomy. Each type has a list of "type-specific" properties
# stored in the properties JSON blob; common fields (system name, mac, ip,
# serial, notes, etc.) live as top-level columns and are shared across types.
INVENTORY_TYPES = ("host", "vm", "network", "ups", "disk", "peripheral", "tablet", "phone", "printer")

INVENTORY_TYPE_PROPERTIES = {
    "host": [],  # all fields are top-level (cpu, ram, os, etc.)
    "vm": [
        # VMs use ALL the host fields (cpu, ram, os, etc. all top-level)
        # AND these VM-specific fields stored in the properties JSON.
        ("hypervisor",     "string", "Hypervisor (Proxmox/KVM/ESXi/etc.)"),
        ("vcpu_count",     "int",    "vCPU count"),
        ("ram_alloc_gb",   "int",    "Allocated RAM (GB)"),
        ("disk_alloc_gb",  "int",    "Allocated disk (GB)"),
        ("autostart",      "bool",   "Auto-starts with host"),
    ],
    "network": [
        ("port_count",     "int",    "Port count"),
        ("poe_watts",      "int",    "PoE budget (W)"),
        ("managed",        "bool",   "Managed"),
        ("uplink_speed",   "string", "Uplink speed"),
    ],
    "ups": [
        ("capacity_va",       "int",    "Capacity (VA)"),
        ("capacity_wh",       "int",    "Capacity (Wh)"),
        ("runtime_min",       "int",    "Runtime (min) at full load"),
        ("battery_age_years", "string", "Battery age"),
    ],
    "disk": [
        ("capacity_gb",  "int",    "Capacity (GB)"),
        ("interface",    "string", "Interface (SATA/NVMe/USB)"),
        ("rpm",          "int",    "Spindle speed (RPM, blank for SSD)"),
        ("used_in",      "string", "Currently installed in"),
        ("health",       "string", "Health status"),
    ],
    "peripheral": [
        ("subtype",      "string", "Type (KVM, monitor, keyboard, etc.)"),
        ("model",        "string", "Model"),
    ],
}


class InventoryDB:
    """SQLite-backed inventory store. Lives in the same database as ping history
    so we have one file to back up / one connection lifecycle to manage."""

    SCHEMA = """
    CREATE TABLE IF NOT EXISTS inventory (
        id           INTEGER PRIMARY KEY AUTOINCREMENT,
        category     TEXT,
        system       TEXT NOT NULL,
        role         TEXT,
        cpu          TEXT,
        ram_gb       REAL,
        gpu          TEXT,
        architecture TEXT,
        os           TEXT,
        cpu_score    INTEGER,
        tdp_watts    INTEGER,
        tpm          TEXT,
        mac          TEXT,
        ip           TEXT,
        serial       TEXT,
        notes        TEXT,
        created_at   INTEGER NOT NULL,
        updated_at   INTEGER NOT NULL
    );
    CREATE UNIQUE INDEX IF NOT EXISTS idx_inv_mac ON inventory(mac) WHERE mac IS NOT NULL AND mac != '';
    CREATE INDEX IF NOT EXISTS idx_inv_category ON inventory(category);
    CREATE INDEX IF NOT EXISTS idx_inv_system ON inventory(system);
    """

    FIELDS = [
        "category", "system", "role", "cpu", "ram_gb", "gpu", "architecture",
        "os", "cpu_score", "tdp_watts", "tpm", "mac", "ip", "serial", "notes",
        "device_type", "properties",
    ]

    def __init__(self, history_db):
        """Reuses the connection from HistoryDB."""
        self.history_db = history_db
        self.lock = history_db.lock  # share the same lock to serialize writes
        self.conn = history_db.conn
        self.conn.executescript(self.SCHEMA)
        if not _column_exists(self.conn, "inventory", "device_type"):
            self.conn.execute("ALTER TABLE inventory ADD COLUMN device_type TEXT DEFAULT 'host' NOT NULL")
            logging.info("InventoryDB: added device_type column")
        if not _column_exists(self.conn, "inventory", "properties"):
            self.conn.execute("ALTER TABLE inventory ADD COLUMN properties TEXT")
            logging.info("InventoryDB: added properties column")
        # Connections table - records edges between inventory devices.
        # CREATE TABLE IF NOT EXISTS is idempotent so re-runs are safe.
        self.conn.executescript("""
            CREATE TABLE IF NOT EXISTS inventory_connections (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                from_device_id  INTEGER NOT NULL,
                to_device_id    INTEGER NOT NULL,
                from_port       TEXT,
                to_port         TEXT,
                connection_type TEXT DEFAULT 'ethernet',
                notes           TEXT,
                created_at      INTEGER NOT NULL,
                FOREIGN KEY (from_device_id) REFERENCES inventory(id) ON DELETE CASCADE,
                FOREIGN KEY (to_device_id)   REFERENCES inventory(id) ON DELETE CASCADE
            );
            CREATE INDEX IF NOT EXISTS idx_conn_from ON inventory_connections(from_device_id);
            CREATE INDEX IF NOT EXISTS idx_conn_to   ON inventory_connections(to_device_id);
        """)
        # SQLite needs PRAGMA foreign_keys=ON for CASCADE to actually work.
        # The HistoryDB connection might not have it on; flip it now.
        self.conn.execute("PRAGMA foreign_keys = ON")
        logging.info("InventoryDB: schema ready")

    @staticmethod
    def normalize_mac(mac):
        """Lowercase + colon-separated. Returns '' for falsy input."""
        if not mac:
            return ""
        s = str(mac).strip().lower()
        # Strip out anything that isn't hex
        clean = "".join(c for c in s if c in "0123456789abcdef")
        if len(clean) != 12:
            # Don't reformat if it's not a valid 12-hex-char MAC
            return s
        return ":".join(clean[i:i+2] for i in range(0, 12, 2))

    def get_device_type_map(self):
        """Return {ip: device_type} for all inventory records with a non-empty IP.
        When multiple records share an IP, the one with the highest id wins."""
        with self.lock:
            cur = self.conn.execute(
                "SELECT ip, device_type FROM inventory"
                " WHERE ip IS NOT NULL AND ip != ''"
                " ORDER BY id ASC"
            )
            return {row[0]: (row[1] or "host") for row in cur.fetchall()}

    def list_all(self):
        with self.lock:
            cur = self.conn.execute(
                "SELECT id, category, system, role, cpu, ram_gb, gpu, architecture, "
                "os, cpu_score, tdp_watts, tpm, mac, ip, serial, notes, "
                "device_type, properties, "
                "created_at, updated_at FROM inventory ORDER BY system COLLATE NOCASE"
            )
            cols = [d[0] for d in cur.description]
            rows = [dict(zip(cols, row)) for row in cur.fetchall()]
            for r in rows:
                self._decode_properties(r)
            return rows

    def get(self, inv_id):
        with self.lock:
            cur = self.conn.execute(
                "SELECT id, category, system, role, cpu, ram_gb, gpu, architecture, "
                "os, cpu_score, tdp_watts, tpm, mac, ip, serial, notes, "
                "device_type, properties, "
                "created_at, updated_at FROM inventory WHERE id = ?", (inv_id,)
            )
            row = cur.fetchone()
            if not row:
                return None
            cols = [d[0] for d in cur.description]
            rec = dict(zip(cols, row))
            self._decode_properties(rec)
            return rec

    def find_by_mac(self, mac):
        """Return inventory record matching a MAC (normalized), or None."""
        norm = self.normalize_mac(mac)
        if not norm:
            return None
        with self.lock:
            cur = self.conn.execute(
                "SELECT id, category, system, role, cpu, ram_gb, gpu, architecture, "
                "os, cpu_score, tdp_watts, tpm, mac, ip, serial, notes, "
                "device_type, properties, "
                "created_at, updated_at FROM inventory WHERE mac = ?", (norm,)
            )
            row = cur.fetchone()
            if not row:
                return None
            cols = [d[0] for d in cur.description]
            rec = dict(zip(cols, row))
            self._decode_properties(rec)
            return rec

    def _decode_properties(self, rec):
        """Decode the properties JSON blob into a dict in-place. If the
        blob is missing/malformed, set properties to {}."""
        raw = rec.get("properties")
        if raw is None or raw == "":
            rec["properties"] = {}
            return
        if isinstance(raw, dict):
            return  # already decoded
        try:
            import json
            rec["properties"] = json.loads(raw)
            if not isinstance(rec["properties"], dict):
                rec["properties"] = {}
        except (ValueError, TypeError):
            rec["properties"] = {}

    def create(self, data):
        """Insert a new record. data is a dict of field values."""
        clean = self._clean_input(data)
        if not clean.get("system"):
            return None, "system name is required"
        # Check for MAC conflict
        if clean.get("mac"):
            existing = self.find_by_mac(clean["mac"])
            if existing:
                return None, f"a record with MAC {clean['mac']} already exists ({existing['system']})"
        ts = int(time.time())
        with self.lock:
            cur = self.conn.execute(
                "INSERT INTO inventory (category, system, role, cpu, ram_gb, gpu, "
                "architecture, os, cpu_score, tdp_watts, tpm, mac, ip, serial, notes, "
                "device_type, properties, "
                "created_at, updated_at) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (clean.get("category"), clean["system"], clean.get("role"),
                 clean.get("cpu"), clean.get("ram_gb"), clean.get("gpu"),
                 clean.get("architecture"), clean.get("os"), clean.get("cpu_score"),
                 clean.get("tdp_watts"), clean.get("tpm"), clean.get("mac"),
                 clean.get("ip"), clean.get("serial"), clean.get("notes"),
                 clean.get("device_type") or "host",
                 clean.get("properties"),
                 ts, ts)
            )
            return cur.lastrowid, None

    def update(self, inv_id, data):
        clean = self._clean_input(data)
        if "system" in clean and not clean["system"]:
            return False, "system name cannot be empty"
        # MAC conflict check (if MAC is being set, ensure no other record has it)
        if clean.get("mac"):
            existing = self.find_by_mac(clean["mac"])
            if existing and existing["id"] != inv_id:
                return False, f"a different record with MAC {clean['mac']} already exists"
        ts = int(time.time())
        sets = []
        vals = []
        for f in self.FIELDS:
            if f in clean:
                sets.append(f"{f} = ?")
                vals.append(clean[f])
        if not sets:
            return True, None  # nothing to update
        sets.append("updated_at = ?")
        vals.append(ts)
        vals.append(inv_id)
        with self.lock:
            cur = self.conn.execute(
                f"UPDATE inventory SET {', '.join(sets)} WHERE id = ?", vals
            )
            if cur.rowcount == 0:
                return False, "record not found"
        return True, None

    # ─── Connections (inventory_connections table) ──────────────────────────
    # Connection types we accept. Anything else gets coerced to "ethernet".
    CONNECTION_TYPES = ("ethernet", "fiber", "wifi", "virtual", "power", "usb", "console", "other")

    def _normalize_conn_type(self, t):
        if t is None: return "ethernet"
        s = str(t).strip().lower()
        return s if s in self.CONNECTION_TYPES else "ethernet"

    def list_connections_for_device(self, device_id):
        """Return all connections involving this device, both directions.

        Each result includes the OTHER device's id and name (joined server-side
        so the UI doesn't need a second call), plus a "direction" field of
        either "out" (this device is the from end) or "in" (this device is
        the to end).
        """
        with self.lock:
            cur = self.conn.execute(
                "SELECT c.id, c.from_device_id, c.to_device_id, "
                "c.from_port, c.to_port, c.connection_type, c.notes, c.created_at, "
                "f.system AS from_name, f.device_type AS from_type, "
                "t.system AS to_name,   t.device_type AS to_type "
                "FROM inventory_connections c "
                "JOIN inventory f ON f.id = c.from_device_id "
                "JOIN inventory t ON t.id = c.to_device_id "
                "WHERE c.from_device_id = ? OR c.to_device_id = ? "
                "ORDER BY c.connection_type, c.created_at",
                (device_id, device_id),
            )
            cols = [d[0] for d in cur.description]
            rows = [dict(zip(cols, row)) for row in cur.fetchall()]
        for r in rows:
            r["direction"] = "out" if r["from_device_id"] == device_id else "in"
        return rows

    def list_all_connections(self):
        """Return all connections in the system. Used by the topology view."""
        with self.lock:
            cur = self.conn.execute(
                "SELECT id, from_device_id, to_device_id, from_port, to_port, "
                "connection_type, notes, created_at FROM inventory_connections "
                "ORDER BY connection_type, created_at"
            )
            cols = [d[0] for d in cur.description]
            return [dict(zip(cols, row)) for row in cur.fetchall()]

    def create_connection(self, data):
        """Create a new connection row. Returns (id, error_msg)."""
        try:
            from_id = int(data.get("from_device_id"))
            to_id   = int(data.get("to_device_id"))
        except (TypeError, ValueError):
            return None, "from_device_id and to_device_id required"
        if from_id == to_id:
            return None, "cannot connect a device to itself"
        # Verify both ends exist
        with self.lock:
            cur = self.conn.execute(
                "SELECT id FROM inventory WHERE id IN (?, ?)", (from_id, to_id)
            )
            ids = {r[0] for r in cur.fetchall()}
            if from_id not in ids or to_id not in ids:
                return None, "one or both devices do not exist"
        ctype = self._normalize_conn_type(data.get("connection_type"))
        from_port = (data.get("from_port") or "").strip() or None
        to_port   = (data.get("to_port") or "").strip() or None
        notes     = (data.get("notes") or "").strip() or None
        ts = int(time.time())
        with self.lock:
            cur = self.conn.execute(
                "INSERT INTO inventory_connections "
                "(from_device_id, to_device_id, from_port, to_port, "
                "connection_type, notes, created_at) "
                "VALUES (?, ?, ?, ?, ?, ?, ?)",
                (from_id, to_id, from_port, to_port, ctype, notes, ts),
            )
            return cur.lastrowid, None

    def update_connection(self, conn_id, data):
        """Update fields of an existing connection. Returns (ok, error_msg)."""
        fields, values = [], []
        if "from_port" in data:
            fields.append("from_port = ?")
            values.append((data.get("from_port") or "").strip() or None)
        if "to_port" in data:
            fields.append("to_port = ?")
            values.append((data.get("to_port") or "").strip() or None)
        if "connection_type" in data:
            fields.append("connection_type = ?")
            values.append(self._normalize_conn_type(data.get("connection_type")))
        if "notes" in data:
            fields.append("notes = ?")
            values.append((data.get("notes") or "").strip() or None)
        if not fields:
            return False, "no fields to update"
        values.append(conn_id)
        with self.lock:
            cur = self.conn.execute(
                "UPDATE inventory_connections SET " + ", ".join(fields)
                + " WHERE id = ?", tuple(values)
            )
            if cur.rowcount == 0:
                return False, "connection not found"
        return True, None

    def delete_connection(self, conn_id):
        with self.lock:
            cur = self.conn.execute(
                "DELETE FROM inventory_connections WHERE id = ?", (conn_id,)
            )
            if cur.rowcount == 0:
                return False, "connection not found"
        return True, None

    def delete(self, inv_id):
        with self.lock:
            cur = self.conn.execute("DELETE FROM inventory WHERE id = ?", (inv_id,))
            if cur.rowcount == 0:
                return False, "record not found"
        return True, None

    def replace_all(self, records):
        """Wipe inventory and bulk-insert. Used by 'Replace all' import mode."""
        with self.lock:
            self.conn.execute("DELETE FROM inventory")
        ok, fail = 0, []
        for rec in records:
            inv_id, err = self.create(rec)
            if err:
                fail.append({"system": rec.get("system", "?"), "error": err})
            else:
                ok += 1
        return ok, fail

    def _clean_input(self, data):
        """Coerce / validate field values."""
        import json as _json
        out = {}
        for f in self.FIELDS:
            if f not in data:
                continue
            v = data[f]
            if v is None or (isinstance(v, str) and v.strip() == ""):
                out[f] = None
                continue
            if f in ("ram_gb",):
                try: out[f] = float(v)
                except (ValueError, TypeError): out[f] = None
            elif f in ("cpu_score", "tdp_watts"):
                try: out[f] = int(float(v))
                except (ValueError, TypeError): out[f] = None
            elif f == "mac":
                out[f] = self.normalize_mac(v)
                if not out[f]:
                    out[f] = None
            elif f == "device_type":
                t = str(v).strip().lower()
                out[f] = t if t in INVENTORY_TYPES else "peripheral"
            elif f == "properties":
                # Accept a dict (preferred), serialize to JSON for storage.
                # Strings are passed through if they parse as JSON dicts.
                if isinstance(v, dict):
                    out[f] = _json.dumps(v)
                elif isinstance(v, str):
                    try:
                        parsed = _json.loads(v)
                        out[f] = _json.dumps(parsed) if isinstance(parsed, dict) else None
                    except (ValueError, TypeError):
                        out[f] = None
                else:
                    out[f] = None
            else:
                out[f] = str(v).strip() if v is not None else None
        return out


def export_inventory_to_xlsx(inventory_db, scope='hosts'):
    """Build an XLSX file in memory containing inventory records.

    scope='hosts' (default): exports only host-type records on a single sheet
      named "Inventory". Filename: netwatch-inventory-hosts-{hostname}-{date}.xlsx
    scope='all': exports all device types, one sheet per type that has records.
      Sheet order follows INV_TYPE_ORDER. Filename: netwatch-inventory-all-…xlsx

    Column layout matches the import format for round-tripping host records.
    Returns (bytes, filename) on success, or (None, error_msg) on failure.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        return None, "openpyxl not available"

    import io
    import socket
    from datetime import datetime as _dt

    COLUMNS = [
        ("Category",             "category"),
        ("System",               "system"),
        ("Role / Status",        "role"),
        ("CPU",                  "cpu"),
        ("RAM_GB",               "ram_gb"),
        ("GPU",                  "gpu"),
        ("Architecture",         "architecture"),
        ("OS",                   "os"),
        ("Estimated_CPU_Score",  "cpu_score"),
        ("Max_TDP_Watts",        "tdp_watts"),
        ("TPM_Version",          "tpm"),
        ("MAC_Primary",          "mac"),
        ("IP_Address",           "ip"),
        ("Service_Tag_Serial",   "serial"),
        ("Notes",                "notes"),
    ]

    SHEET_NAMES = {
        'host': 'Hosts', 'vm': 'VMs', 'network': 'Network',
        'ups': 'UPS', 'disk': 'Disks', 'peripheral': 'Peripherals',
        'tablet': 'Tablets', 'phone': 'Phones', 'printer': 'Printers',
    }
    TYPE_ORDER = ['host', 'vm', 'network', 'ups', 'disk', 'peripheral', 'tablet', 'phone', 'printer']

    def _write_sheet(ws, records):
        for col_idx, (header, _) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
        for row_idx, rec in enumerate(records, start=2):
            for col_idx, (_, field) in enumerate(COLUMNS, start=1):
                ws.cell(row=row_idx, column=col_idx, value=rec.get(field))
        for col_idx, (header, field) in enumerate(COLUMNS, start=1):
            max_len = len(header)
            for rec in records:
                val = rec.get(field)
                if val is not None and len(str(val)) > max_len:
                    max_len = len(str(val))
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
        ws.freeze_panes = "A2"

    try:
        hostname = socket.gethostname() or "unknown"
        date_str = _dt.now().strftime("%Y-%m-%d")

        all_records = inventory_db.list_all()

        if scope == 'all':
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # remove default blank sheet

            # Group records by device_type
            by_type = {}
            for r in all_records:
                dt = r.get('device_type') or 'host'
                by_type.setdefault(dt, []).append(r)

            # Write sheets in TYPE_ORDER, then any unrecognised types
            ordered = [t for t in TYPE_ORDER if t in by_type]
            extras  = [t for t in by_type if t not in TYPE_ORDER]
            for dt in ordered + extras:
                sheet_name = SHEET_NAMES.get(dt, dt.title())
                ws = wb.create_sheet(title=sheet_name)
                _write_sheet(ws, by_type[dt])

            filename = f"netwatch-inventory-all-{hostname}-{date_str}.xlsx"
        else:
            # scope == 'hosts' (default)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Inventory"
            records = [r for r in all_records
                       if (r.get("device_type") or "host") == "host"]
            _write_sheet(ws, records)
            filename = f"netwatch-inventory-hosts-{hostname}-{date_str}.xlsx"

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue(), filename

    except Exception as e:
        return None, f"export failed: {e}"


def import_inventory_from_xlsx(inventory_db, xlsx_bytes, mode="add"):
    """Parse an uploaded xlsx and insert records.

    Expected columns (by header name, case-insensitive, flexible):
      Category, System, Role / Status, CPU, RAM_GB, GPU, Architecture, OS,
      Estimated_CPU_Score, Max_TDP_Watts, TPM_Version, MAC_Primary,
      IP_Address, Service_Tag_Serial

    mode='add'      -> insert new records, skip ones whose MAC or
                       (system+ram+cpu) tuple matches existing
    mode='replace'  -> wipe inventory first, then insert all
    Returns (added_count, skipped_count, errors_list).
    """
    try:
        import openpyxl
    except ImportError:
        return 0, 0, [{"row": 0, "error": "openpyxl not installed"}]

    import io
    try:
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    except Exception as e:
        return 0, 0, [{"row": 0, "error": f"could not parse xlsx: {e}"}]

    ws = wb.active
    # Read first row as headers
    rows = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        return 0, 0, [{"row": 0, "error": "spreadsheet is empty"}]

    # Map header names to our field names (case-insensitive, flexible)
    HEADER_MAP = {
        "category": "category",
        "system": "system",
        "role / status": "role",
        "role/status": "role",
        "role": "role",
        "cpu": "cpu",
        "ram_gb": "ram_gb",
        "ram (gb)": "ram_gb",
        "ram": "ram_gb",
        "gpu": "gpu",
        "architecture": "architecture",
        "arch": "architecture",
        "os": "os",
        "estimated_cpu_score": "cpu_score",
        "cpu_score": "cpu_score",
        "cpu score": "cpu_score",
        "max_tdp_watts": "tdp_watts",
        "tdp": "tdp_watts",
        "tdp_watts": "tdp_watts",
        "tpm_version": "tpm",
        "tpm": "tpm",
        "mac_primary": "mac",
        "mac": "mac",
        "mac address": "mac",
        "ip_address": "ip",
        "ip": "ip",
        "service_tag_serial": "serial",
        "serial": "serial",
        "service tag": "serial",
        "notes": "notes",
    }

    col_to_field = {}
    for idx, header in enumerate(header_row):
        if header is None: continue
        key = str(header).strip().lower()
        field = HEADER_MAP.get(key)
        if field:
            col_to_field[idx] = field

    if "system" not in col_to_field.values():
        return 0, 0, [{"row": 0, "error": "no 'System' column found in spreadsheet"}]

    # Build list of records
    records = []
    for row_idx, row in enumerate(rows, start=2):  # start=2 because row 1 was header
        rec = {}
        for col_idx, value in enumerate(row):
            field = col_to_field.get(col_idx)
            if field and value is not None:
                rec[field] = value
        if not rec.get("system"):
            continue  # skip blank rows
        records.append(rec)

    if mode == "replace":
        ok, fail = inventory_db.replace_all(records)
        return ok, 0, fail

    # 'add' mode: skip duplicates by MAC, otherwise insert
    added = 0
    skipped = 0
    errors = []
    existing = inventory_db.list_all()
    existing_macs = {InventoryDB.normalize_mac(r.get("mac")) for r in existing if r.get("mac")}
    existing_systems = {(r.get("system") or "").lower() for r in existing}
    for rec in records:
        mac_norm = InventoryDB.normalize_mac(rec.get("mac"))
        sys_lower = (rec.get("system") or "").lower()
        if mac_norm and mac_norm in existing_macs:
            skipped += 1
            continue
        if not mac_norm and sys_lower in existing_systems:
            skipped += 1
            continue
        inv_id, err = inventory_db.create(rec)
        if err:
            errors.append({"system": rec.get("system"), "error": err})
        else:
            added += 1
            if mac_norm: existing_macs.add(mac_norm)
            existing_systems.add(sys_lower)
    return added, skipped, errors


# ============================================================================
# Backup
# ============================================================================

BACKUP_MANIFEST_VERSION = 1


def create_backup_tarball(config_path, auth_path):
    """Build a complete backup tarball in memory and return (bytes, filename).

    The SQLite snapshot uses sqlite3's online .backup API for consistency.
    Other files are read normally.

    `config_path` is the path to hosts.yaml; the netwatch directory and
    db path are derived from it. `auth_path` is the auth.json location
    (may not exist yet if no users are configured).
    """
    import io
    import json as _json
    import sqlite3
    import socket
    import tarfile
    import tempfile
    from datetime import datetime as _dt

    netwatch_dir = os.path.dirname(os.path.abspath(config_path))
    db_path      = os.path.join(netwatch_dir, "netwatch.db")
    monitor_path = os.path.join(netwatch_dir, "monitor.py")
    hostname     = socket.gethostname() or "unknown"
    iso_now      = _dt.now().strftime("%Y-%m-%dT%H-%M-%S")
    filename     = f"netwatch-backup-{hostname}-{iso_now}.tar.gz"

    manifest = {
        "manifest_version": BACKUP_MANIFEST_VERSION,
        "netwatch_version": VERSION,
        "created_at":       int(time.time()),
        "created_iso":      _dt.now().isoformat(),
        "source_hostname":  hostname,
        "files":            {},
    }

    # 1) Make a consistent SQLite snapshot to a temp file. We can't
    # tar.add() the live db directly because WAL writes might be active.
    snapshot_path = None
    if os.path.isfile(db_path):
        fd, snapshot_path = tempfile.mkstemp(prefix="nw_backup_", suffix=".db")
        os.close(fd)
        try:
            src = sqlite3.connect(db_path)
            try:
                dst = sqlite3.connect(snapshot_path)
                try:
                    src.backup(dst)
                finally:
                    dst.close()
            finally:
                src.close()
        except Exception:
            # If snapshot fails for any reason, clean up and re-raise
            try: os.unlink(snapshot_path)
            except OSError: pass
            raise

    # 2) Build the tarball in memory.
    try:
        buf = io.BytesIO()
        with tarfile.open(fileobj=buf, mode="w:gz") as tar:
            def _add(real_path, arcname, mode_override=None):
                if not os.path.isfile(real_path):
                    return
                ti = tar.gettarinfo(real_path, arcname=arcname)
                if mode_override is not None:
                    ti.mode = mode_override
                # Strip uid/gid - they're meaningless across machines
                ti.uid = 0; ti.gid = 0
                ti.uname = ""; ti.gname = ""
                with open(real_path, "rb") as f:
                    tar.addfile(ti, f)
                manifest["files"][os.path.basename(arcname)] = os.path.getsize(real_path)

            _add(monitor_path, "netwatch/monitor.py")
            _add(config_path,  "netwatch/hosts.yaml")
            _add(auth_path,    "netwatch/auth.json", mode_override=0o600)
            if snapshot_path:
                # Snapshot lands under the original db filename
                ti = tar.gettarinfo(snapshot_path, arcname="netwatch/netwatch.db")
                ti.uid = 0; ti.gid = 0
                ti.uname = ""; ti.gname = ""
                with open(snapshot_path, "rb") as f:
                    tar.addfile(ti, f)
                manifest["files"]["netwatch.db"] = os.path.getsize(snapshot_path)

            # Manifest goes in last so all file sizes are populated
            manifest_bytes = _json.dumps(manifest, indent=2).encode("utf-8")
            ti = tarfile.TarInfo(name="netwatch/metadata.json")
            ti.size = len(manifest_bytes)
            ti.mtime = int(time.time())
            ti.mode = 0o644
            tar.addfile(ti, io.BytesIO(manifest_bytes))

        return buf.getvalue(), filename, manifest
    finally:
        # Always clean up the snapshot file, even if tar.add fails
        if snapshot_path:
            try: os.unlink(snapshot_path)
            except OSError: pass


# ============================================================================
# ARP-based MAC detection
# ============================================================================

# Lock for serializing writes to hosts.yaml from the ping thread.
# Uses the same lock the host_manager uses for config reloads, but we need
# our own here since this module-level helper can't reach into HostManager.
_ARP_WRITE_LOCK = threading.Lock()

# Track which IPs we've already auto-saved a MAC for in this session.
# Prevents rewriting hosts.yaml on every successful ping.
_ARP_SAVED_THIS_SESSION = set()


def _detect_mac_for_ip(ip):
    """Read /proc/net/arp and return the MAC for ip, or None if not present
    or not yet resolved.

    Format is 6 fixed-position columns:
        IP address       HW type     Flags       HW address            Mask     Device

    Flags 0x2 (ATF_COM) means the entry is complete & resolved. Other values
    mean it's incomplete (no response yet) or has an error.
    """
    try:
        with open("/proc/net/arp") as f:
            lines = f.readlines()
    except OSError:
        return None
    # Skip header line
    for line in lines[1:]:
        parts = line.split()
        if len(parts) < 4:
            continue
        arp_ip, hw_type, flags, mac = parts[0], parts[1], parts[2], parts[3]
        if arp_ip != ip:
            continue
        # Only return resolved entries
        try:
            flag_int = int(flags, 16)
        except ValueError:
            continue
        if not (flag_int & 0x2):
            continue
        if mac == "00:00:00:00:00:00" or not mac:
            continue
        return mac.lower()
    return None


def _normalise_mac(s):
    """Normalise a MAC string for comparison. Returns lowercase colon form."""
    if not s:
        return ""
    clean = "".join(c for c in str(s).lower() if c in "0123456789abcdef")
    if len(clean) != 12:
        return str(s).lower().strip()
    return ":".join(clean[i:i+2] for i in range(0, 12, 2))


def _save_detected_mac(config_path, host_ip, detected_mac):
    """Write back to hosts.yaml: set specs.mac and specs.mac_auto for the
    host with the given IP. Atomic via temp-file-then-rename. Returns True
    on success.

    Race-aware: we re-read the config under our write lock so we don't
    clobber concurrent edits from POST /api/hosts. The HTTP handler also
    uses load_yaml + save under its own lock, so we coordinate through
    the file itself - last writer wins, but neither corrupts.
    """
    with _ARP_WRITE_LOCK:
        try:
            cfg = load_yaml(config_path) or {}
        except Exception as e:
            logging.warning(f"ARP detect: could not read {config_path}: {e}")
            return False
        hosts = cfg.get("hosts", []) or []
        changed = False
        for h in hosts:
            if not isinstance(h, dict):
                continue
            if h.get("ip") != host_ip:
                continue
            specs = h.get("specs")
            if not isinstance(specs, dict):
                specs = {}
                h["specs"] = specs
            existing_mac = specs.get("mac")
            if existing_mac:
                # Don't overwrite a user-set MAC. Caller already decided
                # whether to call us; this is just defensive.
                return False
            specs["mac"] = detected_mac
            specs["mac_auto"] = True
            changed = True
            break
        if not changed:
            return False
        # Atomic write
        tmp = config_path + ".tmp"
        try:
            import yaml
            with open(tmp, "w") as f:
                yaml.safe_dump(cfg, f, default_flow_style=False, sort_keys=False)
            os.chmod(tmp, 0o600)  # hosts.yaml carries the OpenRouter key + ntfy topic
            os.replace(tmp, config_path)
            logging.info(f"ARP detect: saved MAC {detected_mac} for {host_ip}")
            return True
        except Exception as e:
            try: os.unlink(tmp)
            except OSError: pass
            logging.warning(f"ARP detect: could not write {config_path}: {e}")
            return False


# ============================================================================
# Alerts (ntfy)
# ============================================================================

NTFY_DEFAULT_SERVER = "https://ntfy.sh"
NTFY_DOWN_THRESHOLD = 3

_DASHBOARD_URL_CACHE = None


def _get_dashboard_url(settings, port):
    """Return clickable dashboard URL for alert click actions."""
    global _DASHBOARD_URL_CACHE
    explicit = (settings or {}).get("dashboard_url")
    if explicit:
        return str(explicit).rstrip("/")
    if _DASHBOARD_URL_CACHE is not None:
        return _DASHBOARD_URL_CACHE if _DASHBOARD_URL_CACHE else None
    import subprocess
    try:
        result = subprocess.run(
            ["ip", "-o", "route", "get", "1.1.1.1"],
            capture_output=True, text=True, timeout=2,
        )
        if result.returncode == 0:
            parts = result.stdout.split()
            if "src" in parts:
                src_ip = parts[parts.index("src") + 1]
                _DASHBOARD_URL_CACHE = f"http://{src_ip}:{port}"
                return _DASHBOARD_URL_CACHE
    except (subprocess.TimeoutExpired, FileNotFoundError, OSError, ValueError):
        pass
    _DASHBOARD_URL_CACHE = ""
    return None


def send_ntfy_alert(settings, title, message, priority="default",
                    tags=None, click_url=None):
    """Send a ntfy notification. Returns True on success."""
    topic = (settings or {}).get("ntfy_topic")
    if not topic:
        return False
    server = (settings or {}).get("ntfy_server") or NTFY_DEFAULT_SERVER
    server = str(server).rstrip("/")
    url = f"{server}/{topic}"

    import urllib.request
    import urllib.error
    headers = {"Title": title, "Priority": priority}
    if tags: headers["Tags"] = tags
    if click_url: headers["Click"] = click_url

    data = (message or "").encode("utf-8")
    req = urllib.request.Request(url, data=data, method="POST", headers=headers)
    try:
        with urllib.request.urlopen(req, timeout=8) as resp:
            ok = 200 <= resp.status < 300
            if ok:
                logging.info(f"ntfy: sent alert '{title}' to {topic}")
            else:
                logging.warning(f"ntfy: server returned {resp.status} for '{title}'")
            return ok
    except urllib.error.URLError as e:
        logging.warning(f"ntfy: could not deliver '{title}': {e}")
        return False
    except Exception as e:
        logging.warning(f"ntfy: unexpected error sending '{title}': {e}")
        return False


def _send_alert_async(settings, title, message, priority, tags, click_url,
                      on_success=None):
    """Fire-and-forget wrapper. on_success runs only if delivery succeeds."""
    def _worker():
        ok = send_ntfy_alert(settings, title, message, priority=priority,
                             tags=tags, click_url=click_url)
        if ok and on_success:
            try: on_success()
            except Exception as e:
                logging.warning(f"ntfy: on_success callback failed: {e}")
    threading.Thread(target=_worker, daemon=True, name="ntfy-alert").start()


# ============================================================================
# NAS Poller (TrueNAS REST API)
# ============================================================================

class NASPoller:
    POLL_INTERVAL_SECONDS = 900    # 15 minutes
    REPLICATION_STALE_HOURS = 25   # grace window for daily replication tasks

    def __init__(self, auth_manager, alert_settings=None, alert_port=None):
        self._auth_manager = auth_manager
        self._alert_settings = alert_settings or {}
        self._alert_port = alert_port
        self._cache = {
            "reachable": False,
            "last_updated": None,
            "error": None,
            "pools": [],
            "replication_tasks": [],
        }
        self._lock = threading.Lock()
        self._alert_state = {}  # condition_id -> bool, True = currently alerting

    def get_cache(self):
        with self._lock:
            import copy
            return copy.deepcopy(self._cache)

    def _get_config(self):
        data = self._auth_manager.data if self._auth_manager else {}
        return data.get("truenas_url", ""), data.get("truenas_api_key", "")

    @staticmethod
    def _fetch(url, api_key, path):
        import urllib.request
        req = urllib.request.Request(
            url.rstrip("/") + path,
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        )
        with urllib.request.urlopen(req, timeout=10) as resp:
            return json.loads(resp.read().decode())

    @staticmethod
    def _parse_vdevs(data_vdevs):
        vdevs = []
        for v in data_vdevs:
            vdev = {
                "type": v.get("type", "DISK"),
                "name": v.get("name", ""),
                "status": v.get("status", "UNKNOWN"),
                "disks": [],
            }
            for child in v.get("children", []):
                vdev["disks"].append({
                    "name": child.get("disk") or child.get("name", ""),
                    "status": child.get("status", "UNKNOWN"),
                })
            if not vdev["disks"] and v.get("disk"):
                vdev["disks"].append({"name": v["disk"], "status": v.get("status", "UNKNOWN")})
            vdevs.append(vdev)
        return vdevs

    @staticmethod
    def _parse_scrub(scan):
        if not scan:
            return {"status": None, "end_time": None, "errors": 0}
        end_raw = scan.get("end_time")
        end_str = None
        if isinstance(end_raw, str):
            end_str = end_raw
        elif isinstance(end_raw, dict):
            ms = end_raw.get("$date", {})
            if isinstance(ms, dict):
                ms = ms.get("$numberLong")
            if ms:
                from datetime import timezone
                end_str = datetime.fromtimestamp(int(ms) / 1000, tz=timezone.utc).isoformat()
        return {"status": scan.get("state"), "end_time": end_str, "errors": scan.get("errors", 0)}

    @staticmethod
    def _next_cron_run(minute, hour, dom, month, dow):
        """Return next ISO datetime string for a simple cron expression (no ranges/steps/lists)."""
        from datetime import timedelta, timezone
        now = datetime.now(tz=timezone.utc).replace(second=0, microsecond=0)
        t = now + timedelta(minutes=1)
        for _ in range(366 * 24 * 60):
            if (month == "*" or t.month == int(month)) and \
               (dom == "*" or t.day == int(dom)) and \
               (dow == "*" or t.weekday() == (int(dow) - 1) % 7) and \
               (hour == "*" or t.hour == int(hour)) and \
               (minute == "*" or t.minute == int(minute)):
                return t.isoformat()
            t += timedelta(minutes=1)
        return None

    @classmethod
    def _parse_pool(cls, raw, scrub_tasks):
        name = raw.get("name", "")
        next_scrub = None
        for st in scrub_tasks:
            if st.get("pool") == name:
                sch = st.get("schedule", {}) or {}
                def _f(k): v = sch.get(k); return "*" if v is None else str(v)
                next_scrub = cls._next_cron_run(
                    _f("minute"), _f("hour"), _f("dom"), _f("month"), _f("dow"),
                )
                break
        return {
            "name": name,
            "status": raw.get("status", "UNKNOWN"),
            "capacity_used_bytes": raw.get("allocated", 0),
            "capacity_total_bytes": raw.get("size", 0),
            "vdevs": cls._parse_vdevs(raw.get("topology", {}).get("data", [])),
            "last_scrub": cls._parse_scrub(raw.get("scan")),
            "next_scrub": next_scrub,
        }

    @staticmethod
    def _parse_replication(raw):
        state = raw.get("state") or {}
        dt_raw = state.get("datetime") or state.get("time_finished")
        last_run = None
        if isinstance(dt_raw, str):
            last_run = dt_raw
        elif isinstance(dt_raw, dict):
            inner = dt_raw.get("$date") or dt_raw.get("$numberLong")
            if isinstance(inner, dict):
                inner = inner.get("$numberLong")
            last_run = inner
        return {
            "id": raw.get("id"),
            "name": raw.get("name", ""),
            "last_run": last_run,
            "last_state": state.get("state") or "UNKNOWN",
        }

    def start(self, stop_event):
        t = threading.Thread(target=self._loop, args=(stop_event,), daemon=True, name="nas-poller")
        t.start()
        return t

    def _loop(self, stop_event):
        self._poll()
        while not stop_event.is_set():
            stop_event.wait(timeout=self.POLL_INTERVAL_SECONDS)
            if not stop_event.is_set():
                self._poll()

    def _poll(self):
        url, api_key = self._get_config()
        if not url or not api_key:
            with self._lock:
                self._cache.update({"reachable": False, "error": "NAS not configured"})
            return
        try:
            pools_raw = self._fetch(url, api_key, "/api/v2.0/pool")
            scrub_tasks = self._fetch(url, api_key, "/api/v2.0/pool/scrub")
            replication_raw = self._fetch(url, api_key, "/api/v2.0/replication")
            pools = [self._parse_pool(p, scrub_tasks) for p in pools_raw]
            tasks = [self._parse_replication(r) for r in replication_raw]
            with self._lock:
                self._cache = {
                    "reachable": True,
                    "last_updated": datetime.now(tz=timezone.utc).isoformat(),
                    "error": None,
                    "pools": pools,
                    "replication_tasks": tasks,
                }
            self._check_alerts(pools, tasks)
        except Exception as e:
            logging.warning(f"NASPoller: poll failed: {e}")
            with self._lock:
                self._cache.update({"reachable": False, "error": str(e)})

    def _fire_alert(self, condition_id, title, message):
        if not self._alert_state.get(condition_id, False):
            self._alert_state[condition_id] = True
            click_url = _get_dashboard_url(self._alert_settings, self._alert_port or 8080)
            _send_alert_async(
                self._alert_settings, title, message,
                priority="high", tags="warning", click_url=click_url,
            )

    def _clear_alert(self, condition_id):
        self._alert_state[condition_id] = False

    def _check_alerts(self, pools, tasks):
        from datetime import timezone, timedelta
        for pool in pools:
            cid = f"pool_health_{pool['name']}"
            if pool["status"] != "ONLINE":
                self._fire_alert(cid, "Netwatch · NAS Alert",
                                 f"Pool \"{pool['name']}\" is {pool['status']}")
            else:
                self._clear_alert(cid)
            cid_scrub = f"scrub_errors_{pool['name']}"
            errors = (pool.get("last_scrub") or {}).get("errors", 0) or 0
            if int(errors) > 0:
                self._fire_alert(cid_scrub, "Netwatch · NAS Alert",
                                 f"Scrub on \"{pool['name']}\" found {errors} error(s)")
            else:
                self._clear_alert(cid_scrub)

        now = datetime.now(tz=timezone.utc)
        stale_delta = timedelta(hours=self.REPLICATION_STALE_HOURS)
        for task in tasks:
            cid = f"replication_{task['id'] or task['name']}"
            ok_states = ("SUCCESS", "FINISHED", "PENDING")
            failed = task["last_state"] not in ok_states
            stale = False
            last = None
            if task.get("last_run"):
                try:
                    last = datetime.fromisoformat(task["last_run"].rstrip("Z"))
                    if last.tzinfo is None:
                        last = last.replace(tzinfo=timezone.utc)
                    stale = (now - last) > stale_delta
                except (ValueError, TypeError):
                    pass
            if failed or stale:
                hours_old = int((now - last).total_seconds() // 3600) if last else 0
                reason = "failed" if failed else f"stale ({hours_old}h)"
                self._fire_alert(cid, "Netwatch · NAS Alert",
                                 f"Replication \"{task['name']}\" {reason}")
            else:
                self._clear_alert(cid)


# ============================================================================
# Wake-on-LAN
# ============================================================================

def _detect_broadcast_address():
    """Detect the directed broadcast address for the Pi's primary interface.

    Works for any subnet size (/24, /22, /16, etc.) by reading the actual
    netmask from `ip -o -f inet addr show`. Falls back to 255.255.255.255
    if detection fails for any reason.

    Returns (broadcast_str, source_str). source_str is one of:
      "auto:<iface>"  - successfully detected from a real interface
      "fallback"      - using 255.255.255.255
    """
    import subprocess
    import ipaddress
    try:
        # Step 1: find which interface the Pi uses to reach the LAN.
        # `ip route get` returns something like:
        #   1.1.1.1 via 192.168.1.1 dev eth0 src 192.168.1.42 ...
        result = subprocess.run(
            ["ip", "-o", "route", "get", "1.1.1.1"],
            capture_output=True, text=True, timeout=2,
        )
        if result.returncode != 0:
            return "255.255.255.255", "fallback"
        parts = result.stdout.split()
        # Find the "dev" token to get the interface name
        if "dev" not in parts:
            return "255.255.255.255", "fallback"
        iface = parts[parts.index("dev") + 1]

        # Step 2: get the CIDR for that interface.
        # `ip -o -f inet addr show eth0` returns lines like:
        #   3: eth0    inet 192.168.1.42/22 brd 192.168.3.255 scope global ...
        # The 'brd' token already gives us the broadcast - prefer it.
        result = subprocess.run(
            ["ip", "-o", "-f", "inet", "addr", "show", iface],
            capture_output=True, text=True, timeout=2,
        )
        if result.returncode != 0 or not result.stdout.strip():
            return "255.255.255.255", "fallback"

        line_parts = result.stdout.split()
        # Try the 'brd' field first (most reliable)
        if "brd" in line_parts:
            bcast = line_parts[line_parts.index("brd") + 1]
            return bcast, f"auto:{iface}"

        # Otherwise compute from the inet x.x.x.x/NN field
        if "inet" in line_parts:
            cidr = line_parts[line_parts.index("inet") + 1]
            net = ipaddress.IPv4Network(cidr, strict=False)
            return str(net.broadcast_address), f"auto:{iface}"

        return "255.255.255.255", "fallback"
    except (subprocess.TimeoutExpired, FileNotFoundError, ValueError, OSError):
        return "255.255.255.255", "fallback"


# Cache the detection result so we log once and don't re-run `ip` per packet.
_WOL_BROADCAST_CACHE = None

def _get_wol_broadcast():
    global _WOL_BROADCAST_CACHE
    if _WOL_BROADCAST_CACHE is None:
        bcast, source = _detect_broadcast_address()
        _WOL_BROADCAST_CACHE = bcast
        if source.startswith("auto:"):
            iface = source.split(":", 1)[1]
            logging.info(f"WoL: directed broadcast detected as {bcast} (via {iface})")
        else:
            logging.warning(
                f"WoL: could not detect subnet broadcast, falling back to {bcast}. "
                "Wake-on-LAN may not reach hosts on a /22 or larger subnet."
            )
    return _WOL_BROADCAST_CACHE


def send_wol_packet(mac_address):
    """Send a Wake-on-LAN magic packet to the given MAC. Returns (ok, msg).

    Uses the auto-detected directed broadcast address for the Pi's subnet,
    which works correctly for /22, /16, etc. - not just /24.
    """
    import socket
    mac = mac_address.replace(":", "").replace("-", "").lower()
    if len(mac) != 12 or not all(c in "0123456789abcdef" for c in mac):
        return False, "Invalid MAC address format"
    try:
        mac_bytes = bytes.fromhex(mac)
    except ValueError:
        return False, "Could not parse MAC address"
    # Magic packet: 6 bytes of 0xFF, followed by the MAC repeated 16 times.
    packet = b"\xff" * 6 + mac_bytes * 16
    bcast = _get_wol_broadcast()
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.setsockopt(socket.SOL_SOCKET, socket.SO_BROADCAST, 1)
        # Port 9 is the standard WoL discard port
        s.sendto(packet, (bcast, 9))
        s.close()
        logging.info(f"WoL: sent magic packet to {mac_address} via {bcast}")
        return True, None
    except OSError as e:
        return False, f"Network error: {e}"


# ============================================================================
# Pi self-monitoring
# ============================================================================

_PI_LOCAL_IPS_CACHE = None

def _get_pi_local_ips():
    """Return the set of local IPs this Pi has, used to identify which monitored
    host represents the Pi itself. Cached on first call."""
    global _PI_LOCAL_IPS_CACHE
    if _PI_LOCAL_IPS_CACHE is not None:
        return _PI_LOCAL_IPS_CACHE
    import subprocess
    ips = set()
    try:
        result = subprocess.run(
            ["ip", "-o", "-f", "inet", "addr"],
            capture_output=True, text=True, timeout=2,
        )
        if result.returncode == 0:
            for line in result.stdout.splitlines():
                parts = line.split()
                if "inet" in parts:
                    cidr = parts[parts.index("inet") + 1]
                    ip = cidr.split("/")[0]
                    ips.add(ip)
    except (subprocess.TimeoutExpired, FileNotFoundError, OSError):
        pass
    # Always include loopback so 127.0.0.1 in hosts.yaml also matches the Pi
    ips.add("127.0.0.1")
    _PI_LOCAL_IPS_CACHE = ips
    return ips


def _is_local_ip(ip):
    return ip in _get_pi_local_ips()


def read_pi_health():
    """Read current Pi system metrics. Returns a dict with whatever we could
    read; missing keys mean that metric is unavailable on this system."""
    health = {}

    # CPU temperature (Pi exposes this at /sys/class/thermal/thermal_zone0/temp
    # in millidegrees C). Fall back to None if unavailable.
    for path in ("/sys/class/thermal/thermal_zone0/temp",):
        try:
            with open(path) as f:
                health["cpu_temp_c"] = round(int(f.read().strip()) / 1000.0, 1)
                break
        except (OSError, ValueError):
            continue

    # Load average
    try:
        with open("/proc/loadavg") as f:
            parts = f.read().split()
            health["load_1m"]  = float(parts[0])
            health["load_5m"]  = float(parts[1])
            health["load_15m"] = float(parts[2])
    except (OSError, ValueError, IndexError):
        pass

    # CPU count for normalising load average
    try:
        health["cpu_count"] = os.cpu_count() or 1
    except Exception:
        pass

    # Memory (read /proc/meminfo)
    try:
        meminfo = {}
        with open("/proc/meminfo") as f:
            for line in f:
                k, _, v = line.partition(":")
                meminfo[k.strip()] = v.strip()
        def _kb(k):
            v = meminfo.get(k, "")
            try:
                return int(v.split()[0]) * 1024
            except (ValueError, IndexError):
                return None
        total = _kb("MemTotal")
        avail = _kb("MemAvailable")
        if total and avail is not None:
            used = total - avail
            health["mem_total_bytes"] = total
            health["mem_used_bytes"]  = used
            health["mem_pct"] = round(used / total * 100, 1)
    except OSError:
        pass

    # Disk usage of /
    try:
        st = os.statvfs("/")
        total = st.f_blocks * st.f_frsize
        free  = st.f_bavail * st.f_frsize
        used  = total - free
        health["disk_total_bytes"] = total
        health["disk_used_bytes"]  = used
        health["disk_pct"] = round(used / total * 100, 1) if total else None
    except OSError:
        pass

    # System uptime
    try:
        with open("/proc/uptime") as f:
            health["uptime_seconds"] = int(float(f.read().split()[0]))
    except (OSError, ValueError):
        pass

    return health


# ============================================================================
# Network discovery (nmap-based)
# ============================================================================

# Cache the most recent scan result. Scans take a few seconds, so we run
# them in a background thread and let the client poll for results.
_DISCOVERY_LOCK = threading.Lock()
_DISCOVERY_STATE = {
    "running":   False,
    "started":   None,    # unix epoch
    "finished":  None,    # unix epoch
    "subnet":    None,
    "results":   [],
    "error":     None,
}


def _check_nmap_available():
    """Return (ok, path_or_error)."""
    try:
        r = subprocess.run(
            ["which", "nmap"], capture_output=True, text=True, timeout=2
        )
        if r.returncode == 0 and r.stdout.strip():
            return True, r.stdout.strip()
        return False, "nmap not installed (try: sudo apt-get install nmap)"
    except (subprocess.TimeoutExpired, FileNotFoundError, OSError) as e:
        return False, f"could not check for nmap: {e}"


def _detect_subnet_cidr():
    """Detect the Pi's primary subnet in CIDR form, e.g. '192.168.0.0/22'.
    Reuses the same logic as _detect_broadcast_address."""
    import ipaddress
    try:
        # Find the outgoing interface
        result = subprocess.run(
            ["ip", "-o", "route", "get", "1.1.1.1"],
            capture_output=True, text=True, timeout=2,
        )
        if result.returncode != 0:
            return None
        parts = result.stdout.split()
        if "dev" not in parts:
            return None
        iface = parts[parts.index("dev") + 1]

        # Get the CIDR for that interface
        result = subprocess.run(
            ["ip", "-o", "-f", "inet", "addr", "show", iface],
            capture_output=True, text=True, timeout=2,
        )
        if result.returncode != 0:
            return None
        line_parts = result.stdout.split()
        if "inet" not in line_parts:
            return None
        cidr = line_parts[line_parts.index("inet") + 1]
        net = ipaddress.IPv4Network(cidr, strict=False)
        return str(net)
    except (subprocess.TimeoutExpired, FileNotFoundError, ValueError, OSError):
        return None


def _run_nmap_scan(subnet):
    """Run `nmap -sn -oX -` against the subnet. Parses XML output and returns
    a list of {ip, mac, hostname, vendor} dicts. Returns (results, error)."""
    import xml.etree.ElementTree as ET

    try:
        # -sn = ping scan (no ports). -oX - = XML to stdout.
        # -T4 = aggressive timing for faster scans on local networks.
        # -n  = no DNS resolution by nmap (we'll do reverse DNS ourselves
        #       per host below for hostname enrichment).
        result = subprocess.run(
            ["nmap", "-sn", "-T4", "-n", "-oX", "-", subnet],
            capture_output=True, text=True, timeout=120,
        )
    except subprocess.TimeoutExpired:
        return [], "scan timed out (>2 minutes)"
    except FileNotFoundError:
        return [], "nmap not installed"
    except OSError as e:
        return [], f"could not run nmap: {e}"

    if result.returncode != 0:
        return [], f"nmap exited with code {result.returncode}: {result.stderr.strip()[:200]}"

    try:
        root = ET.fromstring(result.stdout)
    except ET.ParseError as e:
        return [], f"could not parse nmap output: {e}"

    hosts = []
    for host_el in root.findall("host"):
        # Skip down hosts (status="down")
        status_el = host_el.find("status")
        if status_el is None or status_el.get("state") != "up":
            continue

        ip = mac = vendor = None
        for addr in host_el.findall("address"):
            addrtype = addr.get("addrtype")
            if addrtype == "ipv4":
                ip = addr.get("addr")
            elif addrtype == "mac":
                mac = addr.get("addr", "").lower()
                vendor = addr.get("vendor")

        if not ip:
            continue

        # Reverse DNS lookup
        hostname = None
        try:
            import socket
            hostname = socket.gethostbyaddr(ip)[0]
        except (socket.herror, socket.gaierror, OSError):
            pass

        hosts.append({
            "ip":       ip,
            "mac":      mac,
            "hostname": hostname,
            "vendor":   vendor,
        })

    # Sort by IP for stable display
    def _ip_key(h):
        try:
            return tuple(int(x) for x in h["ip"].split("."))
        except (ValueError, AttributeError):
            return (999, 999, 999, 999)
    hosts.sort(key=_ip_key)
    return hosts, None


def start_discovery_scan():
    """Kick off a scan in a background thread if one isn't already running.
    Returns (started, message)."""
    with _DISCOVERY_LOCK:
        if _DISCOVERY_STATE["running"]:
            return False, "scan already in progress"

        ok, info = _check_nmap_available()
        if not ok:
            return False, info

        subnet = _detect_subnet_cidr()
        if not subnet:
            return False, "could not auto-detect subnet"

        _DISCOVERY_STATE.update({
            "running":  True,
            "started":  int(time.time()),
            "finished": None,
            "subnet":   subnet,
            "results":  [],
            "error":    None,
        })

    def _worker():
        results, error = _run_nmap_scan(subnet)
        with _DISCOVERY_LOCK:
            _DISCOVERY_STATE.update({
                "running":  False,
                "finished": int(time.time()),
                "results":  results,
                "error":    error,
            })
        if error:
            logging.warning(f"Discovery scan failed: {error}")
        else:
            logging.info(f"Discovery scan finished: {len(results)} hosts found in {subnet}")

    threading.Thread(target=_worker, daemon=True, name="discovery-scan").start()
    return True, f"started scan on {subnet}"


def get_discovery_state():
    """Return a snapshot of current scan state for the API."""
    with _DISCOVERY_LOCK:
        return dict(_DISCOVERY_STATE)


# ============================================================================
# Dashboard HTML (loaded from dashboard.html at startup)
# ============================================================================

def _load_dashboard_html(base_dir):
    path = os.path.join(base_dir, "dashboard.html")
    with open(path, encoding="utf-8") as f:
        # {{VERSION}} markers cache-bust the /static/ asset URLs on upgrades
        return f.read().replace("{{VERSION}}", VERSION)



# ============================================================================
# HTTP server
# ============================================================================

def build_topology_payload(inventory_db, host_manager):
    """Bundle inventory records + connections + linked-host status into a
    single payload for the topology view. Doing this server-side cuts the
    frontend from 3 round trips to 1 and lets us join MAC -> host status
    without serialising the full host list."""
    if not inventory_db:
        return {"nodes": [], "edges": []}

    # Build a MAC -> host status lookup
    host_by_mac = {}
    if host_manager:
        for h in host_manager.list_hosts():
            d = h.to_dict()
            mac = (d.get("specs") or {}).get("mac")
            norm = InventoryDB.normalize_mac(mac) if mac else ""
            if norm:
                host_by_mac[norm] = {
                    "name":   d.get("name"),
                    "ip":     d.get("ip"),
                    "is_up":  d.get("is_up"),
                    "status": d.get("status"),
                }

    nodes = []
    for rec in inventory_db.list_all():
        norm_mac = InventoryDB.normalize_mac(rec.get("mac")) if rec.get("mac") else ""
        linked = host_by_mac.get(norm_mac) if norm_mac else None
        nodes.append({
            "id":          rec["id"],
            "name":        rec.get("system") or "(unnamed)",
            "category":    rec.get("category"),
            "device_type": rec.get("device_type") or "host",
            "linked_host": linked,
            # Status inherits from linked host. Devices without a linked
            # monitored host (peripherals, switches we don't monitor) show
            # as UNKNOWN which renders as a neutral border.
            "status":      (linked["status"] if linked else "UNKNOWN"),
            "is_up":       (linked["is_up"] if linked else None),
            "ip":          rec.get("ip"),
            "mac":         rec.get("mac"),
        })

    edges = []
    for c in inventory_db.list_all_connections():
        edges.append({
            "id":              c["id"],
            "source":          c["from_device_id"],
            "target":          c["to_device_id"],
            "from_port":       c["from_port"],
            "to_port":         c["to_port"],
            "connection_type": c["connection_type"],
        })

    return {"nodes": nodes, "edges": edges}


# Settings keys safe to expose via /api/status. Everything else (API keys,
# ntfy topic) stays server-side; the AI panel uses /api/ai-config instead.
SETTINGS_PUBLIC_KEYS = ("default_interval", "ping_timeout", "history_window",
                        "refresh_rate", "history_days")


def build_api_payload(host_manager, settings, incident_log=None, inventory_db=None):
    hosts = host_manager.list_hosts()
    events = incident_log.list_incidents() if incident_log else []
    device_types = inventory_db.get_device_type_map() if inventory_db else {}
    return {
        "generated": datetime.now().isoformat(),
        "settings":  {k: settings[k] for k in SETTINGS_PUBLIC_KEYS if k in settings},
        "summary": {
            "total":   len(hosts),
            "up":      sum(1 for h in hosts if h.is_up),
            "down":    sum(1 for h in hosts if not h.is_up and h.last_checked and h.always_on),
            "idle":    sum(1 for h in hosts if not h.is_up and h.last_checked and not h.always_on),
            "pending": sum(1 for h in hosts if not h.last_checked),
        },
        "hosts": [
            {**h.to_dict(), "device_type": device_types.get(h.ip, "host")}
            for h in hosts
        ],
        "events": events,
    }


_STATIC_FILES = {
    'main.css':    'text/css; charset=utf-8',
    'fonts.css':   'text/css; charset=utf-8',
    'utils.js':    'application/javascript; charset=utf-8',
    'core.js':     'application/javascript; charset=utf-8',
    'topology.js': 'application/javascript; charset=utf-8',
    'inventory.js':'application/javascript; charset=utf-8',
    'auth.js':     'application/javascript; charset=utf-8',
    'ai-panel.js': 'application/javascript; charset=utf-8',
    'nas.js':      'application/javascript; charset=utf-8',
    'd3.v7.min.js':    'application/javascript; charset=utf-8',
    'dmsans-300.woff2':'font/woff2',
    'dmsans-400.woff2':'font/woff2',
    'dmsans-500.woff2':'font/woff2',
    'dmsans-600.woff2':'font/woff2',
    'dmmono-400.woff2':'font/woff2',
    'dmmono-500.woff2':'font/woff2',
    'favicon.svg':     'image/svg+xml',
    'favicon-alert.svg':'image/svg+xml',
    'manifest.json':   'application/manifest+json',
    'icon-192.png':    'image/png',
    'icon-512.png':    'image/png',
    'apple-touch-icon.png':'image/png',
}

# ── Route handler functions (module-level; testable without HTTP) ─────────────

def _h_get_status(host_manager, settings, incident_log, inventory_db) -> tuple:
    return 200, build_api_payload(host_manager, settings, incident_log, inventory_db)


def _h_get_ai_config(settings: dict) -> tuple:
    api_key = settings.get("openrouter_api_key", "")
    if not api_key.strip():
        return 404, {"error": "ai_not_configured"}
    return 200, {
        "api_key": api_key,
        "model": settings.get("ai_model", "openrouter/free"),
    }


def _h_get_hosts(config_path: str) -> tuple:
    try:
        cfg = load_yaml(config_path) or {}
        return 200, {"hosts": cfg.get("hosts", [])}
    except Exception as e:
        return 500, {"error": f"Could not read config: {e}"}


def _h_get_pi_health() -> tuple:
    try:
        return 200, read_pi_health()
    except Exception as e:
        logging.exception("Error reading Pi health")
        return 500, {"error": str(e)}


def _h_get_nas(nas_poller) -> tuple:
    if nas_poller is None:
        return 503, {"reachable": False, "error": "NAS poller not available"}
    return 200, nas_poller.get_cache()


def _h_get_auth_status(auth_manager, current_user_fn) -> tuple:
    user, is_admin = current_user_fn() if auth_manager else (None, False)
    return 200, {
        "logged_in":      bool(user),
        "username":       user,
        "admin":          is_admin,
        "setup_required": bool(auth_manager and not auth_manager.has_users),
    }


def _h_get_auth_users(auth_manager) -> tuple:
    if not auth_manager:
        return 404, {"error": "auth disabled"}
    return 200, {"users": auth_manager.list_users()}


def _h_get_inventory(inventory_db, host_manager) -> tuple:
    try:
        items = inventory_db.list_all() if inventory_db else []
        host_map = {}
        if host_manager:
            for h in [h.to_dict() for h in host_manager.list_hosts()]:
                mac = (h.get("specs", {}) or {}).get("mac")
                if mac:
                    key = InventoryDB.normalize_mac(mac)
                    if key:
                        host_map[key] = {
                            "name":       h.get("name"),
                            "ip":         h.get("ip"),
                            "is_up":      h.get("is_up"),
                            "status":     h.get("status"),
                            "uptime_pct": h.get("uptime_pct"),
                        }
        for item in items:
            m = InventoryDB.normalize_mac(item.get("mac"))
            item["linked_host"] = host_map.get(m) if m else None
        return 200, {"items": items}
    except Exception as e:
        logging.exception("inventory list error")
        return 500, {"error": str(e)}


def _h_get_inventory_record(path: str, inventory_db, host_manager) -> tuple:
    try:
        inv_id = int(path.split("/")[-1])
    except ValueError:
        return 400, {"error": "invalid id"}
    if not inventory_db:
        return 500, {"error": "inventory not available"}
    rec = inventory_db.get(inv_id)
    if not rec:
        return 404, {"error": "not found"}
    m = InventoryDB.normalize_mac(rec.get("mac"))
    rec["linked_host"] = None
    if m and host_manager:
        for h in [hh.to_dict() for hh in host_manager.list_hosts()]:
            h_mac = (h.get("specs", {}) or {}).get("mac")
            if InventoryDB.normalize_mac(h_mac) == m:
                rec["linked_host"] = {
                    "name":       h.get("name"),
                    "ip":         h.get("ip"),
                    "is_up":      h.get("is_up"),
                    "status":     h.get("status"),
                    "uptime_pct": h.get("uptime_pct"),
                }
                break
    return 200, rec


def _h_get_topology(inventory_db, host_manager) -> tuple:
    if not inventory_db:
        return 500, {"error": "inventory not available"}
    try:
        return 200, build_topology_payload(inventory_db, host_manager)
    except Exception as e:
        logging.exception("topology fetch error")
        return 500, {"error": str(e)}


def _h_get_connections(inventory_db) -> tuple:
    if not inventory_db:
        return 500, {"error": "inventory not available"}
    try:
        return 200, {"items": inventory_db.list_all_connections()}
    except Exception as e:
        logging.exception("connections list error")
        return 500, {"error": str(e)}


def _h_get_connections_for_device(path: str, inventory_db) -> tuple:
    if not inventory_db:
        return 500, {"error": "inventory not available"}
    try:
        inv_id = int(path.split("/")[-2])
    except (ValueError, IndexError):
        return 400, {"error": "invalid id"}
    try:
        return 200, {"items": inventory_db.list_connections_for_device(inv_id)}
    except Exception as e:
        logging.exception("connections fetch error")
        return 500, {"error": str(e)}


def _h_get_discover(config_path: str) -> tuple:
    try:
        state = get_discovery_state()
        cfg = load_yaml(config_path) or {}
        known_ips = {h.get("ip") for h in cfg.get("hosts", []) if isinstance(h, dict)}
        state["results"] = [
            {**r, "already_monitored": r["ip"] in known_ips}
            for r in state.get("results", [])
        ]
        return 200, state
    except Exception as e:
        logging.exception("Error reading discovery state")
        return 500, {"error": str(e)}


def _h_get_history(path: str, history_db) -> tuple:
    if history_db is None:
        return 500, {"error": "history not available"}
    from urllib.parse import urlparse, parse_qs
    qs = parse_qs(urlparse(path).query)
    ip = (qs.get("ip", [""])[0] or "").strip()
    if not ip:
        return 400, {"error": "ip required"}
    try:
        hours = max(1, min(int(qs.get("hours", ["24"])[0]), 168))
    except ValueError:
        return 400, {"error": "hours must be an integer"}
    try:
        days = max(1, min(int(qs.get("days", ["60"])[0]), 365))
    except ValueError:
        return 400, {"error": "days must be an integer"}
    try:
        series = history_db.history_series(ip, hours=hours)
        daily = history_db.daily_history(ip, days=days)
    except Exception as e:
        logging.exception("history fetch error")
        return 500, {"error": str(e)}
    return 200, {"ip": ip, "hours": hours, **series, "daily": daily}


def _h_post_inventory_create(body: dict, inventory_db) -> tuple:
    if not inventory_db:
        return 500, {"error": "inventory not available"}
    inv_id, err = inventory_db.create(body)
    if err:
        return 400, {"error": err}
    return 200, {"ok": True, "id": inv_id}


def _h_post_inventory_update(path: str, body: dict, inventory_db) -> tuple:
    if not inventory_db:
        return 500, {"error": "inventory not available"}
    try:
        inv_id = int(path.split("/")[-1])
    except ValueError:
        return 400, {"error": "invalid id"}
    ok, err = inventory_db.update(inv_id, body)
    if not ok:
        return 400, {"error": err}
    return 200, {"ok": True}


def _h_post_inventory_delete(path: str, inventory_db) -> tuple:
    if not inventory_db:
        return 500, {"error": "inventory not available"}
    try:
        inv_id = int(path.split("/")[-2])
    except (ValueError, IndexError):
        return 400, {"error": "invalid id"}
    ok, err = inventory_db.delete(inv_id)
    if not ok:
        return 404, {"error": err}
    return 200, {"ok": True}


def _h_post_connection_create(path: str, body: dict, inventory_db) -> tuple:
    if not inventory_db:
        return 500, {"error": "inventory not available"}
    try:
        inv_id = int(path.split("/")[-2])
    except (ValueError, IndexError):
        return 400, {"error": "invalid id"}
    body = dict(body)
    body["from_device_id"] = inv_id
    new_id, err = inventory_db.create_connection(body)
    if err:
        return 400, {"error": err}
    return 200, {"ok": True, "id": new_id}


def _h_post_connection_update(path: str, body: dict, inventory_db) -> tuple:
    if not inventory_db:
        return 500, {"error": "inventory not available"}
    try:
        conn_id = int(path.split("/")[-1])
    except ValueError:
        return 400, {"error": "invalid id"}
    ok, err = inventory_db.update_connection(conn_id, body)
    if not ok:
        return 400, {"error": err}
    return 200, {"ok": True}


def _h_post_connection_delete(path: str, inventory_db) -> tuple:
    if not inventory_db:
        return 500, {"error": "inventory not available"}
    try:
        conn_id = int(path.split("/")[-2])
    except (ValueError, IndexError):
        return 400, {"error": "invalid id"}
    ok, err = inventory_db.delete_connection(conn_id)
    if not ok:
        return 404, {"error": err}
    return 200, {"ok": True}


def _h_post_discover() -> tuple:
    try:
        started, msg = start_discovery_scan()
        if started:
            return 200, {"ok": True, "message": msg}
        return 400, {"error": msg}
    except Exception as e:
        logging.exception("Error starting discovery scan")
        return 500, {"error": str(e)}


def _h_post_detect_mac(body: dict) -> tuple:
    ip = (body.get("ip") or "").strip()
    if not ip:
        return 400, {"error": "ip required"}
    try:
        mac = _detect_mac_for_ip(ip)
        if mac:
            return 200, {"ok": True, "mac": mac}
        return 404, {"error": "not in ARP cache (host may be offline or not yet pinged)"}
    except Exception as e:
        logging.exception("detect-mac error")
        return 500, {"error": str(e)}


def _h_post_wake(body: dict, host_manager, inventory_db) -> tuple:
    target_ip = body.get("ip", "").strip()
    if not target_ip:
        return 400, {"error": "ip is required"}
    target_host = next((h for h in host_manager.list_hosts() if h.ip == target_ip), None)
    if not target_host:
        return 404, {"error": "Host not found"}
    mac = (target_host.specs or {}).get("mac", "")
    if not mac and inventory_db:
        try:
            for rec in inventory_db.list_all():
                if rec.get("ip") == target_ip and rec.get("mac"):
                    mac = rec["mac"]
                    logging.info("WoL: using MAC from inventory record %s for %s",
                                 rec.get("id"), target_ip)
                    break
        except Exception as e:
            logging.warning("WoL inventory MAC lookup failed: %s", e)
    if not mac:
        return 400, {"error": "No MAC address configured for this host (in hosts.yaml or inventory)"}
    ok, err = send_wol_packet(mac)
    if ok:
        return 200, {"ok": True, "message": f"Magic packet sent to {mac}"}
    return 500, {"error": err or "Failed to send magic packet"}


def _h_post_hosts(body: dict, config_path: str, host_manager, settings: dict) -> tuple:
    new_hosts = body.get("hosts", [])
    if not isinstance(new_hosts, list):
        return 400, {"error": "'hosts' must be a list"}
    ok, err = validate_hosts_config({"hosts": new_hosts})
    if not ok:
        return 400, {"error": err}
    try:
        save_hosts_config(config_path, new_hosts)
        logging.info(f"hosts.yaml updated via web: {len(new_hosts)} hosts")
        host_manager.reload_from_config(new_hosts, settings.get("default_interval", 30))
        return 200, {"ok": True, "count": len(new_hosts)}
    except Exception as e:
        logging.exception("Error saving hosts")
        return 500, {"error": str(e)}


def _h_post_auth_users(body: dict, auth_manager) -> tuple:
    username = body.get("username", "")
    password = body.get("password", "")
    is_admin = bool(body.get("admin", False))
    ok, err = auth_manager.create_user(username, password, admin=is_admin)
    if not ok:
        return 400, {"error": err}
    return 200, {"ok": True}


def _h_post_auth_password(body: dict, user: str, auth_manager) -> tuple:
    current = body.get("current", "")
    new_pw = body.get("new", "")
    if not auth_manager.verify_password(user, current):
        return 401, {"error": "current password is incorrect"}
    ok, err = auth_manager.change_password(user, new_pw)
    if not ok:
        return 400, {"error": err}
    return 200, {"ok": True}


def _h_post_auth_user_delete(path: str, auth_manager) -> tuple:
    username = path[len("/api/auth/users/"):]
    if not username:
        return 400, {"error": "username required"}
    ok, err = auth_manager.delete_user(username)
    if not ok:
        return 400, {"error": err}
    return 200, {"ok": True}


def make_handler(host_manager, settings, config_path, incident_log=None, auth_manager=None, inventory_db=None, dashboard_html="", history_db=None, nas_poller=None):
    class Handler(BaseHTTPRequestHandler):
        def log_message(self, fmt, *args): pass

        def _client_ip(self):
            # Just use the direct connection IP - no proxy support for now
            return self.client_address[0] if self.client_address else "unknown"

        def _current_user(self):
            """Returns (username, is_admin) or (None, False) if not logged in."""
            if not auth_manager:
                return None, False
            cookies = parse_cookies(self.headers.get("Cookie", ""))
            return auth_manager.verify_session_cookie(cookies.get("nw_session", ""))

        def _require_auth(self, admin_only=False):
            """Returns True if request is authorised, else writes 401 and returns False.
            If no users exist yet, returns False with a 'setup_required' response so
            the frontend can prompt for first-run setup."""
            if not auth_manager:
                return True  # auth disabled entirely
            if not auth_manager.has_users:
                self._send_json(401, {"error": "setup_required",
                                      "message": "No users configured yet. Set up the first admin user."})
                return False
            user, is_admin = self._current_user()
            if not user:
                self._send_json(401, {"error": "auth_required"})
                return False
            if admin_only and not is_admin:
                self._send_json(403, {"error": "admin_required"})
                return False
            return True

        def _set_session_cookie(self, username):
            cookie = auth_manager.make_session_cookie(username)
            max_age = auth_manager.SESSION_DAYS * 86400
            self.send_header("Set-Cookie",
                f"nw_session={cookie}; Path=/; Max-Age={max_age}; HttpOnly; SameSite=Strict")

        def _clear_session_cookie(self):
            self.send_header("Set-Cookie",
                "nw_session=; Path=/; Max-Age=0; HttpOnly; SameSite=Strict")


        def _read_json_body(self, max_bytes=1024 * 1024):
            """Read a JSON request body with size cap. Returns (data, error_response).

            On success: (parsed_dict, None).
            On error: (None, True) and an HTTP error response has been written.
            """
            try:
                length = int(self.headers.get("Content-Length", 0))
            except (TypeError, ValueError):
                self._send_json(400, {"error": "invalid Content-Length"})
                return None, True
            if length > max_bytes:
                self._send_json(413, {"error": f"request body too large (max {max_bytes} bytes)"})
                return None, True
            try:
                body = self.rfile.read(length).decode()
                data = json.loads(body)
            except json.JSONDecodeError:
                self._send_json(400, {"error": "invalid JSON"})
                return None, True
            except (UnicodeDecodeError, ValueError):
                self._send_json(400, {"error": "invalid request body"})
                return None, True
            if not isinstance(data, dict):
                self._send_json(400, {"error": "expected a JSON object"})
                return None, True
            return data, None

        def _send_json(self, code, obj):
            body = json.dumps(obj).encode()
            self.send_response(code)
            self.send_header("Content-Type", "application/json")
            self.send_header("Content-Length", len(body))
            self.end_headers()
            self.wfile.write(body)

        def do_GET(self):
            if self.path in ("/", "/index.html"):
                body = dashboard_html.encode()
                self.send_response(200)
                self.send_header("Content-Type", "text/html; charset=utf-8")
                self.send_header("Content-Length", len(body))
                self.send_header("Cache-Control", "no-cache")
                self.end_headers()
                self.wfile.write(body)
                return
            if self.path.startswith('/static/'):
                fname = self.path[8:].split('?')[0]
                if fname not in _STATIC_FILES:
                    self._send_json(404, {'error': 'not found'})
                    return
                base_dir = os.path.dirname(os.path.abspath(config_path))
                fpath = os.path.join(base_dir, 'static', fname)
                try:
                    with open(fpath, 'rb') as f:
                        body = f.read()
                    self.send_response(200)
                    self.send_header('Content-Type', _STATIC_FILES[fname])
                    self.send_header('Content-Length', len(body))
                    # URLs carry ?v={VERSION}, so a day of caching is safe
                    self.send_header('Cache-Control', 'public, max-age=86400')
                    self.end_headers()
                    self.wfile.write(body)
                except FileNotFoundError:
                    self._send_json(404, {'error': f'static file not found: {fname}'})
                return
            if self.path == "/api/status":
                if not self._require_auth(): return
                self._send_json(*_h_get_status(host_manager, settings, incident_log, inventory_db))
                return
            if self.path == "/api/ai-config":
                if not self._require_auth(): return
                self._send_json(*_h_get_ai_config(settings))
                return
            if self.path == "/api/hosts":
                if not self._require_auth(): return
                self._send_json(*_h_get_hosts(config_path))
                return
            if self.path == "/api/pi-health":
                if not self._require_auth(): return
                self._send_json(*_h_get_pi_health())
                return
            if self.path == "/api/nas":
                if not self._require_auth(): return
                self._send_json(*_h_get_nas(nas_poller))
                return
            if self.path == "/api/auth/status":
                self._send_json(*_h_get_auth_status(auth_manager, self._current_user))
                return
            if self.path == "/api/auth/users":
                if not self._require_auth(admin_only=True): return
                self._send_json(*_h_get_auth_users(auth_manager))
                return
            if self.path == "/api/inventory-export" or self.path.startswith("/api/inventory-export?"):
                if not self._require_auth(admin_only=True): return
                if not inventory_db:
                    self._send_json(500, {"error": "inventory not available"}); return
                try:
                    from urllib.parse import urlparse as _up, parse_qs as _pqs
                    _scope = _pqs(_up(self.path).query).get('scope', ['hosts'])[0]
                    if _scope not in ('hosts', 'all'):
                        _scope = 'hosts'
                    data, result = export_inventory_to_xlsx(inventory_db, scope=_scope)
                    if data is None:
                        self._send_json(500, {"error": result}); return
                    self.send_response(200)
                    self.send_header("Content-Type",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    self.send_header("Content-Length", str(len(data)))
                    self.send_header("Content-Disposition", f'attachment; filename="{result}"')
                    self.end_headers()
                    self.wfile.write(data)
                    logging.info(f"Inventory export: {result} ({len(data)} bytes)")
                except Exception as e:
                    logging.exception("inventory export error")
                    self._send_json(500, {"error": str(e)})
                return
            if self.path == "/api/inventory":
                if not self._require_auth(): return
                self._send_json(*_h_get_inventory(inventory_db, host_manager))
                return
            if self.path == "/api/topology":
                if not self._require_auth(): return
                self._send_json(*_h_get_topology(inventory_db, host_manager))
                return
            if self.path == "/api/connections":
                if not self._require_auth(): return
                self._send_json(*_h_get_connections(inventory_db))
                return
            if (self.path.startswith("/api/inventory/") and self.path.endswith("/connections")):
                if not self._require_auth(): return
                self._send_json(*_h_get_connections_for_device(self.path, inventory_db))
                return
            if self.path.startswith("/api/inventory/") and self.path != "/api/inventory/":
                if not self._require_auth(): return
                self._send_json(*_h_get_inventory_record(self.path, inventory_db, host_manager))
                return
            if self.path == "/api/discover":
                if not self._require_auth(): return
                self._send_json(*_h_get_discover(config_path))
                return
            if self.path.startswith("/api/history"):
                if not self._require_auth(): return
                self._send_json(*_h_get_history(self.path, history_db))
                return
            self.send_response(404)
            self.end_headers()

        def do_POST(self):
            # Auth routes that set/clear cookies stay inline
            if self.path == "/api/auth/setup":
                if not auth_manager:
                    self._send_json(400, {"error": "auth disabled"}); return
                client_ip = self._client_ip()
                if client_ip not in ("127.0.0.1", "::1", "localhost"):
                    self._send_json(403, {
                        "error": "setup must be performed from localhost",
                        "message": "SSH to the Pi and run: curl -X POST http://localhost:8080/api/auth/setup -H 'Content-Type: application/json' -d '{\"username\":\"admin\",\"password\":\"...\"}'"
                    }); return
                if auth_manager.has_users:
                    self._send_json(400, {"error": "setup already complete"}); return
                try:
                    length = int(self.headers.get("Content-Length", 0))
                    body = self.rfile.read(length).decode()
                    data = json.loads(body)
                    username = data.get("username", "")
                    password = data.get("password", "")
                    ok, err = auth_manager.create_user(username, password, admin=True)
                    if not ok:
                        self._send_json(400, {"error": err}); return
                    self.send_response(200)
                    self.send_header("Content-Type", "application/json")
                    self._set_session_cookie(username)
                    self.end_headers()
                    self.wfile.write(json.dumps({"ok": True, "username": username}).encode())
                except json.JSONDecodeError:
                    self._send_json(400, {"error": "invalid JSON"})
                except Exception as e:
                    logging.exception("setup error")
                    self._send_json(500, {"error": str(e)})
                return

            if self.path == "/api/auth/login":
                if not auth_manager:
                    self._send_json(400, {"error": "auth disabled"}); return
                ip = self._client_ip()
                if auth_manager.is_locked_out(ip):
                    self._send_json(429, {"error": "too many failed attempts, try again in 15 minutes"}); return
                try:
                    length = int(self.headers.get("Content-Length", 0))
                    body = self.rfile.read(length).decode()
                    data = json.loads(body)
                    username = data.get("username", "")
                    password = data.get("password", "")
                    if auth_manager.verify_password(username, password):
                        auth_manager.record_successful_login(ip)
                        is_admin = auth_manager.is_admin(username.strip().lower())
                        self.send_response(200)
                        self.send_header("Content-Type", "application/json")
                        self._set_session_cookie(username.strip().lower())
                        self.end_headers()
                        self.wfile.write(json.dumps({
                            "ok": True,
                            "username": username.strip().lower(),
                            "admin": is_admin,
                        }).encode())
                    else:
                        auth_manager.record_failed_attempt(ip)
                        self._send_json(401, {"error": "invalid username or password"})
                except json.JSONDecodeError:
                    self._send_json(400, {"error": "invalid JSON"})
                except Exception as e:
                    logging.exception("login error")
                    self._send_json(500, {"error": str(e)})
                return

            if self.path == "/api/auth/logout":
                self.send_response(200)
                self.send_header("Content-Type", "application/json")
                self._clear_session_cookie()
                self.end_headers()
                self.wfile.write(b'{"ok":true}')
                return

            if self.path == "/api/inventory":
                if not self._require_auth(): return
                data, err = self._read_json_body()
                if err: return
                self._send_json(*_h_post_inventory_create(data, inventory_db))
                return

            if self.path.startswith("/api/inventory/") and self.path.endswith("/delete"):
                if not self._require_auth(): return
                self._send_json(*_h_post_inventory_delete(self.path, inventory_db))
                return

            if (self.path.startswith("/api/inventory/") and self.path.endswith("/connections")):
                if not self._require_auth(): return
                data, err = self._read_json_body()
                if err: return
                self._send_json(*_h_post_connection_create(self.path, data, inventory_db))
                return

            if (self.path.startswith("/api/connections/") and self.path.endswith("/delete")):
                if not self._require_auth(): return
                self._send_json(*_h_post_connection_delete(self.path, inventory_db))
                return

            if self.path.startswith("/api/connections/"):
                if not self._require_auth(): return
                data, err = self._read_json_body()
                if err: return
                self._send_json(*_h_post_connection_update(self.path, data, inventory_db))
                return

            if self.path.startswith("/api/inventory/"):
                if not self._require_auth(): return
                data, err = self._read_json_body()
                if err: return
                self._send_json(*_h_post_inventory_update(self.path, data, inventory_db))
                return

            if self.path == "/api/inventory-import":
                if not self._require_auth(): return
                if not inventory_db:
                    self._send_json(500, {"error": "inventory not available"}); return
                try:
                    ctype = self.headers.get("Content-Type", "")
                    length = int(self.headers.get("Content-Length", 0))
                    if length > 10 * 1024 * 1024:
                        self._send_json(400, {"error": "file too large (10MB max)"}); return
                    body = self.rfile.read(length)
                    if "multipart/form-data" not in ctype:
                        self._send_json(400, {"error": "expected multipart/form-data"}); return
                    boundary = None
                    for part in ctype.split(";"):
                        part = part.strip()
                        if part.startswith("boundary="):
                            boundary = part[9:].strip('"')
                    if not boundary:
                        self._send_json(400, {"error": "missing boundary"}); return
                    delimiter = ("--" + boundary).encode()
                    parts = body.split(delimiter)
                    file_bytes = None
                    mode = "add"
                    for p in parts:
                        if not p or p == b"--" or p.strip() in (b"--\r\n", b""):
                            continue
                        sep = p.find(b"\r\n\r\n")
                        if sep < 0: continue
                        headers_blob = p[:sep].decode("latin-1", errors="replace")
                        content = p[sep + 4:]
                        if content.endswith(b"\r\n"): content = content[:-2]
                        name = None
                        for hline in headers_blob.split("\r\n"):
                            if hline.lower().startswith("content-disposition"):
                                for piece in hline.split(";"):
                                    piece = piece.strip()
                                    if piece.startswith("name="):
                                        name = piece[5:].strip('"')
                        if name == "file": file_bytes = content
                        elif name == "mode":
                            try:
                                mode = content.decode().strip()
                                if mode not in ("add", "replace"): mode = "add"
                            except Exception: mode = "add"
                    if file_bytes is None:
                        self._send_json(400, {"error": "no file uploaded"}); return
                    added, skipped, errors = import_inventory_from_xlsx(inventory_db, file_bytes, mode=mode)
                    self._send_json(200, {"ok": True, "added": added, "skipped": skipped,
                                          "errors": errors[:20], "mode": mode})
                except Exception as e:
                    logging.exception("inventory import error")
                    self._send_json(500, {"error": str(e)})
                return

            if self.path == "/api/backup":
                if not self._require_auth(admin_only=True): return
                try:
                    auth_path_local = auth_manager.path if auth_manager else None
                    data, filename, manifest = create_backup_tarball(config_path, auth_path_local)
                    self.send_response(200)
                    self.send_header("Content-Type", "application/gzip")
                    self.send_header("Content-Length", str(len(data)))
                    self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
                    self.send_header("X-Netwatch-Backup-Version", str(manifest["manifest_version"]))
                    self.send_header("X-Netwatch-Source", manifest["source_hostname"])
                    self.end_headers()
                    self.wfile.write(data)
                    logging.info(f"Backup downloaded: {filename} ({len(data)} bytes)")
                except Exception as e:
                    logging.exception("Backup failed")
                    self._send_json(500, {"error": f"backup failed: {e}"})
                return

            if self.path == "/api/auth/users":
                if not self._require_auth(admin_only=True): return
                data, err = self._read_json_body()
                if err: return
                self._send_json(*_h_post_auth_users(data, auth_manager))
                return

            if self.path == "/api/auth/password":
                if not self._require_auth(): return
                user, _ = self._current_user()
                data, err = self._read_json_body()
                if err: return
                self._send_json(*_h_post_auth_password(data, user, auth_manager))
                return

            if self.path.startswith("/api/auth/users/"):
                if not self._require_auth(admin_only=True): return
                self._send_json(*_h_post_auth_user_delete(self.path, auth_manager))
                return

            if self.path == "/api/discover":
                if not self._require_auth(): return
                self._send_json(*_h_post_discover())
                return

            if self.path == "/api/detect-mac":
                if not self._require_auth(): return
                data, err = self._read_json_body()
                if err: return
                self._send_json(*_h_post_detect_mac(data))
                return

            if self.path == "/api/wake":
                if not self._require_auth(): return
                data, err = self._read_json_body()
                if err: return
                self._send_json(*_h_post_wake(data, host_manager, inventory_db))
                return

            if self.path == "/api/hosts":
                if not self._require_auth(admin_only=True): return
                data, err = self._read_json_body()
                if err: return
                self._send_json(*_h_post_hosts(data, config_path, host_manager, settings))
                return

            self.send_response(404)
            self.end_headers()

    return Handler


def start_web_server(host_manager, settings, config_path, port, stop_event, incident_log=None, auth_manager=None, inventory_db=None, dashboard_html="", history_db=None, nas_poller=None):
    server = ThreadingHTTPServer(("0.0.0.0", port), make_handler(host_manager, settings, config_path, incident_log, auth_manager, inventory_db, dashboard_html, history_db, nas_poller=nas_poller))
    server.timeout = 1
    logging.info(f"Web dashboard: http://0.0.0.0:{port}")
    while not stop_event.is_set():
        server.handle_request()
    server.server_close()


# ============================================================================
# TUI
# ============================================================================

def safe_addstr(win, y, x, text, attr=0):
    try:
        max_y, max_x = win.getmaxyx()
        if y < 0 or y >= max_y or x < 0 or x >= max_x: return
        allowed = max_x - x - 1
        if allowed <= 0: return
        win.addstr(y, x, text[:allowed], attr)
    except curses.error:
        pass


def clamp(text, width):
    if len(text) >= width:
        return text[:width - 1] + "\u2026"
    return text


def draw_hline(win, y, x, width, char="\u2500", attr=0):
    safe_addstr(win, y, x, char * width, attr)


def draw_box(win, y, x, h, w, attr=0):
    if w < 2 or h < 2: return
    safe_addstr(win, y,         x,         "\u256d", attr)
    safe_addstr(win, y,         x + w - 1, "\u256e", attr)
    safe_addstr(win, y + h - 1, x,         "\u2570", attr)
    safe_addstr(win, y + h - 1, x + w - 1, "\u256f", attr)
    for i in range(1, w - 1):
        safe_addstr(win, y,         x + i, "\u2500", attr)
        safe_addstr(win, y + h - 1, x + i, "\u2500", attr)
    for i in range(1, h - 1):
        safe_addstr(win, y + i, x,         "\u2502", attr)
        safe_addstr(win, y + i, x + w - 1, "\u2502", attr)


def uptime_bar(pct, width=10):
    if pct is None:
        return "\u2500" * width
    filled = int(round(pct / 100 * width))
    return "\u2588" * filled + "\u2591" * (width - filled)


def draw_tui(stdscr, host_manager, refresh_rate, port, stop_event):
    curses.start_color()
    curses.use_default_colors()
    curses.init_pair(1,  82,  -1)
    curses.init_pair(2,  196, -1)
    curses.init_pair(3,  214, -1)
    curses.init_pair(4,  39,  -1)
    curses.init_pair(5,  243, -1)
    curses.init_pair(6,  255, -1)
    curses.init_pair(7,  208, -1)
    curses.init_pair(8,  51,  -1)
    C_UP    = curses.color_pair(1) | curses.A_BOLD
    C_DOWN  = curses.color_pair(2) | curses.A_BOLD
    C_WAIT  = curses.color_pair(3) | curses.A_BOLD
    C_HDR   = curses.color_pair(4)
    C_HDRB  = curses.color_pair(4) | curses.A_BOLD
    C_MUTED = curses.color_pair(5)
    C_TEXT  = curses.color_pair(6)
    C_LATC  = curses.color_pair(7)
    C_SPARK = curses.color_pair(8)
    curses.curs_set(0)
    stdscr.nodelay(True)
    COL_IND, COL_NAME, COL_IP, COL_STAT, COL_LAT, COL_BAR, COL_PCT, COL_SPK, COL_CHK = 1, 4, 24, 42, 50, 62, 74, 82, 105

    while not stop_event.is_set():
        stdscr.erase()
        max_y, max_x = stdscr.getmaxyx()
        now_str = datetime.now().strftime("%a %Y-%m-%d  %H:%M:%S")
        hosts = host_manager.list_hosts()
        up_hosts = [h for h in hosts if h.is_up]
        dn_hosts = [h for h in hosts if not h.is_up and h.last_checked]
        latencies = [h.last_latency_ms for h in up_hosts if h.last_latency_ms]
        avg_lat = (sum(latencies) / len(latencies)) if latencies else None
        uptimes = [h.uptime_pct for h in hosts if h.uptime_pct is not None]
        avg_up = (sum(uptimes) / len(uptimes)) if uptimes else None

        row = 0
        safe_addstr(stdscr, row, 1, BRAND, C_HDRB)
        safe_addstr(stdscr, row, 1 + len(BRAND), f"  v{VERSION} \u00b7 homelab monitor", C_MUTED)
        url_str = f"http://0.0.0.0:{port}"
        safe_addstr(stdscr, row, max(0, max_x - len(now_str) - len(url_str) - 4), url_str, C_HDR)
        safe_addstr(stdscr, row, max(0, max_x - len(now_str) - 1), now_str, C_MUTED)
        row += 1
        draw_hline(stdscr, row, 0, max_x, "\u2500", C_MUTED)
        row += 1

        if max_x >= 60:
            num_cards = 5
            card_w = max_x // num_cards
            cards = [
                ("hosts up",    f"{len(up_hosts)} / {len(hosts)}",          C_UP),
                ("hosts down",  str(len(dn_hosts)),                          C_DOWN if dn_hosts else C_MUTED),
                ("avg latency", f"{avg_lat:.1f} ms" if avg_lat else "- ms",  C_LATC),
                ("avg uptime",  f"{avg_up:.1f}%" if avg_up else "-%",        C_UP if avg_up and avg_up >= 90 else C_WAIT),
                ("dashboard",   f":{port}",                                   C_HDR),
            ]
            for i, (label, val, color) in enumerate(cards):
                cx = i * card_w
                bw = card_w if i < num_cards - 1 else max_x - cx
                draw_box(stdscr, row, cx, 4, bw, C_MUTED)
                safe_addstr(stdscr, row + 1, cx + 2, label.upper(), C_MUTED)
                safe_addstr(stdscr, row + 2, cx + 2, val, color)
            row += 4

        draw_hline(stdscr, row, 0, max_x, "\u2500", C_MUTED)
        row += 1
        safe_addstr(stdscr, row, COL_IND,  "\u25cf",         C_MUTED)
        safe_addstr(stdscr, row, COL_NAME, "HOST",           C_HDR)
        safe_addstr(stdscr, row, COL_IP,   "IP ADDRESS",     C_HDR)
        safe_addstr(stdscr, row, COL_STAT, "STATUS",         C_HDR)
        safe_addstr(stdscr, row, COL_LAT,  "LATENCY",        C_HDR)
        safe_addstr(stdscr, row, COL_BAR,  "UPTIME",         C_HDR)
        safe_addstr(stdscr, row, COL_PCT,  "%",              C_HDR)
        if max_x > COL_SPK + 10:
            safe_addstr(stdscr, row, COL_SPK, "HISTORY",     C_HDR)
        if max_x > COL_CHK + 8:
            safe_addstr(stdscr, row, COL_CHK, "LAST PING",   C_HDR)
        row += 1
        draw_hline(stdscr, row, 0, max_x, "\u2500", C_MUTED)
        row += 1

        groups = {}
        for h in hosts:
            groups.setdefault(h.group, []).append(h)

        for group_name, group_hosts in groups.items():
            if row >= max_y - 2: break
            prefix = "\u2500\u2500 "
            suffix_len = max(0, max_x - len(prefix) - len(group_name) - 4)
            safe_addstr(stdscr, row, 0, prefix, C_MUTED)
            safe_addstr(stdscr, row, len(prefix), f" {group_name.upper()} ", C_HDRB)
            safe_addstr(stdscr, row, len(prefix) + len(group_name) + 2, " " + "\u2500" * suffix_len, C_MUTED)
            row += 1
            for host in group_hosts:
                if row >= max_y - 2: break
                with host.lock:
                    is_up   = host.is_up
                    pend    = host.last_checked is None
                    name    = clamp(host.name, 18)
                    ip      = host.ip
                    status  = host.status_str
                    latency = host.latency_str
                    upct    = host.uptime_pct
                    uptime  = host.uptime_str
                    spark   = host.spark_str(20)
                    checked = host.checked_str
                if pend:
                    ind_attr = C_WAIT; st_attr = C_WAIT; nm_attr = C_TEXT
                elif is_up:
                    ind_attr = C_UP;   st_attr = C_UP;   nm_attr = C_TEXT
                elif status == "IDLE":
                    ind_attr = C_MUTED; st_attr = C_MUTED; nm_attr = C_MUTED
                else:
                    ind_attr = C_DOWN; st_attr = C_DOWN; nm_attr = C_DOWN
                bar_str = uptime_bar(upct, 10)
                if upct is None:
                    bar_attr = C_MUTED
                elif upct >= 95:
                    bar_attr = C_UP
                elif upct >= 80:
                    bar_attr = C_WAIT
                else:
                    bar_attr = C_DOWN
                safe_addstr(stdscr, row, COL_IND,  "\u25cf",         ind_attr)
                safe_addstr(stdscr, row, COL_NAME, f"{name:<18}",    nm_attr)
                safe_addstr(stdscr, row, COL_IP,   f"{ip:<16}",      C_MUTED)
                safe_addstr(stdscr, row, COL_STAT, f"{status:<7}",   st_attr)
                safe_addstr(stdscr, row, COL_LAT,  f"{latency:<11}", C_LATC if is_up else C_MUTED)
                safe_addstr(stdscr, row, COL_BAR,  bar_str,          bar_attr)
                safe_addstr(stdscr, row, COL_PCT,  f" {uptime:<6}",  bar_attr)
                if max_x > COL_SPK + 10:
                    for i, ch in enumerate(spark):
                        if COL_SPK + i >= max_x - 1: break
                        sp_color = C_SPARK if ch == "\u2588" else (C_DOWN if ch == "\u2581" else C_MUTED)
                        safe_addstr(stdscr, row, COL_SPK + i, ch, sp_color)
                if max_x > COL_CHK + 8:
                    safe_addstr(stdscr, row, COL_CHK, checked, C_MUTED)
                row += 1

        if max_y >= 3:
            draw_hline(stdscr, max_y - 2, 0, max_x, "\u2500", C_MUTED)
            keys = " [q] quit  [r] refresh "
            info = f" dashboard: http://0.0.0.0:{port}  \u00b7  {len(hosts)} hosts  \u00b7  {refresh_rate}s interval "
            safe_addstr(stdscr, max_y - 1, 0, keys, C_MUTED)
            safe_addstr(stdscr, max_y - 1, max(0, max_x - len(info) - 1), info, C_MUTED)

        stdscr.refresh()
        for _ in range(refresh_rate * 10):
            key = stdscr.getch()
            if key in (ord('q'), ord('Q')):
                stop_event.set()
                return
            if key in (ord('r'), ord('R')):
                break
            time.sleep(0.1)


# ============================================================================
# Entry point
# ============================================================================

def main():
    parser = argparse.ArgumentParser(description="Netwatch - homelab ping monitor")
    parser.add_argument("--config", default="hosts.yaml")
    parser.add_argument("--no-tui", action="store_true", help="Headless / systemd mode")
    parser.add_argument("--no-web", action="store_true", help="Disable web dashboard")
    parser.add_argument("--port",   type=int, default=8080, help="Web server port")
    parser.add_argument("--log",    default="monitor.log")
    args = parser.parse_args()

    from logging.handlers import RotatingFileHandler
    _log_handler = RotatingFileHandler(args.log, maxBytes=10 * 1024 * 1024, backupCount=3)
    _log_handler.setFormatter(logging.Formatter("%(asctime)s %(message)s", datefmt="%Y-%m-%d %H:%M:%S"))
    logging.basicConfig(level=logging.INFO, handlers=[_log_handler])

    config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), args.config)
    base_dir = os.path.dirname(os.path.abspath(__file__))
    dashboard_html = _load_dashboard_html(base_dir)
    config   = load_yaml(config_path)
    settings = config.get("settings", {})
    default_interval = settings.get("default_interval", 30)
    ping_timeout     = settings.get("ping_timeout", 2)
    history_window   = settings.get("history_window", 100)
    refresh_rate     = settings.get("refresh_rate", 5)

    stop_event = threading.Event()

    # Authentication
    config_dir = os.path.dirname(os.path.abspath(config_path))
    auth_path = os.path.join(config_dir, "auth.json")
    db_path = os.path.join(config_dir, "netwatch.db")
    auth_manager = AuthManager(auth_path, db_path=db_path)
    if auth_manager.has_users:
        print(f"[netwatch] Auth enabled - {len(auth_manager.list_users())} user(s) configured")
    else:
        print(f"[netwatch] Auth NOT YET CONFIGURED - visit the dashboard to set up the first admin user")

    # SQLite-backed history (persists pings & incidents across restarts)
    retention_days = int(settings.get("history_days", 30))
    history_db = HistoryDB(db_path, retention_days=retention_days)
    print(f"[netwatch] History DB -> {db_path} (retention {retention_days} days)")
    inventory_db = InventoryDB(history_db)
    inv_count = len(inventory_db.list_all())
    print(f"[netwatch] Inventory  -> {inv_count} record(s)")

    # Daily prune task
    pt = threading.Thread(target=_prune_loop, args=(history_db, stop_event), daemon=True, name="prune")
    pt.start()

    # Ping flush task (batched inserts land every 30s)
    ft = threading.Thread(target=_flush_loop, args=(history_db, stop_event), daemon=True, name="ping-flush")
    ft.start()

    # systemd sends SIGTERM on stop; route it through the KeyboardInterrupt
    # path so buffered pings get flushed and the WAL is checkpointed.
    import signal
    def _sigterm(_sig, _frame):
        raise KeyboardInterrupt
    signal.signal(signal.SIGTERM, _sigterm)

    incident_log = IncidentLog(history_db=history_db)
    host_manager = HostManager(
        config_path, ping_timeout, history_window, stop_event,
        incident_log, history_db,
        alert_settings=settings, alert_port=args.port,
    )
    host_manager.load_initial(config.get("hosts", []), default_interval)

    nas_poller = NASPoller(auth_manager, alert_settings=settings, alert_port=args.port)
    _nas_url, _ = nas_poller._get_config()
    if _nas_url:
        nas_poller.start(stop_event)
        print(f"[netwatch] NAS poller -> polling TrueNAS every {NASPoller.POLL_INTERVAL_SECONDS}s")

    if not args.no_web:
        web_settings = {**settings, "default_interval": default_interval}
        wt = threading.Thread(
            target=start_web_server,
            args=(host_manager, web_settings, config_path, args.port, stop_event, incident_log, auth_manager, inventory_db, dashboard_html, history_db),
            kwargs={"nas_poller": nas_poller},
            daemon=True
        )
        wt.start()
        print(f"[netwatch] Dashboard -> http://0.0.0.0:{args.port}")

    try:
        if args.no_tui:
            print(f"[netwatch] Monitoring {len(host_manager.list_hosts())} hosts")
            print("[netwatch] Headless mode. Ctrl+C to stop.")
            try:
                while True: time.sleep(1)
            except KeyboardInterrupt:
                stop_event.set()
        else:
            try:
                curses.wrapper(draw_tui, host_manager, refresh_rate, args.port, stop_event)
            except KeyboardInterrupt:
                pass
            finally:
                stop_event.set()
                print("\n[netwatch] Stopped.")
    finally:
        auth_manager.close()
        history_db.close()


if __name__ == "__main__":
    main()
