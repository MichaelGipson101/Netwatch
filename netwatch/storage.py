import os
import json
import time
import logging
import sqlite3
import threading

from netwatch import VERSION


def _column_exists(conn: "sqlite3.Connection", table: str, column: str) -> bool:
    """Return True if `column` exists in `table`. Both must be code-controlled identifiers."""
    rows = conn.execute(f"PRAGMA table_info({table})").fetchall()
    return any(row[1] == column for row in rows)


# ============================================================================
# Persistent history (SQLite)
# ============================================================================

class HistoryDB:
    """Thread-safe SQLite store for ping results and incidents.

    Schema:
        pings(id, host_ip, timestamp, is_up, latency_ms)
        incidents(id, host_ip, host_name, host_group, started, ended, duration_seconds)

    All timestamps are unix epoch seconds (integer).
    Pruning runs nightly to drop data older than retention_days.
    """

    SCHEMA = """
    CREATE TABLE IF NOT EXISTS pings (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        host_ip     TEXT NOT NULL,
        timestamp   INTEGER NOT NULL,
        is_up       INTEGER NOT NULL,
        latency_ms  REAL
    );
    CREATE INDEX IF NOT EXISTS idx_pings_host_time ON pings(host_ip, timestamp);
    CREATE INDEX IF NOT EXISTS idx_pings_time ON pings(timestamp);

    CREATE TABLE IF NOT EXISTS incidents (
        id               INTEGER PRIMARY KEY AUTOINCREMENT,
        host_ip          TEXT NOT NULL,
        host_name        TEXT NOT NULL,
        host_group       TEXT NOT NULL,
        started          INTEGER NOT NULL,
        ended            INTEGER,
        duration_seconds INTEGER
    );
    CREATE INDEX IF NOT EXISTS idx_incidents_host ON incidents(host_ip);
    CREATE INDEX IF NOT EXISTS idx_incidents_started ON incidents(started);
    CREATE INDEX IF NOT EXISTS idx_incidents_ended ON incidents(ended);

    CREATE TABLE IF NOT EXISTS ping_daily (
        day          TEXT NOT NULL,
        host_ip      TEXT NOT NULL,
        total        INTEGER NOT NULL,
        up           INTEGER NOT NULL,
        latency_avg  REAL,
        latency_min  REAL,
        latency_max  REAL,
        PRIMARY KEY (day, host_ip)
    );

    CREATE TABLE IF NOT EXISTS briefs (
        id            INTEGER PRIMARY KEY AUTOINCREMENT,
        created_ts    INTEGER NOT NULL,
        subject       TEXT    NOT NULL,
        stats_json    TEXT    NOT NULL,
        narrative     TEXT    NOT NULL,
        analysis_json TEXT
    );
    CREATE INDEX IF NOT EXISTS idx_briefs_ts ON briefs(created_ts);

    CREATE TABLE IF NOT EXISTS power_readings (
        id         INTEGER PRIMARY KEY AUTOINCREMENT,
        timestamp  INTEGER NOT NULL,
        watts      REAL,
        voltage    REAL,
        current_a  REAL,
        energy_kwh REAL
    );
    CREATE INDEX IF NOT EXISTS idx_power_ts ON power_readings(timestamp);
    """

    FLUSH_MAX = 200          # safety flush if the 30s flusher falls behind
    BUFFER_HARD_CAP = 5000   # drop oldest beyond this if SQLite is wedged

    def __init__(self, db_path, retention_days=30):
        import sqlite3
        self.db_path = db_path
        self.retention_days = retention_days
        self.lock = threading.Lock()
        self._ping_buffer = []
        # check_same_thread=False because we share the connection across threads,
        # and we serialize writes with self.lock.
        self.conn = sqlite3.connect(db_path, check_same_thread=False, isolation_level=None)
        self.conn.execute("PRAGMA journal_mode=WAL")
        self.conn.execute("PRAGMA synchronous=NORMAL")
        # Cap the WAL file: without this SQLite never shrinks the -wal file
        # below its high-water mark (it hit 226MB via daily VACUUM).
        self.conn.execute("PRAGMA journal_size_limit=16777216")  # 16MB
        self.conn.executescript(self.SCHEMA)
        if not _column_exists(self.conn, "incidents", "alert_sent"):
            self.conn.execute("ALTER TABLE incidents ADD COLUMN alert_sent INTEGER DEFAULT 0")
            logging.info("HistoryDB: added alert_sent column to incidents")
        logging.info(f"HistoryDB: opened {db_path} (retention {retention_days} days)")

    def close(self):
        with self.lock:
            try:
                self._flush_pings_locked()
            except Exception:
                pass
            try:
                self.conn.close()
            except Exception:
                pass

    # ── Pings ───────────────────────────────────────────────────────────────

    def record_ping(self, host_ip, is_up, latency_ms):
        ts = int(time.time())
        with self.lock:
            self._ping_buffer.append((host_ip, ts, 1 if is_up else 0, latency_ms))
            if len(self._ping_buffer) >= self.FLUSH_MAX:
                self._flush_pings_locked()

    def _flush_pings_locked(self):
        """Write all buffered pings in one transaction. Caller holds self.lock.
        Batching cuts WAL write amplification ~10-30x vs per-ping commits."""
        if not self._ping_buffer:
            return
        if len(self._ping_buffer) > self.BUFFER_HARD_CAP:
            dropped = len(self._ping_buffer) - self.BUFFER_HARD_CAP
            del self._ping_buffer[:dropped]
            logging.warning(f"HistoryDB: dropped {dropped} buffered pings (DB unavailable?)")
        self.conn.execute("BEGIN")
        try:
            self.conn.executemany(
                "INSERT INTO pings (host_ip, timestamp, is_up, latency_ms) VALUES (?, ?, ?, ?)",
                self._ping_buffer,
            )
            self.conn.execute("COMMIT")
        except Exception:
            try: self.conn.execute("ROLLBACK")
            except Exception: pass
            raise
        self._ping_buffer.clear()

    def flush_pings(self):
        with self.lock:
            self._flush_pings_locked()

    def recent_pings(self, host_ip, limit=100):
        """Return up to `limit` most-recent pings for a host, oldest first.
        Used to repopulate the in-memory history deque on startup."""
        with self.lock:
            self._flush_pings_locked()
            cur = self.conn.execute(
                "SELECT is_up, latency_ms FROM pings "
                "WHERE host_ip = ? ORDER BY timestamp DESC LIMIT ?",
                (host_ip, limit),
            )
            rows = cur.fetchall()
        # Reverse so it's oldest-first (matches deque chronological order)
        return [(bool(r[0]), r[1]) for r in reversed(rows)]

    def latest_ping(self, host_ip):
        """Return (is_up, latency_ms, timestamp) of the most recent ping for
        the host, or None if no pings recorded."""
        with self.lock:
            self._flush_pings_locked()
            cur = self.conn.execute(
                "SELECT is_up, latency_ms, timestamp FROM pings "
                "WHERE host_ip = ? ORDER BY timestamp DESC LIMIT 1",
                (host_ip,),
            )
            row = cur.fetchone()
        if row is None:
            return None
        return (bool(row[0]), row[1], row[2])

    def history_series(self, host_ip, hours=24, target_points=180):
        """Bucketed latency/uptime series for charting, oldest first.

        Buckets are aligned absolute-time windows of span/target_points
        (min 60s). avg/min/max ignore down pings (NULL latency); up_pct is
        the fraction of up pings in the bucket."""
        hours = max(1, min(int(hours), 168))
        span = hours * 3600
        bucket = max(60, span // target_points)
        since = int(time.time()) - span
        with self.lock:
            self._flush_pings_locked()
            cur = self.conn.execute(
                "SELECT (timestamp / ?) * ? AS bucket_ts, "
                "AVG(latency_ms), MIN(latency_ms), MAX(latency_ms), AVG(is_up), COUNT(*) "
                "FROM pings WHERE host_ip = ? AND timestamp >= ? "
                "GROUP BY bucket_ts ORDER BY bucket_ts",
                (bucket, bucket, host_ip, since),
            )
            rows = cur.fetchall()
        points = [
            {"t": r[0],
             "avg": round(r[1], 2) if r[1] is not None else None,
             "min": round(r[2], 2) if r[2] is not None else None,
             "max": round(r[3], 2) if r[3] is not None else None,
             "up_pct": round((r[4] or 0) * 100, 1),
             "n": r[5]}
            for r in rows
        ]
        return {"bucket_seconds": bucket, "points": points}

    # ── Daily rollups ───────────────────────────────────────────────────────

    def rollup_days(self):
        """Aggregate complete (past, local-time) days into ping_daily.
        Idempotent (INSERT OR REPLACE re-rolls partial days). Runs daily
        from _prune_loop, BEFORE prune deletes the raw rows. ~29 rows/day,
        kept forever — months of uptime trends for pennies of storage."""
        with self.lock:
            self._flush_pings_locked()
            cur = self.conn.execute(
                "INSERT OR REPLACE INTO ping_daily "
                "(day, host_ip, total, up, latency_avg, latency_min, latency_max) "
                "SELECT date(timestamp, 'unixepoch', 'localtime'), host_ip, "
                "COUNT(*), SUM(is_up), AVG(latency_ms), MIN(latency_ms), MAX(latency_ms) "
                "FROM pings "
                "WHERE date(timestamp, 'unixepoch', 'localtime') < date('now', 'localtime') "
                "GROUP BY date(timestamp, 'unixepoch', 'localtime'), host_ip"
            )
            n = cur.rowcount
        if n:
            logging.info(f"HistoryDB: rolled up {n} host-day row(s)")
        return n

    def daily_history(self, host_ip, days=60):
        """Daily uptime/latency rollups for a host, oldest first."""
        days = max(1, min(int(days), 365))
        with self.lock:
            cur = self.conn.execute(
                "SELECT day, total, up, latency_avg, latency_min, latency_max "
                "FROM ping_daily WHERE host_ip = ? AND day >= date('now', 'localtime', ?) "
                "ORDER BY day",
                (host_ip, f"-{days} days"),
            )
            rows = cur.fetchall()
        return [
            {"day": r[0], "total": r[1], "up": r[2],
             "uptime_pct": round(r[2] / r[1] * 100, 2) if r[1] else None,
             "latency_avg": round(r[3], 2) if r[3] is not None else None,
             "latency_min": round(r[4], 2) if r[4] is not None else None,
             "latency_max": round(r[5], 2) if r[5] is not None else None}
            for r in rows
        ]

    # ── Incidents ───────────────────────────────────────────────────────────

    def open_incident(self, host_ip, host_name, host_group, started_at=None):
        """Open a new incident if there's no ongoing one for this host."""
        ts = int(started_at) if started_at is not None else int(time.time())
        with self.lock:
            # Check for an existing open incident
            cur = self.conn.execute(
                "SELECT id FROM incidents WHERE host_ip = ? AND ended IS NULL LIMIT 1",
                (host_ip,),
            )
            if cur.fetchone():
                return  # already an ongoing incident
            self.conn.execute(
                "INSERT INTO incidents (host_ip, host_name, host_group, started) "
                "VALUES (?, ?, ?, ?)",
                (host_ip, host_name, host_group, ts),
            )

    def close_incident(self, host_ip):
        """Close any open incident for this host."""
        ts = int(time.time())
        with self.lock:
            cur = self.conn.execute(
                "SELECT id, started FROM incidents WHERE host_ip = ? AND ended IS NULL LIMIT 1",
                (host_ip,),
            )
            row = cur.fetchone()
            if row is None:
                return
            inc_id, started = row
            duration = ts - started
            self.conn.execute(
                "UPDATE incidents SET ended = ?, duration_seconds = ? WHERE id = ?",
                (ts, duration, inc_id),
            )

    def list_incidents(self, limit=100):
        """Return incidents most-recent-first as a list of dicts. Ongoing ones
        come first; resolved ones follow ordered by start time descending."""
        from datetime import datetime as _dt
        now = int(time.time())
        with self.lock:
            cur = self.conn.execute(
                "SELECT id, host_ip, host_name, host_group, started, ended, duration_seconds "
                "FROM incidents ORDER BY ended IS NULL DESC, started DESC LIMIT ?",
                (limit,),
            )
            rows = cur.fetchall()
        result = []
        for inc_id, host_ip, host_name, host_group, started, ended, duration in rows:
            ongoing = ended is None
            dur = (now - started) if ongoing else (duration or 0)
            st = _dt.fromtimestamp(started)
            # Time-only for today's events; month+day prefix once a midnight
            # has passed so the list never shows ambiguous bare times.
            # NOTE: started_str is SERVER-local; clients grouping by started_ts
            # (browser-local) should derive display labels from started_ts.
            fmt = "%H:%M:%S" if st.date() == _dt.now().date() else "%b %d %H:%M"
            result.append({
                "host_ip":          host_ip,
                "host_name":        host_name,
                "host_group":       host_group,
                "started_ts":       started,
                "started_str":      st.strftime(fmt),
                "started_iso":      st.isoformat(),
                "ended_iso":        _dt.fromtimestamp(ended).isoformat() if ended else None,
                "duration_seconds": dur,
                "ongoing":          ongoing,
            })
        return result

    def update_incident_host_info(self, host_ip, host_name, host_group):
        """Keep host_name/host_group up to date on existing open incidents
        when a host gets renamed or moved between groups."""
        with self.lock:
            self.conn.execute(
                "UPDATE incidents SET host_name = ?, host_group = ? "
                "WHERE host_ip = ? AND ended IS NULL",
                (host_name, host_group, host_ip),
            )

    def mark_incident_alerted(self, host_ip):
        """Set alert_sent=1 for the open incident for host_ip."""
        with self.lock:
            self.conn.execute(
                "UPDATE incidents SET alert_sent = 1 "
                "WHERE host_ip = ? AND ended IS NULL",
                (host_ip,),
            )

    def get_open_incident_alert_status(self, host_ip):
        """Return alert_sent (bool) for the open incident, or None."""
        with self.lock:
            cur = self.conn.execute(
                "SELECT alert_sent FROM incidents "
                "WHERE host_ip = ? AND ended IS NULL LIMIT 1",
                (host_ip,),
            )
            row = cur.fetchone()
        if row is None:
            return None
        return bool(row[0])

    def get_last_closed_incident_alert_status(self, host_ip):
        """Return alert_sent (bool) for the most recently closed incident."""
        with self.lock:
            cur = self.conn.execute(
                "SELECT alert_sent FROM incidents "
                "WHERE host_ip = ? AND ended IS NOT NULL "
                "ORDER BY ended DESC LIMIT 1",
                (host_ip,),
            )
            row = cur.fetchone()
        if row is None:
            return None
        return bool(row[0])

    # ── Briefs ──────────────────────────────────────────────────────────────

    def insert_brief(self, created_ts, subject, stats_json, narrative, analysis_json=None):
        with self.lock:
            self.conn.execute(
                "INSERT INTO briefs (created_ts, subject, stats_json, narrative, analysis_json) "
                "VALUES (?, ?, ?, ?, ?)",
                (created_ts, subject, stats_json, narrative, analysis_json),
            )

    def get_briefs(self, days=7):
        cutoff = int(time.time()) - days * 86400
        with self.lock:
            cur = self.conn.execute(
                "SELECT id, created_ts, subject, stats_json, narrative "
                "FROM briefs WHERE created_ts >= ? ORDER BY created_ts DESC",
                (cutoff,),
            )
            rows = cur.fetchall()
        result = []
        for id_, created_ts, subject, stats_json, narrative in rows:
            try:
                stats = json.loads(stats_json)
            except Exception:
                stats = {}
            result.append({
                "id": id_,
                "created_ts": created_ts,
                "subject": subject,
                "stats": stats,
                "narrative": narrative,
            })
        return result

    # ── Power readings ───────────────────────────────────────────────────────

    def insert_power_reading(self, ts, watts, voltage, current_a, energy_kwh):
        with self.lock:
            self.conn.execute(
                "INSERT INTO power_readings (timestamp, watts, voltage, current_a, energy_kwh) "
                "VALUES (?, ?, ?, ?, ?)",
                (ts, watts, voltage, current_a, energy_kwh),
            )

    def get_power_readings(self, days=7):
        cutoff = int(time.time()) - days * 86400
        with self.lock:
            rows = self.conn.execute(
                "SELECT timestamp, watts, voltage, current_a, energy_kwh "
                "FROM power_readings WHERE timestamp >= ? ORDER BY timestamp ASC",
                (cutoff,),
            ).fetchall()
        return [
            {"timestamp": r[0], "watts": r[1], "voltage": r[2],
             "current_a": r[3], "energy_kwh": r[4]}
            for r in rows
        ]

    # ── Pruning ─────────────────────────────────────────────────────────────

    def prune(self):
        """Delete rows older than retention_days. Returns (pings_deleted, incidents_deleted)."""
        cutoff = int(time.time()) - self.retention_days * 86400
        with self.lock:
            r1 = self.conn.execute(
                "DELETE FROM pings WHERE timestamp < ?", (cutoff,)
            )
            pings_deleted = r1.rowcount
            r2 = self.conn.execute(
                "DELETE FROM incidents WHERE ended IS NOT NULL AND ended < ?", (cutoff,)
            )
            incidents_deleted = r2.rowcount
            r3 = self.conn.execute(
                "DELETE FROM briefs WHERE created_ts < ?",
                (int(time.time()) - 7 * 86400,),
            )
            briefs_deleted = r3.rowcount
            r4 = self.conn.execute(
                "DELETE FROM power_readings WHERE timestamp < ?", (cutoff,)
            )
            power_deleted = r4.rowcount
            # No VACUUM: with fixed retention the DB is steady-state and
            # freed pages get reused. Daily VACUUM rewrote the whole DB
            # through the WAL (~300MB/day of SD writes) for nothing.
            # Manual reclaim if ever needed: sqlite3 netwatch.db VACUUM.
            self.conn.execute("PRAGMA wal_checkpoint(TRUNCATE)")
        if pings_deleted or incidents_deleted or briefs_deleted or power_deleted:
            logging.info(
                f"HistoryDB: pruned {pings_deleted} pings, "
                f"{incidents_deleted} incidents, "
                f"{briefs_deleted} briefs, "
                f"{power_deleted} power readings"
            )
        return pings_deleted, incidents_deleted


def _flush_loop(history_db, stop_event):
    """Flush buffered pings every 30s; final flush on shutdown."""
    while not stop_event.wait(30):
        try:
            history_db.flush_pings()
        except Exception as e:
            logging.warning(f"HistoryDB flush failed: {e}")
    try:
        history_db.flush_pings()
    except Exception:
        pass


def _prune_loop(history_db, stop_event):
    """Run prune() once a day until stop_event is set."""
    SECONDS_PER_DAY = 86400
    # Run first prune ~60s after startup so the system isn't busy at boot
    elapsed = SECONDS_PER_DAY - 60
    while not stop_event.is_set():
        if elapsed >= SECONDS_PER_DAY:
            rollup_ok = True
            try:
                history_db.rollup_days()
            except Exception as e:
                logging.warning(f"HistoryDB rollup failed: {e}")
                rollup_ok = False
            if rollup_ok:
                try:
                    history_db.prune()
                except Exception as e:
                    logging.warning(f"HistoryDB prune failed: {e}")
            elapsed = 0
        time.sleep(5)
        elapsed += 5


# ============================================================================
# Inventory (CMDB)
# ============================================================================

# Inventory type taxonomy. Each type has a list of "type-specific" properties
# stored in the properties JSON blob; common fields (system name, mac, ip,
# serial, notes, etc.) live as top-level columns and are shared across types.
INVENTORY_TYPES = ("host", "vm", "network", "ups", "disk", "peripheral", "tablet", "phone", "printer")

INVENTORY_TYPE_PROPERTIES = {
    "host": [],  # all fields are top-level (cpu, ram, os, etc.)
    "vm": [
        # VMs use ALL the host fields (cpu, ram, os, etc. all top-level)
        # AND these VM-specific fields stored in the properties JSON.
        ("hypervisor",     "string", "Hypervisor (Proxmox/KVM/ESXi/etc.)"),
        ("vcpu_count",     "int",    "vCPU count"),
        ("ram_alloc_gb",   "int",    "Allocated RAM (GB)"),
        ("disk_alloc_gb",  "int",    "Allocated disk (GB)"),
        ("autostart",      "bool",   "Auto-starts with host"),
        ("proxmox_vmid",   "int",    "Proxmox VMID"),
    ],
    "network": [
        ("port_count",     "int",    "Port count"),
        ("poe_watts",      "int",    "PoE budget (W)"),
        ("managed",        "bool",   "Managed"),
        ("uplink_speed",   "string", "Uplink speed"),
    ],
    "ups": [
        ("capacity_va",       "int",    "Capacity (VA)"),
        ("capacity_wh",       "int",    "Capacity (Wh)"),
        ("runtime_min",       "int",    "Runtime (min) at full load"),
        ("battery_age_years", "string", "Battery age"),
    ],
    "disk": [
        ("capacity_gb",  "int",    "Capacity (GB)"),
        ("interface",    "string", "Interface (SATA/NVMe/USB)"),
        ("rpm",          "int",    "Spindle speed (RPM, blank for SSD)"),
        ("used_in",      "string", "Currently installed in"),
        ("health",       "string", "Health status"),
    ],
    "peripheral": [
        ("subtype",      "string", "Type (KVM, monitor, keyboard, etc.)"),
        ("model",        "string", "Model"),
    ],
}


class InventoryDB:
    """SQLite-backed inventory store. Lives in the same database as ping history
    so we have one file to back up / one connection lifecycle to manage."""

    SCHEMA = """
    CREATE TABLE IF NOT EXISTS inventory (
        id           INTEGER PRIMARY KEY AUTOINCREMENT,
        category     TEXT,
        system       TEXT NOT NULL,
        role         TEXT,
        cpu          TEXT,
        ram_gb       REAL,
        gpu          TEXT,
        architecture TEXT,
        os           TEXT,
        cpu_score    INTEGER,
        tdp_watts    INTEGER,
        tpm          TEXT,
        mac          TEXT,
        ip           TEXT,
        serial       TEXT,
        notes        TEXT,
        created_at   INTEGER NOT NULL,
        updated_at   INTEGER NOT NULL
    );
    CREATE UNIQUE INDEX IF NOT EXISTS idx_inv_mac ON inventory(mac) WHERE mac IS NOT NULL AND mac != '';
    CREATE INDEX IF NOT EXISTS idx_inv_category ON inventory(category);
    CREATE INDEX IF NOT EXISTS idx_inv_system ON inventory(system);
    """

    FIELDS = [
        "category", "system", "role", "cpu", "ram_gb", "gpu", "architecture",
        "os", "cpu_score", "tdp_watts", "tpm", "mac", "ip", "serial", "notes",
        "device_type", "properties",
    ]

    def __init__(self, history_db):
        """Reuses the connection from HistoryDB."""
        self.history_db = history_db
        self.lock = history_db.lock  # share the same lock to serialize writes
        self.conn = history_db.conn
        self.conn.executescript(self.SCHEMA)
        if not _column_exists(self.conn, "inventory", "device_type"):
            self.conn.execute("ALTER TABLE inventory ADD COLUMN device_type TEXT DEFAULT 'host' NOT NULL")
            logging.info("InventoryDB: added device_type column")
        if not _column_exists(self.conn, "inventory", "properties"):
            self.conn.execute("ALTER TABLE inventory ADD COLUMN properties TEXT")
            logging.info("InventoryDB: added properties column")
        # Connections table - records edges between inventory devices.
        # CREATE TABLE IF NOT EXISTS is idempotent so re-runs are safe.
        self.conn.executescript("""
            CREATE TABLE IF NOT EXISTS inventory_connections (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                from_device_id  INTEGER NOT NULL,
                to_device_id    INTEGER NOT NULL,
                from_port       TEXT,
                to_port         TEXT,
                connection_type TEXT DEFAULT 'ethernet',
                notes           TEXT,
                created_at      INTEGER NOT NULL,
                FOREIGN KEY (from_device_id) REFERENCES inventory(id) ON DELETE CASCADE,
                FOREIGN KEY (to_device_id)   REFERENCES inventory(id) ON DELETE CASCADE
            );
            CREATE INDEX IF NOT EXISTS idx_conn_from ON inventory_connections(from_device_id);
            CREATE INDEX IF NOT EXISTS idx_conn_to   ON inventory_connections(to_device_id);
        """)
        # SQLite needs PRAGMA foreign_keys=ON for CASCADE to actually work.
        # The HistoryDB connection might not have it on; flip it now.
        self.conn.execute("PRAGMA foreign_keys = ON")
        logging.info("InventoryDB: schema ready")

    @staticmethod
    def normalize_mac(mac):
        """Lowercase + colon-separated. Returns '' for falsy input."""
        if not mac:
            return ""
        s = str(mac).strip().lower()
        # Strip out anything that isn't hex
        clean = "".join(c for c in s if c in "0123456789abcdef")
        if len(clean) != 12:
            # Don't reformat if it's not a valid 12-hex-char MAC
            return s
        return ":".join(clean[i:i+2] for i in range(0, 12, 2))

    def get_device_type_map(self):
        """Return {ip: device_type} for all inventory records with a non-empty IP.
        When multiple records share an IP, the one with the highest id wins."""
        with self.lock:
            cur = self.conn.execute(
                "SELECT ip, device_type FROM inventory"
                " WHERE ip IS NOT NULL AND ip != ''"
                " ORDER BY id ASC"
            )
            return {row[0]: (row[1] or "host") for row in cur.fetchall()}

    def list_all(self):
        with self.lock:
            cur = self.conn.execute(
                "SELECT id, category, system, role, cpu, ram_gb, gpu, architecture, "
                "os, cpu_score, tdp_watts, tpm, mac, ip, serial, notes, "
                "device_type, properties, "
                "created_at, updated_at FROM inventory ORDER BY system COLLATE NOCASE"
            )
            cols = [d[0] for d in cur.description]
            rows = [dict(zip(cols, row)) for row in cur.fetchall()]
            for r in rows:
                self._decode_properties(r)
            return rows

    def get(self, inv_id):
        with self.lock:
            cur = self.conn.execute(
                "SELECT id, category, system, role, cpu, ram_gb, gpu, architecture, "
                "os, cpu_score, tdp_watts, tpm, mac, ip, serial, notes, "
                "device_type, properties, "
                "created_at, updated_at FROM inventory WHERE id = ?", (inv_id,)
            )
            row = cur.fetchone()
            if not row:
                return None
            cols = [d[0] for d in cur.description]
            rec = dict(zip(cols, row))
            self._decode_properties(rec)
            return rec

    def find_by_mac(self, mac):
        """Return inventory record matching a MAC (normalized), or None."""
        norm = self.normalize_mac(mac)
        if not norm:
            return None
        with self.lock:
            cur = self.conn.execute(
                "SELECT id, category, system, role, cpu, ram_gb, gpu, architecture, "
                "os, cpu_score, tdp_watts, tpm, mac, ip, serial, notes, "
                "device_type, properties, "
                "created_at, updated_at FROM inventory WHERE mac = ?", (norm,)
            )
            row = cur.fetchone()
            if not row:
                return None
            cols = [d[0] for d in cur.description]
            rec = dict(zip(cols, row))
            self._decode_properties(rec)
            return rec

    def _decode_properties(self, rec):
        """Decode the properties JSON blob into a dict in-place. If the
        blob is missing/malformed, set properties to {}."""
        raw = rec.get("properties")
        if raw is None or raw == "":
            rec["properties"] = {}
            return
        if isinstance(raw, dict):
            return  # already decoded
        try:
            import json
            rec["properties"] = json.loads(raw)
            if not isinstance(rec["properties"], dict):
                rec["properties"] = {}
        except (ValueError, TypeError):
            rec["properties"] = {}

    def create(self, data):
        """Insert a new record. data is a dict of field values."""
        clean = self._clean_input(data)
        if not clean.get("system"):
            return None, "system name is required"
        # Check for MAC conflict
        if clean.get("mac"):
            existing = self.find_by_mac(clean["mac"])
            if existing:
                return None, f"a record with MAC {clean['mac']} already exists ({existing['system']})"
        ts = int(time.time())
        with self.lock:
            try:
                cur = self.conn.execute(
                    "INSERT INTO inventory (category, system, role, cpu, ram_gb, gpu, "
                    "architecture, os, cpu_score, tdp_watts, tpm, mac, ip, serial, notes, "
                    "device_type, properties, "
                    "created_at, updated_at) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (clean.get("category"), clean["system"], clean.get("role"),
                     clean.get("cpu"), clean.get("ram_gb"), clean.get("gpu"),
                     clean.get("architecture"), clean.get("os"), clean.get("cpu_score"),
                     clean.get("tdp_watts"), clean.get("tpm"), clean.get("mac"),
                     clean.get("ip"), clean.get("serial"), clean.get("notes"),
                     clean.get("device_type") or "host",
                     clean.get("properties"),
                     ts, ts)
                )
                return cur.lastrowid, None
            except sqlite3.IntegrityError:
                return None, f"a record with MAC {clean.get('mac')} already exists"

    def update(self, inv_id, data):
        clean = self._clean_input(data)
        if "system" in clean and not clean["system"]:
            return False, "system name cannot be empty"
        # MAC conflict check (if MAC is being set, ensure no other record has it)
        if clean.get("mac"):
            existing = self.find_by_mac(clean["mac"])
            if existing and existing["id"] != inv_id:
                return False, f"a different record with MAC {clean['mac']} already exists"
        ts = int(time.time())
        sets = []
        vals = []
        for f in self.FIELDS:
            if f in clean:
                sets.append(f"{f} = ?")
                vals.append(clean[f])
        if not sets:
            return True, None  # nothing to update
        sets.append("updated_at = ?")
        vals.append(ts)
        vals.append(inv_id)
        with self.lock:
            cur = self.conn.execute(
                f"UPDATE inventory SET {', '.join(sets)} WHERE id = ?", vals
            )
            if cur.rowcount == 0:
                return False, "record not found"
        return True, None

    # ─── Connections (inventory_connections table) ──────────────────────────
    # Connection types we accept. Anything else gets coerced to "ethernet".
    CONNECTION_TYPES = ("ethernet", "fiber", "wifi", "virtual", "power", "usb", "console", "other")

    def _normalize_conn_type(self, t):
        if t is None: return "ethernet"
        s = str(t).strip().lower()
        return s if s in self.CONNECTION_TYPES else "ethernet"

    def list_connections_for_device(self, device_id):
        """Return all connections involving this device, both directions.

        Each result includes the OTHER device's id and name (joined server-side
        so the UI doesn't need a second call), plus a "direction" field of
        either "out" (this device is the from end) or "in" (this device is
        the to end).
        """
        with self.lock:
            cur = self.conn.execute(
                "SELECT c.id, c.from_device_id, c.to_device_id, "
                "c.from_port, c.to_port, c.connection_type, c.notes, c.created_at, "
                "f.system AS from_name, f.device_type AS from_type, "
                "t.system AS to_name,   t.device_type AS to_type "
                "FROM inventory_connections c "
                "JOIN inventory f ON f.id = c.from_device_id "
                "JOIN inventory t ON t.id = c.to_device_id "
                "WHERE c.from_device_id = ? OR c.to_device_id = ? "
                "ORDER BY c.connection_type, c.created_at",
                (device_id, device_id),
            )
            cols = [d[0] for d in cur.description]
            rows = [dict(zip(cols, row)) for row in cur.fetchall()]
        for r in rows:
            r["direction"] = "out" if r["from_device_id"] == device_id else "in"
        return rows

    def list_all_connections(self):
        """Return all connections in the system. Used by the topology view."""
        with self.lock:
            cur = self.conn.execute(
                "SELECT id, from_device_id, to_device_id, from_port, to_port, "
                "connection_type, notes, created_at FROM inventory_connections "
                "ORDER BY connection_type, created_at"
            )
            cols = [d[0] for d in cur.description]
            return [dict(zip(cols, row)) for row in cur.fetchall()]

    def create_connection(self, data):
        """Create a new connection row. Returns (id, error_msg)."""
        try:
            from_id = int(data.get("from_device_id"))
            to_id   = int(data.get("to_device_id"))
        except (TypeError, ValueError):
            return None, "from_device_id and to_device_id required"
        if from_id == to_id:
            return None, "cannot connect a device to itself"
        # Verify both ends exist
        with self.lock:
            cur = self.conn.execute(
                "SELECT id FROM inventory WHERE id IN (?, ?)", (from_id, to_id)
            )
            ids = {r[0] for r in cur.fetchall()}
            if from_id not in ids or to_id not in ids:
                return None, "one or both devices do not exist"
        ctype = self._normalize_conn_type(data.get("connection_type"))
        from_port = (data.get("from_port") or "").strip() or None
        to_port   = (data.get("to_port") or "").strip() or None
        notes     = (data.get("notes") or "").strip() or None
        ts = int(time.time())
        with self.lock:
            cur = self.conn.execute(
                "INSERT INTO inventory_connections "
                "(from_device_id, to_device_id, from_port, to_port, "
                "connection_type, notes, created_at) "
                "VALUES (?, ?, ?, ?, ?, ?, ?)",
                (from_id, to_id, from_port, to_port, ctype, notes, ts),
            )
            return cur.lastrowid, None

    def update_connection(self, conn_id, data):
        """Update fields of an existing connection. Returns (ok, error_msg)."""
        fields, values = [], []
        if "from_port" in data:
            fields.append("from_port = ?")
            values.append((data.get("from_port") or "").strip() or None)
        if "to_port" in data:
            fields.append("to_port = ?")
            values.append((data.get("to_port") or "").strip() or None)
        if "connection_type" in data:
            fields.append("connection_type = ?")
            values.append(self._normalize_conn_type(data.get("connection_type")))
        if "notes" in data:
            fields.append("notes = ?")
            values.append((data.get("notes") or "").strip() or None)
        if not fields:
            return False, "no fields to update"
        values.append(conn_id)
        with self.lock:
            cur = self.conn.execute(
                "UPDATE inventory_connections SET " + ", ".join(fields)
                + " WHERE id = ?", tuple(values)
            )
            if cur.rowcount == 0:
                return False, "connection not found"
        return True, None

    def delete_connection(self, conn_id):
        with self.lock:
            cur = self.conn.execute(
                "DELETE FROM inventory_connections WHERE id = ?", (conn_id,)
            )
            if cur.rowcount == 0:
                return False, "connection not found"
        return True, None

    def delete(self, inv_id):
        with self.lock:
            cur = self.conn.execute("DELETE FROM inventory WHERE id = ?", (inv_id,))
            if cur.rowcount == 0:
                return False, "record not found"
        return True, None

    def replace_all(self, records):
        """Wipe inventory and bulk-insert atomically. Used by 'Replace all' import mode.

        DELETE and all INSERTs run inside a single BEGIN/COMMIT while holding
        self.lock so concurrent readers never see a partially-empty table.
        """
        ok, fail = 0, []
        ts = int(time.time())
        # Pre-validate and clean records before acquiring the lock so the
        # locked section is as short as possible.
        cleaned = []
        for rec in records:
            clean = self._clean_input(rec)
            if not clean.get("system"):
                fail.append({"system": rec.get("system", "?"), "error": "system name is required"})
            else:
                cleaned.append((rec, clean))
        with self.lock:
            self.conn.execute("BEGIN")
            try:
                self.conn.execute("DELETE FROM inventory")
                for rec, clean in cleaned:
                    if clean.get("mac"):
                        row = self.conn.execute(
                            "SELECT system FROM inventory WHERE mac = ?", (clean["mac"],)
                        ).fetchone()
                        if row:
                            fail.append({"system": clean["system"],
                                         "error": f"MAC {clean['mac']} already used by '{row[0]}'"})
                            continue
                    self.conn.execute(
                        "INSERT INTO inventory (category, system, role, cpu, ram_gb, gpu, "
                        "architecture, os, cpu_score, tdp_watts, tpm, mac, ip, serial, notes, "
                        "device_type, properties, created_at, updated_at) "
                        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                        (clean.get("category"), clean["system"], clean.get("role"),
                         clean.get("cpu"), clean.get("ram_gb"), clean.get("gpu"),
                         clean.get("architecture"), clean.get("os"), clean.get("cpu_score"),
                         clean.get("tdp_watts"), clean.get("tpm"), clean.get("mac"),
                         clean.get("ip"), clean.get("serial"), clean.get("notes"),
                         clean.get("device_type") or "host", clean.get("properties"),
                         ts, ts),
                    )
                    ok += 1
                self.conn.execute("COMMIT")
            except Exception:
                try: self.conn.execute("ROLLBACK")
                except Exception: pass
                raise
        return ok, fail

    def _clean_input(self, data):
        """Coerce / validate field values."""
        import json as _json
        out = {}
        for f in self.FIELDS:
            if f not in data:
                continue
            v = data[f]
            if v is None or (isinstance(v, str) and v.strip() == ""):
                out[f] = None
                continue
            if f in ("ram_gb",):
                try: out[f] = float(v)
                except (ValueError, TypeError): out[f] = None
            elif f in ("cpu_score", "tdp_watts"):
                try: out[f] = int(float(v))
                except (ValueError, TypeError): out[f] = None
            elif f == "mac":
                out[f] = self.normalize_mac(v)
                if not out[f]:
                    out[f] = None
            elif f == "device_type":
                t = str(v).strip().lower()
                out[f] = t if t in INVENTORY_TYPES else "peripheral"
            elif f == "properties":
                # Accept a dict (preferred), serialize to JSON for storage.
                # Strings are passed through if they parse as JSON dicts.
                if isinstance(v, dict):
                    out[f] = _json.dumps(v)
                elif isinstance(v, str):
                    try:
                        parsed = _json.loads(v)
                        out[f] = _json.dumps(parsed) if isinstance(parsed, dict) else None
                    except (ValueError, TypeError):
                        out[f] = None
                else:
                    out[f] = None
            else:
                out[f] = str(v).strip() if v is not None else None
        return out


def export_inventory_to_xlsx(inventory_db, scope='hosts'):
    """Build an XLSX file in memory containing inventory records.

    scope='hosts' (default): exports only host-type records on a single sheet
      named "Inventory". Filename: netwatch-inventory-hosts-{hostname}-{date}.xlsx
    scope='all': exports all device types, one sheet per type that has records.
      Sheet order follows INV_TYPE_ORDER. Filename: netwatch-inventory-all-…xlsx

    Column layout matches the import format for round-tripping host records.
    Returns (bytes, filename) on success, or (None, error_msg) on failure.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        return None, "openpyxl not available"

    import io
    import socket
    from datetime import datetime as _dt

    COLUMNS = [
        ("Category",             "category"),
        ("System",               "system"),
        ("Role / Status",        "role"),
        ("CPU",                  "cpu"),
        ("RAM_GB",               "ram_gb"),
        ("GPU",                  "gpu"),
        ("Architecture",         "architecture"),
        ("OS",                   "os"),
        ("Estimated_CPU_Score",  "cpu_score"),
        ("Max_TDP_Watts",        "tdp_watts"),
        ("TPM_Version",          "tpm"),
        ("MAC_Primary",          "mac"),
        ("IP_Address",           "ip"),
        ("Service_Tag_Serial",   "serial"),
        ("Notes",                "notes"),
    ]

    SHEET_NAMES = {
        'host': 'Hosts', 'vm': 'VMs', 'network': 'Network',
        'ups': 'UPS', 'disk': 'Disks', 'peripheral': 'Peripherals',
        'tablet': 'Tablets', 'phone': 'Phones', 'printer': 'Printers',
    }
    TYPE_ORDER = ['host', 'vm', 'network', 'ups', 'disk', 'peripheral', 'tablet', 'phone', 'printer']

    def _write_sheet(ws, records):
        for col_idx, (header, _) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
        for row_idx, rec in enumerate(records, start=2):
            for col_idx, (_, field) in enumerate(COLUMNS, start=1):
                ws.cell(row=row_idx, column=col_idx, value=rec.get(field))
        for col_idx, (header, field) in enumerate(COLUMNS, start=1):
            max_len = len(header)
            for rec in records:
                val = rec.get(field)
                if val is not None and len(str(val)) > max_len:
                    max_len = len(str(val))
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
        ws.freeze_panes = "A2"

    try:
        hostname = socket.gethostname() or "unknown"
        date_str = _dt.now().strftime("%Y-%m-%d")

        all_records = inventory_db.list_all()

        if scope == 'all':
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # remove default blank sheet

            # Group records by device_type
            by_type = {}
            for r in all_records:
                dt = r.get('device_type') or 'host'
                by_type.setdefault(dt, []).append(r)

            # Write sheets in TYPE_ORDER, then any unrecognised types
            ordered = [t for t in TYPE_ORDER if t in by_type]
            extras  = [t for t in by_type if t not in TYPE_ORDER]
            for dt in ordered + extras:
                sheet_name = SHEET_NAMES.get(dt, dt.title())
                ws = wb.create_sheet(title=sheet_name)
                _write_sheet(ws, by_type[dt])

            filename = f"netwatch-inventory-all-{hostname}-{date_str}.xlsx"
        else:
            # scope == 'hosts' (default)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Inventory"
            records = [r for r in all_records
                       if (r.get("device_type") or "host") == "host"]
            _write_sheet(ws, records)
            filename = f"netwatch-inventory-hosts-{hostname}-{date_str}.xlsx"

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue(), filename

    except Exception as e:
        return None, f"export failed: {e}"


def import_inventory_from_xlsx(inventory_db, xlsx_bytes, mode="add"):
    """Parse an uploaded xlsx and insert records.

    Expected columns (by header name, case-insensitive, flexible):
      Category, System, Role / Status, CPU, RAM_GB, GPU, Architecture, OS,
      Estimated_CPU_Score, Max_TDP_Watts, TPM_Version, MAC_Primary,
      IP_Address, Service_Tag_Serial

    mode='add'      -> insert new records, skip ones whose MAC or
                       (system+ram+cpu) tuple matches existing
    mode='replace'  -> wipe inventory first, then insert all
    Returns (added_count, skipped_count, errors_list).
    """
    try:
        import openpyxl
    except ImportError:
        return 0, 0, [{"row": 0, "error": "openpyxl not installed"}]

    import io
    try:
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    except Exception as e:
        return 0, 0, [{"row": 0, "error": f"could not parse xlsx: {e}"}]

    ws = wb.active
    # Read first row as headers
    rows = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        return 0, 0, [{"row": 0, "error": "spreadsheet is empty"}]

    # Map header names to our field names (case-insensitive, flexible)
    HEADER_MAP = {
        "category": "category",
        "system": "system",
        "role / status": "role",
        "role/status": "role",
        "role": "role",
        "cpu": "cpu",
        "ram_gb": "ram_gb",
        "ram (gb)": "ram_gb",
        "ram": "ram_gb",
        "gpu": "gpu",
        "architecture": "architecture",
        "arch": "architecture",
        "os": "os",
        "estimated_cpu_score": "cpu_score",
        "cpu_score": "cpu_score",
        "cpu score": "cpu_score",
        "max_tdp_watts": "tdp_watts",
        "tdp": "tdp_watts",
        "tdp_watts": "tdp_watts",
        "tpm_version": "tpm",
        "tpm": "tpm",
        "mac_primary": "mac",
        "mac": "mac",
        "mac address": "mac",
        "ip_address": "ip",
        "ip": "ip",
        "service_tag_serial": "serial",
        "serial": "serial",
        "service tag": "serial",
        "notes": "notes",
    }

    col_to_field = {}
    for idx, header in enumerate(header_row):
        if header is None: continue
        key = str(header).strip().lower()
        field = HEADER_MAP.get(key)
        if field:
            col_to_field[idx] = field

    if "system" not in col_to_field.values():
        return 0, 0, [{"row": 0, "error": "no 'System' column found in spreadsheet"}]

    # Build list of records
    records = []
    for row_idx, row in enumerate(rows, start=2):  # start=2 because row 1 was header
        rec = {}
        for col_idx, value in enumerate(row):
            field = col_to_field.get(col_idx)
            if field and value is not None:
                rec[field] = value
        if not rec.get("system"):
            continue  # skip blank rows
        records.append(rec)

    if mode == "replace":
        ok, fail = inventory_db.replace_all(records)
        return ok, 0, fail

    # 'add' mode: skip duplicates by MAC, otherwise insert
    added = 0
    skipped = 0
    errors = []
    existing = inventory_db.list_all()
    existing_macs = {InventoryDB.normalize_mac(r.get("mac")) for r in existing if r.get("mac")}
    existing_systems = {(r.get("system") or "").lower() for r in existing}
    for rec in records:
        mac_norm = InventoryDB.normalize_mac(rec.get("mac"))
        sys_lower = (rec.get("system") or "").lower()
        if mac_norm and mac_norm in existing_macs:
            skipped += 1
            continue
        if not mac_norm and sys_lower in existing_systems:
            skipped += 1
            continue
        inv_id, err = inventory_db.create(rec)
        if err:
            errors.append({"system": rec.get("system"), "error": err})
        else:
            added += 1
            if mac_norm: existing_macs.add(mac_norm)
            existing_systems.add(sys_lower)
    return added, skipped, errors


# ============================================================================
# Backup
# ============================================================================

BACKUP_MANIFEST_VERSION = 1


def create_backup_tarball(config_path, auth_path):
    """Build a complete backup tarball in memory and return (bytes, filename).

    The SQLite snapshot uses sqlite3's online .backup API for consistency.
    Other files are read normally.

    `config_path` is the path to hosts.yaml; the netwatch directory and
    db path are derived from it. `auth_path` is the auth.json location
    (may not exist yet if no users are configured).
    """
    import io
    import json as _json
    import sqlite3
    import socket
    import tarfile
    import tempfile
    from datetime import datetime as _dt

    netwatch_dir = os.path.dirname(os.path.abspath(config_path))
    db_path      = os.path.join(netwatch_dir, "netwatch.db")
    monitor_path = os.path.join(netwatch_dir, "monitor.py")
    hostname     = socket.gethostname() or "unknown"
    iso_now      = _dt.now().strftime("%Y-%m-%dT%H-%M-%S")
    filename     = f"netwatch-backup-{hostname}-{iso_now}.tar.gz"

    manifest = {
        "manifest_version": BACKUP_MANIFEST_VERSION,
        "netwatch_version": VERSION,
        "created_at":       int(time.time()),
        "created_iso":      _dt.now().isoformat(),
        "source_hostname":  hostname,
        "files":            {},
    }

    # 1) Make a consistent SQLite snapshot to a temp file. We can't
    # tar.add() the live db directly because WAL writes might be active.
    snapshot_path = None
    if os.path.isfile(db_path):
        fd, snapshot_path = tempfile.mkstemp(prefix="nw_backup_", suffix=".db")
        os.close(fd)
        try:
            src = sqlite3.connect(db_path)
            try:
                dst = sqlite3.connect(snapshot_path)
                try:
                    src.backup(dst)
                finally:
                    dst.close()
            finally:
                src.close()
        except Exception:
            # If snapshot fails for any reason, clean up and re-raise
            try: os.unlink(snapshot_path)
            except OSError: pass
            raise

    # 2) Build the tarball in memory.
    try:
        buf = io.BytesIO()
        with tarfile.open(fileobj=buf, mode="w:gz") as tar:
            def _add(real_path, arcname, mode_override=None):
                if not os.path.isfile(real_path):
                    return
                ti = tar.gettarinfo(real_path, arcname=arcname)
                if mode_override is not None:
                    ti.mode = mode_override
                # Strip uid/gid - they're meaningless across machines
                ti.uid = 0; ti.gid = 0
                ti.uname = ""; ti.gname = ""
                with open(real_path, "rb") as f:
                    tar.addfile(ti, f)
                manifest["files"][os.path.basename(arcname)] = os.path.getsize(real_path)

            _add(monitor_path, "netwatch/monitor.py")
            _add(config_path,  "netwatch/hosts.yaml")
            _add(auth_path,    "netwatch/auth.json", mode_override=0o600)
            if snapshot_path:
                # Snapshot lands under the original db filename
                ti = tar.gettarinfo(snapshot_path, arcname="netwatch/netwatch.db")
                ti.uid = 0; ti.gid = 0
                ti.uname = ""; ti.gname = ""
                with open(snapshot_path, "rb") as f:
                    tar.addfile(ti, f)
                manifest["files"]["netwatch.db"] = os.path.getsize(snapshot_path)

            # Manifest goes in last so all file sizes are populated
            manifest_bytes = _json.dumps(manifest, indent=2).encode("utf-8")
            ti = tarfile.TarInfo(name="netwatch/metadata.json")
            ti.size = len(manifest_bytes)
            ti.mtime = int(time.time())
            ti.mode = 0o644
            tar.addfile(ti, io.BytesIO(manifest_bytes))

        return buf.getvalue(), filename, manifest
    finally:
        # Always clean up the snapshot file, even if tar.add fails
        if snapshot_path:
            try: os.unlink(snapshot_path)
            except OSError: pass


def restore_backup(tarball_path, config_path, force=False):
    """Restore hosts.yaml, auth.json, and netwatch.db from a backup tarball
    built by create_backup_tarball().

    Deliberately does NOT extract the tarball's bundled monitor.py: the
    tarball is meant to be a fully self-contained emergency artifact usable
    on its own, but when --restore runs after a fresh git clone (the normal
    redeploy path), overwriting freshly-cloned code with whatever version
    made the backup would silently downgrade it.

    Returns (ok, message). Never raises for expected failure conditions
    (missing/invalid tarball, conflicting destination files) - callers can
    print the message and exit without a traceback.
    """
    import tarfile

    if not os.path.isfile(tarball_path):
        return False, f"Backup file not found: {tarball_path}"

    try:
        tar = tarfile.open(tarball_path, "r:gz")
    except (tarfile.TarError, OSError) as e:
        return False, f"Could not open backup tarball: {e}"

    with tar:
        try:
            manifest_member = tar.getmember("netwatch/metadata.json")
        except KeyError:
            return False, "Not a valid netwatch backup (missing netwatch/metadata.json)"

        manifest = json.loads(tar.extractfile(manifest_member).read().decode("utf-8"))

        warning = ""
        backup_version = manifest.get("manifest_version", 0)
        if backup_version > BACKUP_MANIFEST_VERSION:
            warning = (
                f"Warning: this backup was made by a newer netwatch version "
                f"(manifest v{backup_version}, this is v{BACKUP_MANIFEST_VERSION}) "
                f"- restore may be incomplete.\n"
            )

        config_dir = os.path.dirname(os.path.abspath(config_path))
        targets = {
            "netwatch/hosts.yaml":  os.path.join(config_dir, "hosts.yaml"),
            "netwatch/auth.json":   os.path.join(config_dir, "auth.json"),
            "netwatch/netwatch.db": os.path.join(config_dir, "netwatch.db"),
        }

        if not force:
            existing = [dest for dest in targets.values() if os.path.exists(dest)]
            if existing:
                listing = "\n".join(f"  - {p}" for p in existing)
                return False, (
                    "Refusing to overwrite existing files (use --force to overwrite):\n"
                    f"{listing}"
                )

        os.makedirs(config_dir, exist_ok=True)
        restored = []
        for arcname, dest in targets.items():
            try:
                member = tar.getmember(arcname)
            except KeyError:
                continue  # e.g. auth.json may be absent if no admin was ever set up
            with tar.extractfile(member) as src, open(dest, "wb") as out:
                out.write(src.read())
            if arcname == "netwatch/auth.json":
                os.chmod(dest, 0o600)
            restored.append(dest)

    files_listing = "\n".join(f"  - {p}" for p in restored)
    message = (
        f"{warning}"
        f"Restored backup from {manifest.get('source_hostname', 'unknown')}, "
        f"created {manifest.get('created_iso', 'unknown')}, "
        f"netwatch v{manifest.get('netwatch_version', 'unknown')}\n"
        f"Files written:\n{files_listing}"
    )
    return True, message

